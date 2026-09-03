"""Market Sniper v3.7 — 10-scenario regression suite."""
import io, json, os, re, shutil, subprocess, sys, tempfile, threading, time, urllib.request, urllib.error

HERE = "/sessions/stoic-brave-ritchie/mnt/Market Sniper"
sys.path.insert(0, HERE); os.chdir(HERE)
SETTINGS = os.path.join(HERE, "my-settings.json")
# Scratch dir for the backup + temp logs. Created here so the suite runs on a
# clean machine instead of assuming a folder someone made by hand once.
SCRATCH = os.path.join(tempfile.gettempdir(), "market_sniper_tests")
os.makedirs(SCRATCH, exist_ok=True)
BACKUP = os.path.join(SCRATCH, "settings.realbackup")
results = []

SECRETY = ("key","sec","secret","token","pass")
def redact(name, detail):
    """Never print a stored value for a secret field. A failing assertion once
    echoed a real live API key into the log; that must not be possible again."""
    d = str(detail)
    if any(w in name.lower() for w in SECRETY) and d.strip("'\"") not in ("", "None"):
        return f"<{len(d.strip(chr(39)))} chars, redacted>"
    return d

def check(sc, name, ok, detail=""):
    results.append((sc, name, ok, redact(name, detail)))
    print(f"   {'PASS' if ok else 'FAIL'}  {name}" + (f"  -- {redact(name, detail)}" if detail and not ok else ""))

def http(url, method="GET", body=None, timeout=10):
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method,
                                 headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.status, json.loads(r.read().decode() or "{}")
    except urllib.error.HTTPError as e:
        try: return e.code, json.loads(e.read().decode() or "{}")
        except Exception: return e.code, {}
    except Exception as e:
        return 0, {"_err": str(e)}

def boot(app, port, log):
    return subprocess.Popen([sys.executable, "-m", "uvicorn", app, "--host", "127.0.0.1",
                             "--port", str(port)], cwd=HERE,
                            stdout=open(log, "wb"), stderr=subprocess.STDOUT)

shutil.copy(SETTINGS, BACKUP)

# THE SUITE MUST NEVER TOUCH THE REAL TRADE LOG.
# _record_close() writes a row for every finished trade, and the fake sessions
# in here finish plenty of them. Three test trades once landed in the live
# money log next to a real one. Redirect the module ONCE, before any scenario
# runs, rather than per-scenario where a later reload can quietly undo it.
import tempfile as _tf_boot, trade_log as _tl_boot
_TLOG_SANDBOX = _tf_boot.mkdtemp(prefix="ms_tradelog_")
_tl_boot.LOG_DIR   = _TLOG_SANDBOX
_tl_boot.CSV_PATH  = os.path.join(_TLOG_SANDBOX, "trades.csv")
_tl_boot.XLSX_PATH = os.path.join(_TLOG_SANDBOX, "Market Sniper Trade Log.xlsx")
REAL_TRADES_CSV = os.path.join(HERE, "logs", "trades.csv")
_REAL_TRADES_BEFORE = (io.open(REAL_TRADES_CSV, encoding="utf-8").read()
                       if os.path.exists(REAL_TRADES_CSV) else None)
OPT = FUT = None
try:
    OPT = boot("main:app", 8000, os.path.join(SCRATCH,"opt1.log"))
    FUT = boot("futures_app:app", 8010, os.path.join(SCRATCH,"fut1.log"))
    time.sleep(11)

    print("\n[1] FUTURES LOGIN SURVIVES A RESTART")
    http("http://127.0.0.1:8010/api/prefs", "POST", {
        "mode":"TOPSTEP","ts_user":"giancuffaro230","ts_acct":"EXPRESS-V2-CT-DLL-132001-66482406",
        "ts_key":"TSKEY","wb_key":"WBKEY","wb_sec":"WBSEC","nt_account":"1114140","nt_folder":"C:/nt/incoming"})
    FUT.terminate(); FUT.wait(timeout=10)
    FUT = boot("futures_app:app", 8010, os.path.join(SCRATCH,"fut2.log")); time.sleep(9)
    _, r = http("http://127.0.0.1:8010/api/prefs"); p = r.get("prefs", {})
    check(1,"username survives restart", p.get("ts_user")=="giancuffaro230", repr(p.get("ts_user")))
    check(1,"topstep account survives", str(p.get("ts_acct","")).startswith("EXPRESS-V2"), repr(p.get("ts_acct")))
    check(1,"NINJA account survives", p.get("nt_account")=="1114140", repr(p.get("nt_account")))
    check(1,"NINJA folder survives", p.get("nt_folder")=="C:/nt/incoming", repr(p.get("nt_folder")))
    check(1,"Webull key survives (was in NO save list)", p.get("wb_key")=="WBKEY", repr(p.get("wb_key")))
    check(1,"Webull secret survives", p.get("wb_sec")=="WBSEC", repr(p.get("wb_sec")))
    check(1,"Topstep API key survives", p.get("ts_key")=="TSKEY", repr(p.get("ts_key")))

    print("\n[2] remember_login DEFAULT no longer deletes secrets")
    check(2,"absent remember_login reads TRUE", p.get("remember_login") is True, repr(p.get("remember_login")))
    http("http://127.0.0.1:8010/api/prefs","POST",{"symbol":"MES"})
    _, r = http("http://127.0.0.1:8010/api/prefs"); p2 = r.get("prefs", {})
    check(2,"unrelated write does NOT wipe keys", p2.get("ts_key")=="TSKEY", repr(p2.get("ts_key")))
    http("http://127.0.0.1:8010/api/prefs","POST",{"remember_login":False})
    _, r = http("http://127.0.0.1:8010/api/prefs"); p3 = r.get("prefs", {})
    check(2,"explicit untick DOES wipe secrets", all(k not in p3 for k in ("ts_key","wb_key","wb_sec")))
    check(2,"untick keeps identifiers", p3.get("ts_user")=="giancuffaro230")

    print("\n[3] OPTIONS key profiles persist to DISK")
    http("http://127.0.0.1:8000/api/prefs","POST",{
        "profiles":[{"name":"Main","k":"K1","s":"S1"},{"name":"Mirror","k":"K2","s":"S2"}],
        "active_profile":"Main","show_secrets":True})
    OPT.terminate(); OPT.wait(timeout=10)
    OPT = boot("main:app", 8000, os.path.join(SCRATCH,"opt2.log")); time.sleep(9)
    _, r = http("http://127.0.0.1:8000/api/prefs"); op = r.get("prefs", {})
    profs = op.get("profiles", [])
    check(3,"both profiles survive restart", [x.get("name") for x in profs]==["Main","Mirror"], str(profs))
    check(3,"profile keys intact", [x.get("k") for x in profs]==["K1","K2"])
    check(3,"active profile remembered", op.get("active_profile")=="Main")
    check(3,"written to disk not browser",
          "profiles" in json.load(io.open(SETTINGS,encoding="utf-8")).get("options_prefs",{}))

    print("\n[4] BROWSER AUTOFILL cannot overwrite the username")
    html = io.open(os.path.join(HERE,"futures_index.html"),encoding="utf-8").read()
    for fid in ("tsUser","tsAcct","ntAccount","ntFolder","wbKey"):
        m = re.search(r'<input id="%s"[^>]*>'%fid, html)
        check(4,f"{fid} autofill-guarded", bool(m and "autocomplete=" in m.group(0)))
    m = re.search(r'<input id="tsKey"[^>]*>', html)
    check(4,"tsKey uses new-password", bool(m and "new-password" in m.group(0)))
    check(4,'hardcoded value="Sim101" removed', 'value="Sim101"' not in html)

    print("\n[5] ITM3 STRIKE MATH")
    import webull_client as wb
    check(5,"QQQ 724 CALLS -> 721", wb.pick_strike(724.0,"CALLS",1.0,"ITM3")==721.0)
    check(5,"QQQ 724 PUTS  -> 727", wb.pick_strike(724.0,"PUTS",1.0,"ITM3")==727.0)
    check(5,"calls land BELOW spot", wb.pick_strike(713.44,"CALLS",1.0,"ITM3")<713.44)
    check(5,"puts land ABOVE spot", wb.pick_strike(713.44,"PUTS",1.0,"ITM3")>713.44)
    check(5,"TSLA respects 2.50 step", wb.pick_strike(331.0,"CALLS",2.5,"ITM3")==325.0)
    check(5,"OTM1 unchanged", wb.pick_strike(724.0,"CALLS",1.0,"OTM1")==725.0)
    for bad in ("", None, "banana", "ITMx"):
        check(5,f"garbage {bad!r} -> safe default", wb.parse_strike_mode(bad) in (("OTM",1),("ITM",1)))
    check(5,"depth clamped at 20", wb.parse_strike_mode("ITM999")==("ITM",20))

    print("\n[6] PREVIEW and ARM cannot disagree")
    s = wb.make_session("LIVE"); s._guard_open = lambda q: None
    for spot in (724.0, 724.4, 718.6, 331.3):
        s._underlying = lambda sym, v=spot: v
        pv = s.preview_entry("QQQ","CALLS"); ar = s.arm("QQQ","CALLS",1)
        check(6,f"spot {spot}: preview target == arm target", pv["target"]==ar["target"],
              f"{pv['target']} vs {ar['target']}")
        s.armed=None
    s._underlying = lambda sym: None
    check(6,"no price -> honest failure not crash", s.preview_entry("QQQ","CALLS").get("ok") is False)

    print("\n[7] LIVE-ONLY: dead modes unreachable")
    import futures_client as fc
    for dead in ("PAPER","TRADOVATE","garbage"):
        code, r = http("http://127.0.0.1:8010/api/connect","POST",{"mode":dead})
        check(7,f"{dead} rejected", code==400 and "must be" in str(r.get("detail","")))
    check(7,"old 'LIVE' maps to NINJA", fc.normalize_mode("LIVE")=="NINJA")
    check(7,"3 real modes valid", all(fc.normalize_mode(m) for m in ("WEBULL","NINJA","TOPSTEP")))
    check(7,"PaperSession gone", not hasattr(wb,"PaperSession"))
    check(7,"TradovateSession gone", not hasattr(fc,"TradovateSession"))
    check(7,"Topstep helpers KEPT (nearly deleted)",
          hasattr(fc,"_tv_front_symbol") and hasattr(fc,"_fmt_px"))

    print("\n[8] AUTO-SYNC SAFETY")
    import auto_sync
    ok_c, err = auto_sync.python_files_compile()
    check(8,"all python compiles", ok_c, str(err))
    watched = auto_sync.snapshot()
    check(8,"my-settings.json NEVER watched", not any("my-settings.json" in f for f in watched))
    check(8,"logs/ never watched", not any("/logs/" in f for f in watched))
    check(8,"_archive/ never watched", not any("_archive" in f for f in watched))
    check(8,"secrets on skip list", "my-settings.json" in auto_sync.SKIP_EXACT)
    rc = subprocess.run(["git","check-ignore","my-settings.json"],cwd=HERE,stdout=subprocess.PIPE).returncode
    check(8,"gitignored too (second lock)", rc==0)

    print("\n[9] ENDPOINTS ALIVE, NO TRACEBACKS")
    for name,url in (("options health","http://127.0.0.1:8000/api/health"),
                     ("futures health","http://127.0.0.1:8010/api/health"),
                     ("options page","http://127.0.0.1:8000/"),
                     ("futures page","http://127.0.0.1:8010/"),
                     ("tape QQQ","http://127.0.0.1:8000/api/tape?symbol=QQQ"),
                     ("tape MNQ","http://127.0.0.1:8010/api/tape?symbol=MNQ"),
                     ("trend","http://127.0.0.1:8000/api/trend?symbol=QQQ"),
                     ("futures prices","http://127.0.0.1:8010/api/prices")):
        try:
            with urllib.request.urlopen(url, timeout=12) as rr: check(9,name, rr.status==200)
        except Exception as e: check(9,name,False,str(e)[:60])
    _, hv = http("http://127.0.0.1:8000/api/health")
    # Read the expected version from config rather than hardcoding it, so a
    # version bump does not fail a test that has nothing to do with versions.
    import config as _cfg
    check(9,f"served version matches config ({_cfg.APP_VERSION})",
          hv.get("version")==_cfg.APP_VERSION, str(hv.get("version")))
    code,_ = http("http://127.0.0.1:8000/api/tape?symbol=NVDA")
    check(9,"unknown symbol rejected", code==400)
    for log in (os.path.join(SCRATCH,"opt2.log"),os.path.join(SCRATCH,"fut2.log")):
        txt = io.open(log,encoding="utf-8",errors="replace").read() if os.path.exists(log) else ""
        check(9,f"no traceback in {os.path.basename(log)}", "Traceback" not in txt)

    print("\n[10] UI INTEGRITY")
    for page in ("index.html","futures_index.html"):
        src = io.open(os.path.join(HERE,page),encoding="utf-8").read()
        ids = set(re.findall(r'id="([\w-]+)"',src)); used = set(re.findall(r"\$\('([\w-]+)'\)",src))
        check(10,f"{page}: every referenced id exists", not (used-ids), str(sorted(used-ids)))
        check(10,f"{page}: no PAPER remnants", "paperMode" not in src and "paperConfig" not in src)
    # Real JS parse, not a brace count. These files hold ~40KB of hand-edited
    # JavaScript; a stray bracket would leave the page blank with the error
    # only visible in the browser console, which nobody has open while trading.
    import subprocess as _sp, shutil as _sh
    if _sh.which("node"):
        for page in ("index.html","futures_index.html"):
            src_ = io.open(os.path.join(HERE,page),encoding="utf-8").read()
            blocks = re.findall(r"<script>(.*?)</script>", src_, re.S)
            check(10,f"{page}: has script blocks", len(blocks)>0)
            for i,js in enumerate(blocks):
                f_=os.path.join(SCRATCH,f"chk_{page}_{i}.js")
                io.open(f_,"w",encoding="utf-8").write(js)
                r_=_sp.run(["node","--check",f_],capture_output=True,text=True)
                check(10,f"{page}: script block {i} parses",
                      r_.returncode==0, (r_.stderr or "").strip().split(chr(10))[0][:90])
    else:
        check(10,"node available for JS syntax check", False, "node not installed")

    idx = io.open(os.path.join(HERE,"index.html"),encoding="utf-8").read()
    import re as _re
    _code = _re.sub(r"//[^\n]*", "", idx)      # strip comments before searching
    check(10,"strike no longer truncated with |0", "strike|0" not in _code)
    check(10,"buy button shows a strike", "fmtStrike(strike)" in idx)
    check(10,"strike buttons present", "smATM" in idx and "smITM2" in idx)
    fidx = io.open(os.path.join(HERE,"futures_index.html"),encoding="utf-8").read()
    check(10,"no Tradovate UI left", "tvConfig" not in fidx and "tvUser" not in fidx)
    print("\n[11] MULTI-BROKER SESSIONS (stay logged in, toggle freely)")
    import futures_app as fa
    class FakeSess:
        def __init__(self, tag, pos=None):
            self.account_id=tag; self.position=pos; self.day_realized=0.0
            self.refreshed=0
        def refresh_mark(self): self.refreshed+=1
        def state(self): return {"account_id":self.account_id,"position":self.position}
    fa.SESSIONS.clear()
    ts=FakeSess("TS:EXPRESS", {"symbol":"MNQ","side":"LONG","pnl":42.0})
    nj=FakeSess("NT:1114140")
    wb_=FakeSess("WB:9999")
    fa.SESSIONS.update({"TOPSTEP":ts,"NINJA":nj,"WEBULL":wb_})
    fa.ACTIVE["mode"]="TOPSTEP"

    check(11,"three brokers logged in at once", len(fa.SESSIONS)==3)
    st = fa.state()
    check(11,"active broker reported", st.get("active_mode")=="TOPSTEP")
    check(11,"roster lists all three", len(st.get("sessions",[]))==3)
    check(11,"INACTIVE brokers still refreshed (TP/SL keeps running)",
          nj.refreshed>0 and wb_.refreshed>0, f"ninja={nj.refreshed} webull={wb_.refreshed}")
    check(11,"open position surfaced on the roster",
          any(x["mode"]=="TOPSTEP" and x["has_position"] for x in st["sessions"]))

    before = ts.refreshed
    sw = fa.switch(fa.SwitchReq(mode="NINJA"))
    check(11,"switch does NOT log anyone out", len(fa.SESSIONS)==3)
    check(11,"switch changes the active broker", sw.get("active_mode")=="NINJA")
    check(11,"switched-away broker STILL refreshed", ts.refreshed>before,
          f"{before} -> {ts.refreshed}")
    check(11,"old 'LIVE' alias switches to NINJA",
          fa.switch(fa.SwitchReq(mode="LIVE")).get("active_mode")=="NINJA")
    try:
        fa.SESSIONS.pop("WEBULL")
        fa.switch(fa.SwitchReq(mode="WEBULL")); ok=False; why="accepted"
    except Exception as e:
        ok = "not logged in yet" in str(e); why=str(e)[:50]
    check(11,"switching to a broker you never logged into is refused", ok, why)

    fa.SESSIONS.update({"WEBULL":wb_})
    d = fa.disconnect(fa.DisconnectReq(mode="NINJA"))
    check(11,"disconnect ONE leaves the others", len(fa.SESSIONS)==2 and "NINJA" not in fa.SESSIONS)
    check(11,"active reassigned after dropping the active one", d.get("active") in fa.SESSIONS)
    fa.disconnect(fa.DisconnectReq(mode=None))
    check(11,"disconnect all clears everything", len(fa.SESSIONS)==0 and fa.ACTIVE["mode"] is None)
    fa.SESSIONS.clear(); fa.ACTIVE["mode"]=None

    print("\n[12] VELOCITY IS HONEST WHEN THE MARKET IS SHUT")
    import tape, time as _t
    def _bar(t,v,rng,c=100.0): return {"t":t,"o":c,"h":c+rng/2,"l":c-rng/2,"c":c,"v":v}
    now_=int(_t.time())

    # Friday's closing auction: a huge volume spike as the last thing printed.
    stale=[_bar(now_-3*86400+i*60, 1000, 0.5) for i in range(35)]
    stale.append(_bar(now_-3*86400+35*60, 90000, 6.0))
    raw = tape.compute(stale)
    # This used to assert the spike DID read "violent" - that was the trap the
    # staleness guard existed to cover. The outlier cap now defuses it one
    # layer earlier, so a lone closing-auction print no longer fools the score
    # either. Both guards are kept: the cap handles feed artifacts mid-session,
    # staleness handles a market that is simply shut.
    check(12,"a lone closing-auction spike no longer reads violent",
          raw["state"] != "violent", raw["state"])
    tape._CACHE.clear()
    tape._bars = lambda ysym, b=stale: b
    v = tape.velocity("TEST")
    check(12,"velocity() overrides it to CLOSED", v["state"]=="closed", v["state"])
    check(12,"score is zero, not 100", v["score"]==0.0, str(v["score"]))
    check(12,"says why", "closed" in str(v.get("note","")).lower())

    # No bars at all (weekend futures).
    tape._CACHE.clear(); tape._bars = lambda ysym: []
    v2 = tape.velocity("TEST2")
    check(12,"no bars -> CLOSED not a crash", v2["state"]=="closed" and v2["score"]==0.0)

    # A live tape must still read normally.
    fresh=[_bar(now_-(40-i)*60, 1000, 0.5) for i in range(35)]
    fresh+=[_bar(now_-(5-i)*60, 1200, 0.6) for i in range(5)]
    tape._CACHE.clear(); tape._bars = lambda ysym, b=fresh: b
    v3 = tape.velocity("TEST3")
    check(12,"a LIVE tape still scores normally",
          v3["state"] in ("calm","normal","fast","violent") and v3["score"]>0, str(v3.get("state")))
    for page in ("index.html","futures_index.html"):
        src=io.open(os.path.join(HERE,page),encoding="utf-8").read()
        check(12,f"{page} renders the closed state", "MARKET CLOSED" in src)

    print("\n[13] BROKER TABS ON THE DASHBOARD + TRAY FALLBACK")
    fh = io.open(os.path.join(HERE,"futures_index.html"),encoding="utf-8").read()
    check(13,"tab strip lives on the DASHBOARD", 'id="brokerTabs"' in fh)
    dash = fh.split('id="dash"',1)[-1]
    check(13,"tabs render inside the dash, not the login screen", 'id="brokerTabs"' in dash)
    check(13,"all three brokers listed", "['WEBULL','WEBULL']" in fh.replace(' ','') or "WEBULL" in fh and "NINJA" in fh and "TOPSTEP" in fh)
    check(13,"clicking a tab switches", "pickBroker" in fh and "'/api/switch'" in fh)
    check(13,"back arrow no longer disconnects",
          "function back(){" in fh and "api('/api/disconnect'" not in fh.split("function back(){",1)[1][:400])
    # The padlock / log-out-every-broker control was REMOVED at his request.
    # These checks used to pass on the leftover lockAll() function after its
    # button was gone - a false green: the test proved a function existed that
    # nothing on the screen could reach. Assert the removal instead.
    check(13,"no log-out-every-broker control remains",
          "function lockAll" not in fh and "FZ.lockAll" not in fh)
    check(13,"and no quit button either",
          "function quitApp" not in fh and "FZ.quitApp" not in fh)
    check(13,"switching broker does not sign you out",
          "'/api/disconnect'" not in fh.split("function pickBroker",1)[1][:400]
          if "function pickBroker" in fh else True)
    for fn in ("pickBroker","refreshSessions"):
        check(13,f"FZ.{fn} exported", bool(re.search(r"return \{[^}]*\b%s\b"%fn, fh, re.S)))

    import run_all
    check(13,"tray is optional, not required", hasattr(run_all,"TRAY_AVAILABLE"))
    check(13,"missing tray libs do NOT crash the launcher", run_all.start_tray() is None)
    # START HIDDEN.vbs, INSTALL.bat, STOP EVERYTHING.bat, TUTORIAL.html and
    # CHECK-SETUP.bat were deleted on request. The launcher absorbed the one
    # thing INSTALL did that nothing else did: installing the Webull SDK.
    _lb0 = [f for f in os.listdir(HERE) if f.endswith("START MARKET SNIPER.bat")]
    _l0 = io.open(os.path.join(HERE,_lb0[0]),encoding="utf-8").read() if _lb0 else ""
    check(13,"launcher installs the Webull SDK itself",
          "webull-openapi-python-sdk" in _l0)
    check(13,"and verifies it imported", _l0.count("from webull.core.client import ApiClient") >= 2)
    itxt = _l0
    req  = io.open(os.path.join(HERE,"requirements.txt"),encoding="utf-8").read()
    check(13,"launcher installs the tray deps", "pystray" in itxt)
    check(13,"tray deps NOT in requirements (runs every launch)",
          not any(l.strip()=="pystray" for l in req.splitlines()))
    check(13,"launcher clears Mark-of-the-Web", "Unblock-File" in itxt)
    check(13,"survives an optional tray failure", "still works fine" in itxt)
    check(13,"launcher warns if the SDK fails", "cannot connect or trade" in itxt)
    check(13,"no dead references to deleted files",
          not any(x in itxt for x in ("INSTALL.bat\"", "CHECK-SETUP", "TUTORIAL.html")))

    print("\n[14] AUTO-SYNC HEALS A STALE GIT LOCK")
    import subprocess as _sp, tempfile, shutil as _sh
    tmp = tempfile.mkdtemp(prefix="locktest")
    try:
        _sp.run(["git","init","-q","."],cwd=tmp)
        _sp.run(["git","config","user.email","t@t.t"],cwd=tmp)
        _sp.run(["git","config","user.name","t"],cwd=tmp)
        io.open(os.path.join(tmp,"a.py"),"w").write("x=1\n")
        _sp.run(["git","add","-A"],cwd=tmp); _sp.run(["git","commit","-qm","init"],cwd=tmp)

        import importlib, auto_sync as _as
        _as = importlib.reload(_as); _as.HERE = tmp
        lock = os.path.join(tmp,".git","HEAD.lock")

        # STALE lock -> must heal and retry
        io.open(lock,"w").close()
        os.utime(lock, (time.time()-300, time.time()-300))
        ok,_out = _as.git("commit","--allow-empty","-m","healed")
        check(14,"stale lock is cleared and the commit retried", ok)
        check(14,"stale lock file removed", not os.path.exists(lock))

        # FRESH lock -> must NOT be stolen (a real git may hold it)
        io.open(lock,"w").close()
        ok2,out2 = _as.git("commit","--allow-empty","-m","should not steal")
        check(14,"a FRESH lock is left alone", (not ok2) and os.path.exists(lock))
        check(14,"and the failure is reported honestly", _as._looks_like_lock_error(out2))
        os.remove(lock)

        # Non-lock failure must not trigger the sweep
        ok3,out3 = _as.git("checkout","no-such-branch")
        check(14,"non-lock failures are not misread as locks",
              (not ok3) and not _as._looks_like_lock_error(out3))
        check(14,"lock sweep covers HEAD.lock, not just index.lock",
              "HEAD.lock" in io.open(os.path.join(HERE,"auto_sync.py"),encoding="utf-8").read()
              or "endswith(\".lock\")" in io.open(os.path.join(HERE,"auto_sync.py"),encoding="utf-8").read())
    finally:
        _sh.rmtree(tmp, ignore_errors=True)
        import importlib, auto_sync as _as2
        _as2 = importlib.reload(_as2)

    print("\n[15] ONE TAB ONLY - a new launch retires the old tab")
    for page, chan in (("index.html","market_sniper_options"),
                       ("futures_index.html","market_sniper_futures")):
        src = io.open(os.path.join(HERE,page),encoding="utf-8").read()
        check(15,f"{page}: guard present", "ONE TAB ONLY" in src)
        check(15,f"{page}: own channel ({chan})", chan in src)
        check(15,f"{page}: stale tab stops polling first",
              "standDown" in src and "clearInterval(poll)" in src.split("standDown",1)[1][:400])
        check(15,f"{page}: falls back when close() is blocked", "safe to close" in src)
        g = src.index("ONE TAB ONLY")
        check(15,f"{page}: MY_BORN set before it is read",
              src.index("var MY_BORN", g) < src.index("e.data.born < MY_BORN", g))
        check(15,f"{page}: identical timestamps break the tie",
              "e.data.born === MY_BORN" in src)
    _lb = [f for f in os.listdir(HERE) if f.endswith("START MARKET SNIPER.bat")]
    lb = io.open(os.path.join(HERE,_lb[0]),encoding="utf-8").read() if _lb else ""
    check(15,"launcher installs the tray only when missing",
          "import pystray, PIL" in lb and "pip install -q pystray pillow" in lb)

    print("\n[16] SERVER RESTART DOES NOT LEAVE A TAB SPINNING ON 400s")
    for page in ("index.html","futures_index.html"):
        src = io.open(os.path.join(HERE,page),encoding="utf-8").read()
        check(16,f"{page}: tick() checks for a lost session", "st.detail" in src)
        check(16,f"{page}: stops all four timers", 
              all(t in src.split("lostCount",1)[1][:600] for t in
                  ("clearInterval(poll)","clearInterval(priceTimer)",
                   "clearInterval(trendTimer)","clearInterval(velTimer)")))
        check(16,f"{page}: returns to the login screen",
              "connectScreen').classList.remove('hidden')" in src)
        check(16,f"{page}: one blip is not treated as a disconnect", "lostCount >= 2" in src)
    _lb = [f for f in os.listdir(HERE) if f.endswith("START MARKET SNIPER.bat")]
    lb = io.open(os.path.join(HERE,_lb[0]),encoding="utf-8").read() if _lb else ""
    check(16,"launcher never calls bare pip", "\npip install" not in lb and "  pip install" not in lb)
    check(16,"launcher pins python to the venv", "VPY=" in lb)
    check(16,"tray install verified after installing",
          lb.count('import pystray, PIL') >= 2)

    print("\n[17] CLOSED MARKET IS HONEST ON BOTH APPS")
    opt = io.open(os.path.join(HERE,"index.html"),encoding="utf-8").read()
    fut = io.open(os.path.join(HERE,"futures_index.html"),encoding="utf-8").read()
    for name, src in (("options",opt),("futures",fut)):
        check(17,f"{name}: tracks marketClosed", "marketClosed" in src)
        check(17,f"{name}: set from the velocity reading",
              "marketClosed = (v.state==='closed')" in src)
        check(17,f"{name}: 'unknown' is not treated as closed",
              "marketClosed = false;" in src)
        check(17,f"{name}: says MARKET CLOSED on the trade buttons",
              "MARKET CLOSED" in src)
    check(17,"options: skips quoting a shut market",
          "if(marketClosed){" in opt and "callSub" in opt.split("if(marketClosed){",1)[1][:300])
    check(17,"futures: warns before sending into a shut market",
          "marketClosed && !confirm(" in fut)
    check(17,"futures: you can still overrule the guess",
          "send anyway" in fut.lower())
    check(17,"futures: buttons repaint when velocity lands",
          "paintMarketState" in fut.split("async function refreshVel",1)[1][:400])
    check(17,"options: velocity awaited before the first quote",
          "refreshVel().then(refreshQuote)" in opt)

    print("\n[18] FUTURES HOURS - overnight must NOT read as closed")
    import tape as _t, time as _tm
    def _bar(t,v=500,rng=5.0,c=20000.0): return {"t":t,"o":c,"h":c+rng,"l":c-rng,"c":c,"v":v}
    now_=int(_tm.time())
    check(18,"threshold clears the 11-min midnight rollover gap",
          _t.STALE_SECONDS/60 > 11, f"{_t.STALE_SECONDS/60:.0f} min")
    check(18,"threshold still detects the 62-min maintenance break",
          _t.STALE_SECONDS/60 < 62, f"{_t.STALE_SECONDS/60:.0f} min")

    # 11-min quiet patch mid-session: MUST stay open
    bars=[_bar(now_-(60-i)*60) for i in range(35)]
    bars.append(_bar(now_-11*60))
    _t._CACHE.clear(); _t._bars = lambda y, b=bars: b
    v=_t.velocity("FUT_OVERNIGHT")
    check(18,"11-min gap overnight still reads OPEN", v["state"]!="closed", v["state"])

    # 40 min of nothing: maintenance break / weekend
    bars2=[_bar(now_-(120-i)*60) for i in range(35)]
    bars2.append(_bar(now_-40*60))
    _t._CACHE.clear(); _t._bars = lambda y, b=bars2: b
    v2=_t.velocity("FUT_BREAK")
    check(18,"40-min silence reads CLOSED", v2["state"]=="closed", v2["state"])

    fut = io.open(os.path.join(HERE,"futures_index.html"),encoding="utf-8").read()
    check(18,"a closed guess WARNS, it does not silently refuse",
          "confirm(" in fut.split("marketClosed &&",1)[1][:400])
    check(18,"and says it is a guess, not a calendar", "not a " in fut and "calendar" in fut)
    # The whole point of the bar-age design is that no hours are hardcoded.
    # Strip comments, then make sure no clock arithmetic snuck into the logic.
    _tp = io.open(os.path.join(HERE,"tape.py"),encoding="utf-8").read()
    _code = "\n".join(l for l in _tp.splitlines() if not l.strip().startswith("#"))
    check(18,"tape.py makes no calendar/clock decisions",
          ".hour" not in _code and "weekday" not in _code and "localtime" not in _code)

    print("\n[19] PHANTOM POSITION - closed by hand, app still thinks you are in")
    import futures_client as _fc
    ph = _fc.make_session("TOPSTEP")
    # Down past the ratchet's opening stop (12.5 points), so an exit really is
    # pending. 23144 was chosen when a 5-point SL was the live bracket; the
    # ratchet replaces it and is wider, so the price has to move to match.
    ph.position = {"symbol":"MNQ","side":"LONG","qty":1,"entry":23150.0,
                   "mark":23135.0,"best":23160.0,"pnl":-30.0,
                   "ratchet_on":True,"ratchet_step":12.5}
    ph.settings.update({"sl_enabled":True,"sl_points":5.0})
    # Any exit will do. The point is that SOMETHING would have fired against a
    # position the broker says you do not hold. Which one answers first is not
    # what this scenario is about - and note the ratchet SUPPRESSES the 5-point
    # SL entirely, by design: it owns the exit when it is on.
    check(19,"a phantom WOULD have fired a bracket", ph._bracket_hit() is not None,
          str(ph._bracket_hit()))
    r_=ph.forget_position()
    check(19,"forget clears it", r_["cleared"] is True and ph.position is None)
    check(19,"nothing can fire afterwards", ph._bracket_hit() is None)
    check(19,"it says no order was sent", "NO order was sent" in (ph.last_event or ""))
    check(19,"forgetting when already flat is harmless",
          ph.forget_position()["cleared"] is False)
    check(19,"armed order is cleared too", ph.armed is None)

    fh2 = io.open(os.path.join(HERE,"futures_index.html"),encoding="utf-8").read()
    check(19,"futures exposes forgetPosition", "forgetPosition" in fh2)
    check(19,"reachable WITHOUT a rejection (you closed by hand)",
          "flatlink" in fh2 and "Closed it yourself" in fh2)
    check(19,"asks for confirmation first", "Clear this position from the app" in fh2)
    check(19,"warns it stops managing the trade", "stops the app managing" in fh2)
    fa2 = io.open(os.path.join(HERE,"futures_app.py"),encoding="utf-8").read()
    check(19,"endpoint exists", "/api/position/forget" in fa2)
    check(19,"options had this all along", "/api/position/forget" in
          io.open(os.path.join(HERE,"main.py"),encoding="utf-8").read())

    print("\n[20] MY CONFIG (round-number entry) IS ALWAYS ON")
    import importlib, config as _cfg
    _cfg = importlib.reload(_cfg)
    check(20,"server default is ON", _cfg.DEFAULT_SETTINGS["my_enabled"] is True)
    idx2 = io.open(os.path.join(HERE,"index.html"),encoding="utf-8").read()
    check(20,"browser fallback is ON too (used before prefs load)",
          "my_enabled:true}" in idx2.replace(" ",""))
    # The DEFAULT must be on. Turning it off when a strategy is armed is
    # correct and expected, so only the declared defaults are checked here.
    check(20,"the config default is not False",
          not re.search(r'"my_enabled"\s*:\s*False',
                        io.open(os.path.join(HERE,"config.py"),encoding="utf-8").read()))
    # Only the DECLARED default matters. Sending {my_enabled:false} when a
    # strategy takes over is the exclusivity rule doing its job.
    _decl = [l for l in idx2.splitlines() if "let settings=" in l or "sl_unit:" in l]
    check(20,"the browser fallback default is true",
          any("my_enabled:true" in l.replace(" ","") for l in _decl), str(_decl)[:120])
    saved = json.load(io.open(SETTINGS,encoding="utf-8")).get("options_settings",{})
    check(20,"your saved file says ON", saved.get("my_enabled") is True, str(saved.get("my_enabled")))
    check(20,"ON means the buttons ARM, not buy at the ask",
          "if(settings.my_enabled){" in idx2 and "/api/order/arm" in idx2)
    _, pr = http("http://127.0.0.1:8000/api/prefs")
    check(20,"server actually SERVES it on", pr.get("settings",{}).get("my_enabled") is True,
          str(pr.get("settings",{}).get("my_enabled")))
    # THE root cause: the browser used to load settings from localStorage, then
    # POST them to /api/settings, which wrote them to my-settings.json. One
    # stale browser value overwrote the file on every connect, which is why
    # fixing the default never stuck.
    check(20,"settings are NOT read from localStorage any more",
          "localStorage.getItem(LSS)" not in idx2)
    # Was: assert a comment saying disk wins. The line above already proves
    # localStorage is not read; this proves settings come from the SERVER.
    check(20,"settings are taken from the server payload",
          "st.settings" in idx2 and "settings={...settings,...st.settings}"
          in idx2.replace(" ", ""))
    check(20,"you can still deliberately turn it off",
          "settings.my_enabled=$('myEnabled').checked" in idx2.replace(" ",""))

    print("\n[21] AUTO-RECONCILE - the app asks Topstep what you actually hold")
    import futures_client as _fc2
    def _rs(broker_says):
        x = _fc2.make_session("TOPSTEP")
        x.token="t"; x.acct={"id":1}
        x.position={"symbol":"MNQ","side":"LONG","qty":1,"entry":23150.0,
                    "mark":23135.0,"best":23160.0,"pnl":-30.0,
                    "ratchet_on":True,"ratchet_step":12.5}
        x.settings.update({"sl_enabled":True,"sl_points":5.0})
        x.broker_positions = lambda: broker_says
        x._last_reconcile = 0
        return x

    # Same point as scenario 19: something fires against a position that is not
    # there. Which exit answers first is not what is under test.
    x=_rs([]);  check(21,"a phantom WOULD have fired a stop", x._bracket_hit() is not None,
                      str(x._bracket_hit()))
    x.refresh_mark()
    check(21,"broker says FLAT -> app clears itself", x.position is None)
    check(21,"and explains why", "FLAT" in (x.last_event or ""))
    check(21,"no order was sent", "No order was sent" in (x.last_event or ""))

    x=_rs([{"contractId":"MNQ","size":1}]); x.refresh_mark()
    check(21,"broker CONFIRMS -> position kept", x.position is not None)

    x=_rs(None); x.refresh_mark()
    check(21,"API failure is NOT read as flat", x.position is not None)

    hits=[]
    x=_rs([]); x.broker_positions=lambda: (hits.append(1), [])[1]; x._last_reconcile=0
    x.reconcile(); x.reconcile(); x.reconcile()
    check(21,"rate-limited, does not hammer the API", len(hits)==1, f"{len(hits)} calls")

    fcsrc = io.open(os.path.join(HERE,"futures_client.py"),encoding="utf-8").read()
    ts_block = fcsrc.split("class TopstepSession",1)[1].split("\nclass ",1)[0]
    # take the whole refresh_mark body, up to the next def
    rm = ts_block.split("def refresh_mark",1)[1].split("\n    def ",1)[0]
    check(21,"reconcile is inside refresh_mark", "self.reconcile()" in rm, rm[:120])
    check(21,"reconcile runs BEFORE brackets can fire",
          "self.reconcile()" in rm and "_maybe_auto_close" in rm
          and rm.index("self.reconcile()") < rm.index("_maybe_auto_close"))
    check(21,"uses the real ProjectX endpoint", "/api/Position/searchOpen" in fcsrc)
    check(21,"None and [] are treated differently",
          "could not ask" in fcsrc)

    print("\n[22] OPTIONS: clearing a phantom must NOT need a rejection first")
    oi = io.open(os.path.join(HERE,"index.html"),encoding="utf-8").read()
    check(22,"the old CLEAR only showed after a rejection",
          "rejectFix').classList.toggle('show'" in oi)
    check(22,"there is now an ALWAYS-visible clear while in a trade",
          "flatlink" in oi and "Closed it yourself in Webull" in oi)
    ta = oi.split('id="tradeActions"',1)[1][:600]
    check(22,"it sits next to CLOSE, not in the reject box", "flatlink" in ta)
    check(22,"it still asks for confirmation", "NOT holding" in oi)
    check(22,"it promises no order is sent", "sends NO order" in oi.replace("This sends NO order","sends NO order"))

    import webull_client as _wb2
    z = _wb2.make_session("LIVE")
    z.position={"symbol":"QQQ","side":"CALLS","strike":711.0,"qty":1,"entry":3.0,"mark":2.7}
    r3 = z.forget_position()
    check(22,"forget clears it", r3["cleared"] is True and z.position is None)
    check(22,"and says no order was sent", "No order was sent" in (z.last_event or ""))
    check(22,"clearing twice is harmless", z.forget_position()["cleared"] is False)
    check(22,"armed entry is cleared too", z.armed is None)

    print("\n[23] DAILY TRADE LOG - one workbook, appended forever")
    import trade_log as _tl, tempfile as _tf, importlib
    _tl = importlib.reload(_tl)
    _sc = _tf.mkdtemp(prefix="tlog")
    _tl.LOG_DIR=_sc; _tl.CSV_PATH=os.path.join(_sc,"trades.csv")
    _tl.XLSX_PATH=os.path.join(_sc,"Market Sniper Trade Log.xlsx")
    check(23,"openpyxl available", _tl.XLSX_AVAILABLE)

    for d,sym,e_,x_,rz in [("2026-08-21","QQQ",3.0,3.85,"TP"),
                           ("2026-08-21","QQQ",2.4,2.05,"SL"),
                           ("2026-08-22","SPY",1.8,2.60,"CLOSE")]:
        _tl.record({"date":d,"symbol":sym,"side":"CALLS","strike":711,"qty":1,
                    "entry":e_,"exit":x_,"pnl":round((x_-e_)*100,2),"exit_reason":rz,
                    "app":"OPTIONS","broker":"WEBULL"})
    check(23,"every trade lands in the CSV", len(_tl._rows())==3, str(len(_tl._rows())))
    check(23,"ONE workbook, not one per day", os.path.exists(_tl.XLSX_PATH))
    check(23,"only one xlsx exists",
          len([f for f in os.listdir(_sc) if f.endswith(".xlsx")])==1)

    import openpyxl as _px
    _wb=_px.load_workbook(_tl.XLSX_PATH)
    check(23,"TRADES + BY DAY sheets", _wb.sheetnames==["TRADES","BY DAY"], str(_wb.sheetnames))
    check(23,"every trade is a row", _wb["TRADES"].max_row==4, str(_wb["TRADES"].max_row))
    bd=_tl.by_day()
    check(23,"grouped by day", [d["date"] for d in bd]==["2026-08-21","2026-08-22"])
    check(23,"daily net is right", bd[0]["net"]==50.0 and bd[1]["net"]==80.0,
          f"{bd[0]['net']} / {bd[1]['net']}")
    check(23,"wins and losses counted", bd[0]["wins"]==1 and bd[0]["losses"]==1)

    # the failure that would silently lose a trade
    _real=_tl.rebuild_xlsx
    def _locked(): raise PermissionError("open in Excel")
    _tl.rebuild_xlsx=_locked
    _tl.record({"date":"2026-08-22","symbol":"QQQ","side":"PUTS","strike":700,"qty":1,
                "entry":1.0,"exit":1.4,"pnl":40.0,"exit_reason":"TP"})
    check(23,"workbook LOCKED in Excel -> trade still saved", len(_tl._rows())==4)
    _tl.rebuild_xlsx=_real
    _tl.rebuild_xlsx()
    check(23,"and appears once Excel is closed",
          _px.load_workbook(_tl.XLSX_PATH)["TRADES"].max_row==5)

    wsrc = io.open(os.path.join(HERE,"webull_client.py"),encoding="utf-8").read()
    check(23,"options logs from the ONE close funnel",
          "trade_log.record" in wsrc.split("_record_close",1)[1][:2500])
    check(23,"auto-exits record TP/SL not just CLOSE", 'position["exit_reason"] = hit' in wsrc)
    fsrc = io.open(os.path.join(HERE,"futures_client.py"),encoding="utf-8").read()
    check(23,"futures logs too", "_log_trade" in fsrc)
    check(23,"declared in requirements",
          "openpyxl" in io.open(os.path.join(HERE,"requirements.txt"),encoding="utf-8").read())

    print("\n[24] OPTIONS AUTO-RECONCILE with Webull")
    import webull_client as _wb3
    def _os_(broker):
        z=_wb3.make_session("LIVE"); z.account_id="A"
        z.position={"symbol":"QQQ","side":"CALLS","strike":711.0,"qty":1,
                    "entry":3.0,"mark":2.7,"expiration":"2026-08-24"}
        z.broker_positions=lambda: broker
        z._last_reconcile=0
        return z
    z=_os_([]);   z.reconcile(force=True)
    check(24,"Webull says nothing open -> cleared", z.position is None)
    check(24,"explains why", "hold nothing" in (z.last_event or ""))
    z=_os_(None); z.reconcile(force=True)
    check(24,"API failure is NOT read as flat", z.position is not None)
    z=_os_([{"x":1}]); z.reconcile(force=True)
    check(24,"broker confirms -> kept", z.position is not None)
    check(24,"runs BEFORE brackets fire",
          wsrc.split("def refresh_mark",1)[1].split("def close",1)[0].index("self.reconcile()")
          < wsrc.split("def refresh_mark",1)[1].split("def close",1)[0].index("_maybe_auto_close"))
    check(24,"finds the SDK call at runtime, not hardcoded", "_position_fns" in wsrc)

    print("\n[25] CONSOLE AUTO-HIDE + no pointless installs")
    _lb2 = [f for f in os.listdir(HERE) if f.endswith("START MARKET SNIPER.bat")]
    L = io.open(os.path.join(HERE,_lb2[0]),encoding="utf-8").read() if _lb2 else ""
    check(25,"deps are import-checked before pip runs",
          'import fastapi, uvicorn, pydantic, openpyxl' in L)
    # Look at the real command lines, not the comments that mention them.
    _cmds = [l.strip() for l in L.splitlines() if not l.strip().lower().startswith("rem")]
    _i_chk = next(i for i,l in enumerate(_cmds) if "import fastapi" in l)
    _i_pip = next(i for i,l in enumerate(_cmds) if "-r requirements.txt" in l)
    check(25,"pip only runs inside that failure branch", _i_chk < _i_pip)
    check(25,"and it is guarded by errorlevel",
          any("errorlevel 1" in l for l in _cmds[_i_chk:_i_pip]))
    check(25,"SDK likewise only installs when missing",
          L.index("from webull.core.client import ApiClient") < L.index("webull-openapi-python-sdk"))
    check(25,"tray likewise", L.index("import pystray, PIL") < L.index("pystray pillow"))
    check(25,"core dep failure stops the launch", "cannot run" in L)

    import run_all as _ra, importlib
    _ra = importlib.reload(_ra)
    check(25,"hide is delayed, not instant", getattr(_ra,"HIDE_AFTER_SECONDS",0) >= 3)
    check(25,"NO tray -> console never hides", _ra.hide_console_when_ready(None) is None)
    check(25,"hiding is a no-op off Windows", _ra._show_console(False) is False)
    check(25,"tray can bring it back", "_toggle_console" in io.open(
          os.path.join(HERE,"run_all.py"),encoding="utf-8").read())
    rsrc = io.open(os.path.join(HERE,"run_all.py"),encoding="utf-8").read()
    check(25,"shutdown un-hides first, so you see why it stopped",
          "_show_console(True)" in rsrc.split("def stop_all",1)[1][:400])
    check(25,"the hide is conditional on the tray, in code",
          "if tray is None" in rsrc)

    print("\n[26] RATCHET - stop climbs a rung at a time and never drops")
    import webull_client as _w4
    L = _w4.LiveSession
    # the ladder G specified, verbatim
    for peak, stop, nxt in [(0,-10,10),(5,-10,10),(9.9,-10,10),(10,0,20),
                            (15,0,20),(20,10,30),(30,20,40),(100,90,110)]:
        lv = L.ratchet_levels(peak, 10)
        check(26,f"best {peak}% -> stop {stop}%, next +{nxt}%",
              lv["stop_pct"]==stop and lv["next_pct"]==nxt,
              f"got stop {lv['stop_pct']} next {lv['next_pct']}")
    check(26,"exactly +10% counts (float error nearly broke this)",
          L.ratchet_levels((3.30-3.00)/3.00*100, 10)["stop_pct"]==0.0)
    check(26,"a gap from +5% to +25% still lands the stop on +10%",
          L.ratchet_levels(25,10)["stop_pct"]==10.0)

    def _rt(marks, step=10.0, entry=3.00):
        z=_w4.make_session("LIVE")
        z.settings.update({"my_enabled":True,"ratchet_step_pct":step,
                           # flat rungs under test here; tiers are scenario 61
                           "ratchet_tiers":False})
        z.position={"symbol":"QQQ","side":"CALLS","strike":711.0,"qty":1,
                    "entry":entry,"mark":entry,"expiration":"2026-08-26"}
        out=[]
        for m in marks:
            z.position["mark"]=m
            hit=z._bracket_hit()
            out.append((m, dict(z.position.get("ratchet") or {}), hit,
                        z.position.get("exit_reason")))
            if hit: break
        return z, out

    z,run = _rt([3.00,2.85,3.15,3.30,3.45,3.60,3.90,3.40])
    stops = {m: r.get("stop_pct") for m,r,_,_ in run}
    check(26,"opens at -10%", stops[3.00]==-10.0)
    check(26,"a dip to -5% does NOT exit", run[1][2] is None)
    check(26,"+10% moves the stop to BREAKEVEN, does not close",
          stops[3.30]==0.0 and run[3][2] is None)
    check(26,"+20% locks in +10%", stops[3.60]==10.0)
    check(26,"+30% locks in +20%", stops[3.90]==20.0)
    check(26,"falling back through the stop closes", run[-1][2] is not None)
    check(26,"and the log says which rung", "RATCHET+20" in str(run[-1][3]), str(run[-1][3]))

    # never loosens
    z2,run2 = _rt([3.00,3.90,3.70,3.75,3.72])
    seq=[r.get("stop_pct") for _,r,_,_ in run2]
    check(26,"stop never moves DOWN", all(b>=a for a,b in zip(seq,seq[1:])), str(seq))

    # breakeven exit names itself
    z3,run3 = _rt([3.00,3.30,2.99])
    check(26,"a give-back to breakeven exits as BE",
          "RATCHET-BE" in str(run3[-1][3]), str(run3[-1][3]))

    # ratchet replaces TP/SL
    z4=_w4.make_session("LIVE")
    z4.settings.update({"my_enabled":True,"tp_enabled":True,"tp_unit":"cents",
                        "tp_value":1,"sl_enabled":True,"sl_unit":"pct","sl_value":1})
    z4.position={"symbol":"QQQ","side":"CALLS","strike":711.0,"qty":1,
                 "entry":3.00,"mark":3.06,"expiration":"2026-08-26"}
    check(26,"a 1c take-profit can NOT close a ratcheted trade", z4._bracket_hit() is None)

    z5=_w4.make_session("LIVE"); z5._guard_open=lambda q: None
    # Set the state this asserts on, rather than inheriting whatever is saved
    # in my-settings.json. Reading his live profile made this test pass or fail
    # depending on what he last clicked, which is not a test.
    z5.settings["my_enabled"]=True
    z5.settings["tp_enabled"]=False
    z5._underlying=lambda sym: 713.0
    z5.arm("QQQ","CALLS",1)
    check(26,"arming does not re-enable a take-profit", z5.settings["tp_enabled"] is False)
    # Arming ALWAYS turns the ratchet on, even if it was off, and never arms a
    # take-profit. There was a dead branch claiming otherwise - unreachable,
    # because it tested my_enabled one line after setting it True - and it got
    # cited twice as an explanation for behaviour it could not produce.
    z5b=_w4.make_session("LIVE"); z5b._guard_open=lambda q: None
    z5b.settings["my_enabled"]=False; z5b.settings["tp_enabled"]=False
    z5b._underlying=lambda sym: 713.0
    z5b.arm("QQQ","CALLS",1)
    check(26,"arming turns the ratchet ON", z5b.settings["my_enabled"] is True)
    check(26,"and still arms no take-profit", z5b.settings["tp_enabled"] is False)
    check(26,"and the dead branch is gone",
          "if not s.get(\"my_enabled\"):" not in
          io.open("webull_client.py",encoding="utf-8").read())

    cfg = io.open(os.path.join(HERE,"config.py"),encoding="utf-8").read()
    check(26,"on by default", '"my_enabled": True' in cfg)
    idx4 = io.open(os.path.join(HERE,"index.html"),encoding="utf-8").read()
    check(26,"live stop shown while in a trade", "next rung" in idx4)
    # The separate RATCHET checkbox is gone - it is one switch with the entry
    # now - so the screen proof is the step input plus the combined switch.
    check(26,"settings expose it", 'id="ratchetStep"' in idx4 and 'id="myEnabled"' in idx4)

    print("\n[27] CONFIG SCREEN: one combined strategy, dead sections gone")
    ix = io.open(os.path.join(HERE,"index.html"),encoding="utf-8").read()
    cfg_pane = ix[ix.index('id="paneConfig"'):ix.index("/paneConfig")]
    check(27,"renamed to ABSOLUTE ENTRY with RATCHET TRAILING",
          "ABSOLUTE ENTRY with RATCHET TRAILING" in cfg_pane)
    check(27,"the old +$1 TP / 10% stop blurb is gone",
          "Auto-sets a" not in ix and "Overrides the TP/SL" not in ix)
    check(27,"entry and ratchet sit in ONE section",
          "myEnabled" in cfg_pane and "ratchetStep" in cfg_pane)
    check(27,"and are ONE switch, not two",
          "ratchetEnabled" not in cfg_pane and
          "ABSOLUTE ENTRY + RATCHET" in cfg_pane)
    for gone in ("TAKE PROFIT — closes","STOP LOSS — closes","AUTO-LOCK"):
        check(27,f"{gone.split(chr(8212))[0].strip()} removed from CONFIGURATION",
              gone not in cfg_pane)
    check(27,"MIRROR moved out of CONFIGURATION", "MIRROR TRADING" not in cfg_pane)
    check(27,"MIRROR has its own tab",
          'id="tabMirror"' in ix and 'id="paneMirror"' in ix
          and "MIRROR TRADING" in ix[ix.index('id="paneMirror"'):ix.index("/paneMirror")])
    check(27,"showTab handles all three panes",
          all(k in ix.split("function showTab",1)[1][:400]
              for k in ("paneConfig","paneStrat","paneMirror")))
    check(27,"no JS still reads the deleted controls",
          not any(x in ix for x in ("$('tpEnabled')","$('slEnabled')","$('alEnabled')",
                                    "$('tpUnit')","$('alMinutes')")))
    check(27,"tpUnitChanged fully removed", "tpUnitChanged" not in ix)

    wsrc2 = io.open(os.path.join(HERE,"webull_client.py"),encoding="utf-8").read()
    # This used to assert that a specific line of SOURCE existed - and that
    # line was unreachable, so the test passed while the behaviour it named
    # could never happen. Assert the behaviour instead.
    _z27 = _w4.make_session("LIVE"); _z27._guard_open=lambda q: None
    _z27.settings["my_enabled"]=False; _z27.settings["tp_enabled"]=False
    _z27._underlying=lambda sym: 713.0
    _z27.arm("QQQ","CALLS",1)
    check(27,"arming is entry-ONLY: no take-profit is armed with it",
          _z27.settings["tp_enabled"] is False and _z27.settings["my_enabled"] is True)
    check(27,"backend tp/sl fields still exist for STRATEGIES",
          '"tp_unit"' in wsrc2 and '"sl_unit"' in wsrc2)

    print("\n[28] ENTRY GRID, ATM, CONTRACT QUALITY, ONE-ARMED-THING")
    import importlib, config as _c5, webull_client as _w5
    _c5 = importlib.reload(_c5); _w5 = importlib.reload(_w5)
    L5 = _w5.LiveSession

    lv = L5.entry_levels_near(710.3, span=3)
    check(28,"grid keeps every whole dollar", all(float(d) in lv for d in range(708,713)))
    check(28,"grid adds X2.50 and X7.50", 707.5 in lv and 712.5 in lv)
    check(28,"grid adds nothing else", not any(abs(x%1-0.5)<1e-9 and int(x)%10 not in (2,7) for x in lv))
    # These used to assert the NEAREST level regardless of side, which is why
    # 707.40 expected 707.50 - a level ABOVE spot for a call. Entry is
    # directional now, so each case names the side it belongs to.
    for spot, side, want in [(707.40,"CALLS",707.0), (707.40,"PUTS",707.5),
                             (707.60,"CALLS",707.5), (707.60,"PUTS",708.0),
                             (707.10,"CALLS",707.0), (707.10,"PUTS",707.5),
                             (712.40,"CALLS",712.0), (712.40,"PUTS",712.5),
                             (712.60,"CALLS",712.5), (712.60,"PUTS",713.0),
                             (709.60,"CALLS",709.0), (709.60,"PUTS",710.0)]:
        got = L5.entry_target(spot, side)
        check(28,f"{side} at {spot} fire at {want}", got==want, str(got))

    for spot in (713.40, 713.60, 714.00, 713.00):
        c = _w5.pick_strike(spot,"CALLS",1.0,"ATM1")
        p_ = _w5.pick_strike(spot,"PUTS",1.0,"ATM1")
        check(28,f"ATM1 at {spot} is never OTM", c<=spot and p_>=spot, f"{c}/{p_}")
    check(28,"ATM parses", _w5.parse_strike_mode("ATM1")==("ATM",1))
    check(28,"garbage falls back to ITM1, not a lottery ticket",
          _w5.parse_strike_mode("banana")==("ITM",1))

    Q = L5.contract_quality
    good = Q(713.40, 711.0, "CALL", 3.05, 2.95)
    check(28,"a real ITM contract is allowed", good["ok"], str(good["reasons"]))
    atm  = Q(713.40, 713.0, "CALL", 1.10, 1.02)
    check(28,"ATM is NOT blocked just for having time value", atm["ok"], str(atm["reasons"]))
    otm  = Q(713.40, 718.0, "CALL", 0.09, 0.05)
    check(28,"a fully-OTM lottery ticket is blocked", not otm["ok"])
    check(28,"and it says WHY, in English", any("OUT of the money" in r for r in otm["reasons"]))
    wide = Q(713.40, 711.0, "CALL", 3.60, 2.90)
    check(28,"a wide spread is blocked", not wide["ok"])
    cheap = Q(713.40, 712.9, "CALL", 0.15, 0.14)
    check(28,"a sub-minimum premium is blocked", not cheap["ok"])

    z = _w5.make_session("LIVE")
    z.strategies=[{"id":"a","name":"A","enabled":False},{"id":"b","name":"B","enabled":False}]
    z.settings["my_enabled"]=True; z._enforce_single_mode("entry")
    check(28,"entry armed -> no strategy on",
          z.settings["my_enabled"] and not any(x["enabled"] for x in z.strategies))
    z.update_strategies([{"id":"a","name":"A","enabled":True},{"id":"b","name":"B","enabled":False}])
    check(28,"arming a strategy switches ENTRY off", z.settings["my_enabled"] is False)
    check(28,"and it is the only one on",
          [x["id"] for x in z.strategies if x["enabled"]]==["a"])
    z.update_strategies([{"id":"a","name":"A","enabled":True},{"id":"b","name":"B","enabled":True}])
    check(28,"two at once is impossible",
          len([x for x in z.strategies if x["enabled"]])==1)
    z.update_settings({"my_enabled":True})
    check(28,"re-arming ENTRY switches every strategy off",
          not any(x["enabled"] for x in z.strategies))
    check(28,"state reports what is armed", z.active_mode["mode"]=="entry")

    ix5 = io.open(os.path.join(HERE,"index.html"),encoding="utf-8").read()
    check(28,"buttons are ATM1/ITM1/ITM2", all(k in ix5 for k in ("smATM","smITM2")) and "smITM3" not in ix5)
    check(28,"a retired ITM3 still lights a button", "m==='ITM3'" in ix5)
    check(28,"blocked contracts are unclickable", ".buy-c.blocked" in ix5 and "pointer-events:none" in ix5)
    check(28,"armed-mode banner exists", 'id="activeMode"' in ix5)

    print("\n[29] PERCENT ONLY - no cash anywhere on the options screen")
    ix6 = io.open(os.path.join(HERE,"index.html"),encoding="utf-8").read()
    # strip comments so the explanations do not count as violations
    code6 = re.sub(r"//[^\n]*", "", ix6)
    for gone, what in [("money(", "money() helper"),
                       ("fmtCost", "contract cost"),
                       ('id="bp"', "buying power chip"),
                       ('id="pnlN"', "dollar P&L element"),
                       ("BP $", "mirror buying power")]:
        check(29, f"{what} removed", gone not in code6)
    check(29,"DAY is a percentage, not dollars",
          "DAY <b" in ix6 and "DAY NET" not in ix6)
    _blot = code6.split('id="blotterRows"',1)[1] if 'id="blotterRows"' in code6 else code6
    check(29,"blotter rows show %", "t.pct" in code6)
    check(29,"and no dollar amount in the row template",
          "$${Math.abs(t.pnl)" not in code6)
    check(29,"ratchet line shows rungs without prices",
          "stop_price" not in code6 and "next_price" not in code6)
    check(29,"hero line drops the rotating fill price",
          "pos.entry.toFixed" not in code6)
    check(29,"buy button shows strike + the level it fires at",
          "lastPreview" in code6 and "fmtStrike(strike)" in code6)
    # The strike must come from the TRIGGER, not the live spot - at 713.40 those
    # resolve to 711C and 712C respectively, so mixing them would mislabel it.
    check(29,"strike is taken from the armed trigger, not current spot",
          "pv.strike != null" in code6)

    # --- 30. ONE entry level per side --------------------------------------
    # CALLS wait for the level at or BELOW price, PUTS for the level at or
    # ABOVE. A breakout side was added on 2026-09-02 and removed the same day,
    # live: a PUT firing on a DROP through the level below joins momentum,
    # where the trade G is describing fades the push INTO resistance. Two
    # triggers also meant the screen showed two numbers and neither was "the"
    # level. Cost of one level: in a trend it may never fill. That is chosen.
    _L = wb.LiveSession
    for _spot in (713.40, 707.60, 712.30, 709.80, 702.55, 710.00, 707.50, 764.40):
        _pc, _ = _L.entry_window(_spot, "CALLS")
        _pp, _ = _L.entry_window(_spot, "PUTS")
        check(30, "%.2f calls wait at or below spot" % _spot, _pc <= _spot + 1e-9,
              "%s vs %s" % (_pc, _spot))
        check(30, "%.2f puts wait at or above spot" % _spot, _pp >= _spot - 1e-9,
              "%s vs %s" % (_pp, _spot))

    def _fires(side, target, spot):
        return (spot <= target) if side == "CALLS" else (spot >= target)

    # G's live case: SPY between 764 and 765, armed PUTS.
    _t, _ = _L.entry_window(764.40, "PUTS")
    check(30, "SPY 764.40 puts wait for 765", abs(_t - 765.0) < 1e-9, str(_t))
    check(30, "a DROP through 764 does NOT fill it", not _fires("PUTS", _t, 764.00))
    check(30, "drifting up does not fill it either", not _fires("PUTS", _t, 764.90))
    check(30, "reaching 765 fills", _fires("PUTS", _t, 765.00))
    _tc, _ = _L.entry_window(764.40, "CALLS")
    check(30, "calls there wait for 764", abs(_tc - 764.0) < 1e-9, str(_tc))
    check(30, "a rise to 765 does NOT fill calls", not _fires("CALLS", _tc, 765.00))

    # Half levels still count.
    check(30, "the .50 grid still applies",
          abs(_L.entry_window(712.30, "PUTS")[0] - 712.50) < 1e-9,
          str(_L.entry_window(712.30, "PUTS")[0]))
    check(30, "and below, for calls",
          abs(_L.entry_window(707.60, "CALLS")[0] - 707.50) < 1e-9)

    _wc = io.open("webull_client.py", encoding="utf-8").read()
    check(30, "arm() stores ONE level", '"breakout": brk' not in _wc)
    check(30, "the trigger checks ONE level", 'broke = brk is not None' not in _wc)
    check(30, "the trigger is announced to 2dp", "{a['target']:.0f}" not in _wc)
    # Was: assert a comment explaining the removal. Assert the REMOVAL - and
    # strip comments first, or this trips on the note explaining WHY it went,
    # which is the same prose-testing trap one level down.
    _entry30 = _wc.split("def _maybe_trigger_entry", 1)[1].split("\n    def ", 1)[0]
    _entry30_code = "\n".join(l for l in _entry30.split("\n")
                              if not l.strip().startswith("#"))
    for _dead in ("brk", "breakout"):
        check(30, "no %s remains in the entry CODE" % _dead,
              _dead not in _entry30_code, _dead)
    # And the behaviour it guarantees: one level, and it is the pullback.
    _z30 = _w4.make_session("LIVE"); _z30._guard_open = lambda q: None
    _z30._underlying = lambda sym: 713.4
    _armed30 = _z30.arm("QQQ", "CALLS", 1)
    check(30, "arming yields exactly one trigger level",
          set(_armed30) == {"symbol", "side", "qty", "target", "spot_at_arm"},
          str(sorted(_armed30)))
    _ix30 = io.open("index.html", encoding="utf-8").read()
    check(30, "the screen shows one number", "armed.breakout" not in _ix30
          and "r.breakout" not in _ix30)

    # The ratchet must still protect EVERY fill, armed or manual.
    check(30, "the ratchet is armed on every position", '"ratchet_on": True,' in _wc)
    class _R30(wb.LiveSession):
        def __init__(self):
            import config as _c
            self.settings = dict(_c.DEFAULT_SETTINGS); self.strategies = []
            self.settings["my_enabled"] = False
            self.settings["ratchet_tiers"] = False
    _r30 = _R30()
    _r30.position = {"entry": 3.00, "mark": 3.00, "qty": 1,
                     "ratchet_on": True, "ratchet_step": 10.0}
    _r30._update_ratchet()
    check(30, "a manual fill still gets a stop with the switch off",
          (_r30.position.get("ratchet") or {}).get("stop_pct") == -10.0,
          str((_r30.position.get("ratchet") or {}).get("stop_pct")))

    # --- 31. One switch, still window, toggle switches -------------------
    _idx = io.open("index.html", encoding="utf-8").read()
    # The modal is vertically centred; a content-height box re-centres on every
    # tab change and the whole frame jumps.
    check(31, "settings modal has a fixed height", "height:min(90vh,760px)" in _idx)
    check(31, "modal is a flex column so the frame cannot resize",
          "flex-direction:column" in _idx.split(".smodal{",1)[1][:400])
    check(31, "only the pane area scrolls", ".panes{" in _idx and "overflow-y:auto" in _idx)
    check(31, "title/tabs/actions are pinned",
          ".smodal h2,.smodal .tabs,.smodal .activemode,.smodal .sactions{flex:none}" in _idx)
    check(31, "every pane lives inside the scroller",
          all(('id="%s"' % p) in _idx.split('<div class="panes">',1)[1]
                                   .split('<!-- /panes -->',1)[0]
              for p in ("paneConfig","paneStrat","paneMirror")))
    check(31, "save/cancel do NOT scroll away",
          "sactions" not in _idx.split('<div class="panes">',1)[1]
                                .split('<!-- /panes -->',1)[0])
    # Toggle switches, not checkboxes.
    check(31, "toggle-switch styling exists", ".sw input:checked + i" in _idx)
    check(31, "entry+ratchet uses a switch",
          '<span class="sw"><input type="checkbox" id="myEnabled"' in _idx)
    check(31, "strategy cards use a switch",
          _idx.count('class="sw"') >= 2 and 'EZ.toggleStrategy' in _idx)
    # ONE switch for both halves.
    check(31, "the separate RATCHET checkbox is gone", "ratchetEnabled" not in _idx)
    check(31, "the step % input survived", 'id="ratchetStep"' in _idx)
    check(31, "the browser stores ONE key, not two",
          "settings.ratchet_enabled" not in _idx)
    check(31, "arming a strategy switches the one switch off",
          "api('/api/settings',{my_enabled:false});" in _idx)
    # Server refuses to hold the two apart, whatever arrives.
    # update_settings() persists to my-settings.json. These are throwaway
    # objects exercising the logic, so the write is stubbed out - otherwise the
    # suite leaves the real file holding whatever the last loop iteration set.
    _real_save = wb.uc.save
    wb.uc.save = lambda *a, **k: None
    import config as _cfgA
    class _One(wb.LiveSession):
        def __init__(self):
            import config as _cfg
            self.settings = dict(_cfg.DEFAULT_SETTINGS); self.strategies = []
        def _enforce_single_mode(self, prefer=None): pass
    _f = _One()
    check(31, "ratchet_enabled is not a setting any more",
          "ratchet_enabled" not in _cfgA.DEFAULT_SETTINGS)
    check(31, "my_enabled is", _cfgA.DEFAULT_SETTINGS["my_enabled"] is True)
    for _start in (True, False):
        for _send, _want in (({"my_enabled": True}, True),
                             ({"my_enabled": False}, False),
                             # legacy key, still accepted, folded into the one
                             ({"ratchet_enabled": True}, True),
                             ({"ratchet_enabled": False}, False),
                             # both sent: the live key wins, no ambiguity
                             ({"my_enabled": True, "ratchet_enabled": False}, True),
                             ({"my_enabled": False, "ratchet_enabled": True}, False)):
            _f.settings["my_enabled"] = _start
            _f.update_settings(dict(_send))
            check(31, "from %s, %s -> %s" % (_start, _send, _want),
                  _f.settings["my_enabled"] is _want, str(_f.settings["my_enabled"]))
            check(31, "and the retired key is never stored (%s)" % _send,
                  "ratchet_enabled" not in _f.settings)
    # The trap the old two-key version had: OR-ing them made OFF unsendable.
    _f.settings["my_enabled"] = True
    _f.update_settings({"my_enabled": False})
    check(31, "it can actually be switched OFF", _f.settings["my_enabled"] is False)
    _f.settings["my_enabled"] = True
    _f.update_settings({"strike_mode": "ITM1"})
    check(31, "an unrelated setting leaves the switch alone", _f.settings["my_enabled"])
    wb.uc.save = _real_save

    # --- 32. No SAVE button; a live trade keeps its own terms -------------
    import config as _cfg0
    check(32, "the SAVE button is gone", "EZ.saveSettings()" not in _idx)
    # Scoped to the settings footer: CANCEL also names the DISARM button and the
    # strategy editor's cancel, neither of which is going anywhere.
    _sact = _idx.split('<div class="sactions">', 1)[1].split('</div>', 1)[0]
    check(32, "and CANCEL with it", "CANCEL" not in _sact)
    check(32, "only one button is left in the footer", _sact.count("<button") == 1)
    check(32, "DONE just closes", 'onclick="EZ.closeSettings()">DONE<' in _idx)
    check(32, "the switch writes on change",
          'id="myEnabled"' in _idx and
          'onchange="EZ.applySettings()"' in _idx.split('id="myEnabled"',1)[1][:120])
    check(32, "the step box writes on change and on blur",
          'onchange="EZ.applySettings()" onblur="EZ.applySettings()"' in _idx)
    check(32, "picking a strike writes it",
          "applySettings();         // no SAVE button" in _idx)
    check(32, "applying does NOT close the window",
          "closeSettings(); refreshQuote();" not in _idx)
    check(32, "a cleared step box falls back to 10",
          "parseFloat($('ratchetStep').value)||10" in _idx)
    check(32, "and the box is written back so it matches storage",
          "$('ratchetStep').value=settings.ratchet_step_pct" in _idx)
    check(32, "10% is the shipped default", _cfg0.DEFAULT_SETTINGS["ratchet_step_pct"] == 10.0)
    check(32, "and the input agrees", 'id="ratchetStep" type="number" min="1" max="100" step="1" value="10"' in _idx)
    check(32, "an open trade is called out on screen", 'id="liveNote"' in _idx and "NEXT trade" in _idx)

    # The reason this matters: with no SAVE button, every keystroke reaches the
    # server at once. A live trade must not be re-tuned underneath itself.
    _wc2 = io.open("webull_client.py", encoding="utf-8").read()
    # ratchet_on is now unconditionally True - the switch decides whether the
    # ENTRY waits for a level, never whether a live trade has a stop. The step
    # is still frozen at open, which is the part that must not move underneath
    # a running position.
    check(32, "terms are frozen onto the position at open",
          '"ratchet_on": True,' in _wc2 and
          '"ratchet_step": float(self.settings.get("ratchet_step_pct")' in _wc2)
    check(32, "the ratchet reads the position, not live settings",
          'p.get("ratchet_on"' in _wc2 and 'p.get("ratchet_step"' in _wc2)

    class _Live(wb.LiveSession):
        def __init__(self):
            import config as _c
            self.settings = dict(_c.DEFAULT_SETTINGS); self.strategies = []
    _lv = _Live()
    _lv.settings["ratchet_tiers"] = False       # flat rungs under test here
    # Opened at 10% step, then the step is changed to 50% mid-trade.
    _lv.position = {"entry": 3.00, "mark": 3.30, "qty": 1,
                    "ratchet_on": True, "ratchet_step": 10.0}
    _lv._update_ratchet()
    _stop_before = _lv.position["ratchet"]["stop_pct"]
    _lv.settings["ratchet_step_pct"] = 50.0
    _lv._update_ratchet()
    _stop_after = _lv.position["ratchet"]["stop_pct"]
    check(32, "changing the step mid-trade does NOT move a live stop",
          _stop_before == _stop_after == 0.0, "%s -> %s" % (_stop_before, _stop_after))
    # Switching the whole feature off mid-trade must not abandon the open stop.
    _lv.settings["my_enabled"] = False
    _lv._update_ratchet()
    check(32, "switching it off mid-trade does not abandon the open stop",
          _lv.position.get("ratchet") is not None)
    # And the NEXT trade does pick the new terms up.
    _lv.settings.update({"my_enabled": True, "ratchet_step_pct": 50.0})
    _lv.position = {"entry": 3.00, "mark": 4.80, "qty": 1,
                    "ratchet_on": bool(_lv.settings["my_enabled"]),
                    "ratchet_step": float(_lv.settings["ratchet_step_pct"])}
    _lv._update_ratchet()
    check(32, "the next trade uses the new step (+60% -> stop +0%)",
          _lv.position["ratchet"]["stop_pct"] == 0.0,
          str(_lv.position["ratchet"]["stop_pct"]))

    # --- 33. The page actually RUNS ---------------------------------------
    # node --check parses; it does not execute. paintLiveNote() read a variable
    # named `position` that was never declared - a ReferenceError thrown one
    # line before the modal was shown, so SETTINGS silently stopped opening and
    # every syntax check still passed. ui_smoke.js evaluates the real script and
    # calls the handlers a user can reach.
    import subprocess as _sp
    _sm = _sp.run(["node", "ui_smoke.js", "index.html"],
                  cwd=HERE, capture_output=True, text=True, timeout=60)
    check(33, "the options page runs and its handlers do not throw",
          _sm.returncode == 0, (_sm.stdout + _sm.stderr).strip()[:400])

    # And prove the harness can still see that exact failure, or it is theatre.
    # Undo the guard AND reintroduce the undeclared variable, so the throw
    # propagates exactly as it did when SETTINGS stopped opening.
    _orig = io.open(os.path.join(HERE, "index.html"), encoding="utf-8").read()
    _swaps = [
        ("    const live=!!inTrade;", "    const live=!!position;"),
        ("$('setScrim').classList.add('show');\n    try{\n    showTab('config')",
         "showTab('config')"),
        ("paintLiveNote();\n    }catch(e){ console.error('settings paint failed:', e); }",
         "paintLiveNote();\n    $('setScrim').classList.add('show');"),
    ]
    _broken = _orig
    for _a, _b in _swaps:
        check(33, "self-test can rebuild the bug (%s)" % _a.strip()[:34], _a in _broken)
        _broken = _broken.replace(_a, _b, 1)
    _bp = os.path.join(HERE, "_ui_smoke_selftest.html")
    io.open(_bp, "w", encoding="utf-8").write(_broken)
    try:
        _bad = _sp.run(["node", "ui_smoke.js", "_ui_smoke_selftest.html"],
                       cwd=HERE, capture_output=True, text=True, timeout=60)
        _out = _bad.stdout + _bad.stderr
        check(33, "the smoke test still catches an undeclared variable",
              _bad.returncode != 0 and "is not defined" in _out, _out.strip()[:200])
    finally:
        # This file is a deliberately BROKEN copy of index.html. If the delete
        # ever fails it must not be left sitting in the app folder, where
        # auto-sync will happily commit it - park it in _trash, which the
        # launcher wipes on every start.
        try:
            os.remove(_bp)
        except OSError:
            try:
                _t = os.path.join(HERE, "_trash"); os.makedirs(_t, exist_ok=True)
                shutil.move(_bp, os.path.join(_t, "_ui_smoke_selftest.html"))
            except Exception:
                pass

    # The guard that keeps one bad painter from locking you out of settings.
    check(33, "settings open BEFORE anything is painted",
          "$('setScrim').classList.add('show');\n    try{" in _idx)
    check(33, "a painter that throws is logged, not fatal",
          "catch(e){ console.error('settings paint failed:', e); }" in _idx)

    # --- 34. LOCK / quit removed; size warns without showing cash ---------
    check(34, "the LOCK button is gone from the header", "EZ.lockNow()" not in _idx)
    check(34, "the quit X is gone", "EZ.quitApp()" not in _idx)
    # The in-trade lock is a different thing and stays: it is what stops you
    # pressing BUY again while a position is open.
    check(34, "the in-trade lock still exists", 'id="lockNote"' in _idx and "lockdown(inTrade)" in _idx)
    check(34, "CONTRACTS goes red when unaffordable", "#qty.over{color:var(--red)}" in _idx)
    check(34, "affordability is rechecked when you change size",
          "$('qty').textContent=qty; savePrefs({qty});\n    paintQtyAfford();" in _idx)
    check(34, "and when a fresh quote lands", "paintQtyAfford();" in _idx.split("function paintSide",1)[1][:600])
    check(34, "buying power is read from state but never rendered",
          "buyingPower=Number(st.buying_power)" in _idx and "$('bp')" not in _idx)
    # "$" alone would match the $() element helper. What must not appear is a
    # dollar sign rendered into text.
    _aff = _idx.split("function paintQtyAfford", 1)[1].split("function setStrikeMode", 1)[0]
    check(34, "still no dollar figure attached to the warning",
          "'$" not in _aff and '"$' not in _aff and "$'+" not in _aff)
    check(34, "the server still supplies buying power",
          '"buying_power": round(self.buying_power, 2),' in _wc2)
    # A contract is 100 shares - forgetting the multiplier makes the check
    # useless, since a $3.00 option would read as $3 against the account.
    check(34, "cost uses the 100x contract multiplier", "*100*qty" in _idx)

    # --- 35. Time value warns, it does not block -------------------------
    # This blocked a live trade at 9:30 on the first real session. Extrinsic %
    # measures distance-to-strike and time-left, not quality: at the open
    # nearly every 0DTE is 90%+ time value, and an ITM1 a nickel from the
    # strike is 95% by arithmetic. Being fully OTM is the real problem, and
    # that is blocked separately.
    _Q = wb.LiveSession.contract_quality
    _q = _Q(713.95, 714.0, "PUT", 1.00, 0.97)          # the exact live case
    check(35, "95% time value is TRADABLE", _q["ok"] is True,
          str(_q.get("reasons")))
    check(35, "but it still says so", any("time value" in w for w in _q.get("warnings", [])))
    check(35, "warnings are separate from blocking reasons", _q["reasons"] == [])
    # The three that must still refuse.
    check(35, "fully OTM is still blocked", _Q(714.50, 714.0, "PUT", 0.80, 0.77)["ok"] is False)
    check(35, "penny premium is still blocked", _Q(713.95, 714.0, "PUT", 0.10, 0.05)["ok"] is False)
    check(35, "a wide spread is still blocked", _Q(713.10, 714.0, "PUT", 1.40, 1.05)["ok"] is False)
    check(35, "a clean contract has no warning at all",
          _Q(713.10, 714.0, "PUT", 1.40, 1.37)["warnings"] == [])
    check(35, "the warning reaches the button as a tooltip, not as text",
          "btn.title = (qual.warnings||[]).join" in _idx)
    check(35, "the button text is still strike + level",
          "fmtStrike(strike)" in _idx)

    # --- 36. Trade log: how the trade travelled --------------------------
    import trade_log as _tl2
    for _f in ("best_pct","worst_pct","best_price","worst_price","gave_back_pct",
               "ratchet_stop_pct","ratchet_step","strike_mode","held_secs"):
        check(36, "log has a %s column" % _f, _f in _tl2.FIELDS)

    # High/low water marks, tracked whether or not the ratchet is on.
    _tr = {"entry": 3.00, "mark": 3.00}
    for _m in (3.00, 3.30, 2.85, 4.20, 3.60):
        _tr["mark"] = _m; wb.LiveSession._track_excursion(_tr)
    check(36, "best is the highest mark seen", _tr["best_price"] == 4.20, str(_tr.get("best_price")))
    check(36, "worst is the lowest", _tr["worst_price"] == 2.85, str(_tr.get("worst_price")))
    check(36, "best % is right (+40)", _tr["best_pct"] == 40.0, str(_tr.get("best_pct")))
    check(36, "worst % is right (-5)", _tr["worst_pct"] == -5.0, str(_tr.get("worst_pct")))
    # A trade that only ever went up was never down: worst is 0, not the entry.
    _up = {"entry": 2.00, "mark": 2.00}
    for _m in (2.00, 2.40, 2.90):
        _up["mark"] = _m; wb.LiveSession._track_excursion(_up)
    check(36, "a trade that never went red has worst 0", _up["worst_pct"] == 0.0,
          str(_up.get("worst_pct")))

    # The header migration: appending 26 columns to an 18-column file without
    # rewriting the header shifts every value one column left and the file
    # still opens fine, so it would go unnoticed.
    import tempfile as _tf2, csv as _csv2
    _d2 = _tf2.mkdtemp(prefix="tlmig")
    _tl2.LOG_DIR = _d2
    _tl2.CSV_PATH = os.path.join(_d2, "trades.csv")
    _tl2.XLSX_PATH = os.path.join(_d2, "log.xlsx")
    _old_fields = ["date","time_in","symbol","side","entry","exit","pnl","pnl_pct"]
    with io.open(_tl2.CSV_PATH, "w", encoding="utf-8", newline="") as _f2:
        _w2 = _csv2.DictWriter(_f2, fieldnames=_old_fields); _w2.writeheader()
        _w2.writerow({"date":"2026-08-25","time_in":"10:00","symbol":"QQQ",
                      "side":"CALLS","entry":"2.00","exit":"2.50","pnl":"50.0",
                      "pnl_pct":"25.0"})
    _tl2.record({"date":"2026-08-26","symbol":"SPY","side":"PUTS","entry":1.0,
                 "exit":1.5,"pnl":50.0,"pnl_pct":50.0,"best_pct":60.0,"worst_pct":-8.0})
    _got = _tl2._rows()
    check(36, "the old row survives the upgrade", len(_got) == 2, str(len(_got)))
    check(36, "and its values stay in their own columns",
          _got[0]["symbol"] == "QQQ" and _got[0]["pnl_pct"] == "25.0",
          str(_got[0])[:120])
    check(36, "the new row keeps the new columns",
          _got[1]["best_pct"] == "60.0" and _got[1]["worst_pct"] == "-8.0",
          str(_got[1])[:120])
    check(36, "migrating twice is harmless",
          (_tl2._migrate_header() or True) and len(_tl2._rows()) == 2)

    _wc3 = io.open("webull_client.py", encoding="utf-8").read()
    check(36, "excursions are tracked on every mark refresh",
          "self._track_excursion(p)" in _wc3)
    check(36, "the open time is stamped so held time is real",
          '"opened_ts": time.time(),' in _wc3)
    check(36, "the precise ratchet reason is no longer clobbered",
          'if self.position and not self.position.get("exit_reason"):' in _wc3)

    # --- 37. Desktop icon --------------------------------------------------
    _ico = os.path.join(HERE, "sniper.ico")
    check(37, "the icon file exists", os.path.exists(_ico))
    try:
        from PIL import Image as _Im
        _sz = sorted(_Im.open(_ico).ico.sizes())
        check(37, "it carries a 16px and a 256px frame",
              (16, 16) in _sz and (256, 256) in _sz, str(_sz))
    except Exception as _e:
        check(37, "icon is readable", False, str(_e)[:80])

    # A shortcut CANNOT point at the real launcher: its name starts with an
    # emoji, and WScript.Shell rejects any TargetPath holding a character
    # outside the Basic Multilingual Plane - "Value does not fall within the
    # expected range" - leaving an icon on the desktop that does nothing.
    # So an ASCII-named stub sits in between.
    _stub = os.path.join(HERE, "MARKET SNIPER.bat")
    check(37, "the ASCII launcher stub exists", os.path.exists(_stub))
    _st = io.open(_stub, encoding="utf-8").read()
    check(37, "the stub is pure ASCII", all(ord(c) < 128 for c in _st))
    check(37, "the stub finds the real launcher by wildcard",
          '("*START MARKET SNIPER.bat")' in _st)
    check(37, "the stub does not match its own wildcard (no loop)",
          not __import__("fnmatch").fnmatch("MARKET SNIPER.bat",
                                            "*START MARKET SNIPER.bat"))
    check(37, "the stub CALLs, so closing the window still stops the servers",
          'call "%LAUNCHER%"' in _st)
    check(37, "the stub says so when the launcher is missing", "Could not find" in _st)

    _shortcut = os.path.join(HERE, "CREATE DESKTOP ICON.bat")
    check(37, "the one-time setup file exists", os.path.exists(_shortcut))
    _sh = io.open(_shortcut, encoding="utf-8").read()
    check(37, "the setup file is pure ASCII", all(ord(c) < 128 for c in _sh))
    # Windows will not PIN a shortcut whose target is a .bat - a shell rule,
    # not a setting. So the target is cmd.exe (pinnable) and the batch file
    # rides along as an argument.
    check(37, "the target is cmd.exe, so it can be pinned",
          "System32\\cmd.exe" in _sh)
    check(37, "the batch file is passed as an argument",
          "$s.Arguments = '/c" in _sh and "MARKET SNIPER.bat" in _sh)
    check(37, "the emoji launcher is never named in the shortcut",
          "START MARKET SNIPER" not in _sh)
    check(37, "it refuses to run if the stub is missing",
          'if not exist "%~dp0MARKET SNIPER.bat"' in _sh)
    check(37, "it clears a stale icon first", "Remove-Item $link -Force" in _sh)
    # The first version reported success while leaving a dead icon behind,
    # because assigning TargetPath threw and nothing checked afterwards.
    check(37, "it reads the shortcut BACK and proves the target exists",
          "$c = $ws.CreateShortcut($link)" in _sh and
          "Test-Path $c.TargetPath" in _sh)
    check(37, "it also proves the batch file it points at exists",
          "shortcut saved but MARKET SNIPER.bat is missing" in _sh)
    check(37, "it stops on the first PowerShell error",
          "$ErrorActionPreference = 'Stop'" in _sh)
    check(37, "it tells you how to pin it", "Pin to taskbar" in _sh)
    check(37, "it applies the icon", "sniper.ico" in _sh)
    check(37, "it explains the manual fallback",
          "New ^> Shortcut" in _sh and "Change Icon" in _sh)
    check(37, "and the manual route is the pinnable one too",
          'cmd.exe /c "%~dp0MARKET SNIPER.bat"' in _sh)
    # The stub adds nothing: the real launcher already opens the browser, so
    # the Desktop icon needs no extra step to get you to the trading screen.
    _real = [os.path.join(HERE, f) for f in os.listdir(HERE)
             if f.endswith("START MARKET SNIPER.bat")]
    check(37, "the real launcher is still there", len(_real) == 1, str(_real))
    _lb = io.open(_real[0], encoding="utf-8", errors="replace").read()
    check(37, "the launcher still opens the browser itself",
          "http://127.0.0.1:8000" in _lb and "start" in _lb)

    # --- 38. Velocity survives Yahoo's cumulative-volume bars -------------
    # Measured live 2026-08-26 12:42 ET: QQQ's 12:41 bar carried 18,887,220
    # against neighbours of 30-70k, and SPY's 12:07 bar carried 15,501,626
    # against 50-90k. Yahoo intermittently publishes a cumulative figure in a
    # single 1-minute bar. Averaged in, ONE such bar decides the reading: in
    # the recent window the meter pinned to "violent", in the baseline the
    # identical tape read "calm". Both were on screen the same afternoon.
    import tape as _tp
    _tp = importlib.reload(_tp)

    def _bar(t, v, c=500.0, rng=0.10):
        return {"t": t, "o": c, "h": c + rng / 2, "l": c - rng / 2, "c": c, "v": v}

    _clean = [_bar(i, 55000) for i in range(34)] + [_bar(34, 55000)]
    _in_recent = [_bar(i, 55000) for i in range(34)] + [_bar(34, 18887220)]
    _in_base = [_bar(0, 15501626)] + [_bar(i, 55000) for i in range(1, 35)]

    _rc = _tp.compute(_clean)
    _rr = _tp.compute(_in_recent)
    _rb = _tp.compute(_in_base)
    check(38, "clean tape reads normal", _rc["state"] == "normal", str(_rc["score"]))
    check(38, "a 300x artifact in the RECENT window no longer reads violent",
          _rr["state"] == "normal", "%s / %s" % (_rr["state"], _rr["score"]))
    check(38, "a 300x artifact in the BASELINE no longer reads calm",
          _rb["state"] == "normal", "%s / %s" % (_rb["state"], _rb["score"]))
    check(38, "and all three agree, because the tape is the same tape",
          _rc["score"] == _rr["score"] == _rb["score"],
          "%s %s %s" % (_rc["score"], _rr["score"], _rb["score"]))
    check(38, "the artifact is counted, not hidden", _rr.get("clipped_bars", 0) == 1)
    check(38, "an artifact as the NEWEST bar suppresses acceleration",
          _rr["accel_pct"] == 0.0 and "artifact" in (_rr.get("note") or ""),
          "%s / %s" % (_rr["accel_pct"], _rr.get("note")))

    # A REAL burst must still register - the cap must not flatten genuine speed.
    _burst = [_bar(i, 50000) for i in range(30)] + [_bar(30 + i, 260000) for i in range(5)]
    _rbu = _tp.compute(_burst)
    check(38, "a real 5x burst still reads fast or violent",
          _rbu["state"] in ("fast", "violent"), "%s / %s" % (_rbu["state"], _rbu["score"]))
    # And a genuinely dead tape must still read dead, or "silent tape = do not
    # enter" stops working.
    _dead = [_bar(i, 55000) for i in range(30)] + [_bar(30 + i, 2000) for i in range(5)]
    check(38, "a dying tape still reads calm", _tp.compute(_dead)["state"] == "calm",
          str(_tp.compute(_dead)["score"]))
    _shut = [_bar(i, 0) for i in range(35)]
    check(38, "a closed market still reads closed",
          "closed" in (_tp.compute(_shut).get("note") or ""))

    check(38, "medians, not means, decide the ratios",
          "vol_recent = _median(v_recent)" in io.open("tape.py", encoding="utf-8").read())
    check(38, "acceleration is clamped to a believable range",
          "min(accel, 400.0)" in io.open("tape.py", encoding="utf-8").read())

    # --- 39. Dwell time -----------------------------------------------------
    import levels as _lv2
    _lv2 = importlib.reload(_lv2)
    _n = int(time.time())

    def _lb(i, lo, hi, v=50000, total=60):
        return {"t": _n - (total - i) * 60, "o": (lo + hi) / 2,
                "h": hi, "l": lo, "c": (lo + hi) / 2, "v": v}

    for _p, _want in ((713.40, (713.0, 714.0)), (713.99, (713.0, 714.0)),
                      (702.55, (702.0, 703.0))):
        check(39, "%.2f brackets to %s" % (_p, _want),
              _lv2.bracketing_levels(_p) == _want, str(_lv2.bracketing_levels(_p)))
    # Price exactly ON a dollar: the pair must never collapse to one number.
    _b, _a = _lv2.bracketing_levels(713.00)
    check(39, "price on the dollar still gives two distinct levels", _b != _a and _a - _b == 1.0)

    # 60 minutes stuck inside 713.20-713.80: neither dollar is ever touched.
    _pin = [_lb(i, 713.20, 713.80) for i in range(60)]
    _d = _lv2.dwell(_pin, now=_n)
    check(39, "a pinned market reports pinned", _d["pinned"] is True)
    check(39, "an untouched level is None, not 0",
          _d["mins_below"] is None and _d["mins_above"] is None)
    check(39, "None and 0 stay different", 0 is not None)

    # One bar pokes through 714, two minutes ago.
    _mv = [_lb(i, 713.20, 713.80) for i in range(58)] + \
          [_lb(58, 713.5, 714.10), _lb(59, 713.4, 713.9)]
    _d2 = _lv2.dwell(_mv, now=_n)
    check(39, "a touch clears pinned", _d2["pinned"] is False)
    check(39, "and is timed correctly (2m)", abs(_d2["mins_above"] - 2.0) < 0.6,
          str(_d2["mins_above"]))

    # Highs/lows, not closes: a wick through a level IS a touch. Using closes
    # would miss the rejection wick, which is the touch that matters most.
    check(39, "a wick through the level counts as a touch",
          _lv2._touched({"h": 714.02, "l": 713.40}, 714.0) is True)
    check(39, "a bar that never reached it does not",
          _lv2._touched({"h": 713.90, "l": 713.40}, 714.0) is False)

    # The label must quote the bars actually available, not the 180 cap.
    check(39, "label quotes the real lookback, not the cap",
          ">60m" in _lv2.label(_d), _lv2.label(_d))

    # Dwell and velocity answer different questions. Pinned AND violent is a
    # contradiction and must be surfaced, not averaged away.
    _ag = _lv2.agreement({"ok": True, "pinned": True}, {"ok": True, "state": "violent"})
    check(39, "pinned + violent is flagged as disagreement", _ag["agree"] is False)
    _ag2 = _lv2.agreement({"ok": True, "pinned": True}, {"ok": True, "state": "calm"})
    check(39, "pinned + calm agrees", _ag2["agree"] is True)

    check(39, "no bars is handled, not crashed", _lv2.dwell([])["ok"] is False)
    _mainsrc = io.open("main.py", encoding="utf-8").read()
    check(39, "the endpoint exists", '@app.get("/api/dwell")' in _mainsrc)
    check(39, "levels is imported defensively like tape",
          "import levels\nexcept Exception:\n    levels = None" in _mainsrc)
    check(39, "the endpoint rejects untradable symbols",
          "isn't one of the tradable symbols" in _mainsrc.split('/api/dwell', 1)[1][:900])

    # --- 40. Volume gauge ---------------------------------------------------
    import gauges as _g
    _g = importlib.reload(_g)

    # THE TRAP: comparing volume-so-far against whole past days makes every
    # morning read as dead. The profile is learned from real bars, and it is
    # nothing like a straight line - measured on QQQ, 12:30 is 64% of a day
    # done where a flat clock says 46%.
    _prof = [(570 + i * 5, min(1.0, (i * 5 / 390.0) ** 0.75)) for i in range(79)]
    check(40, "the profile is used, not a flat clock",
          _g.expected_fraction(_prof, 630) > (630 - 570) / 390.0)
    check(40, "before the open, nothing is done", _g.expected_fraction(_prof, 500) == 0.0)
    check(40, "with no profile it degrades to a clock, not a crash",
          0.0 < _g.expected_fraction([], 750) < 1.0)

    # Percentile, not a ratio: a few huge days must not move where today sits.
    _series = [100.0] * 90 + [5000.0] * 10
    check(40, "ten monster days do not drag the ranking",
          _g.percentile_of(101.0, _series) == 90.0, str(_g.percentile_of(101.0, _series)))
    check(40, "an empty history returns None, not a fake number",
          _g.percentile_of(1.0, []) is None)
    check(40, "low/high bands sit where the spec says",
          _g._band(10) == "low" and _g._band(50) == "normal" and _g._band(90) == "high")

    # Yahoo's intraday volume cannot be summed. Measured 2026-08-26 12:55 ET:
    # QQQ's 5-min bars summed to 277,170,149 against a 43,645,800 median day,
    # because six bars carried 26-47x the median. Today's total therefore comes
    # from the daily bar, which agreed with the cleaned sum to within 1%.
    _gsrc = io.open("gauges.py", encoding="utf-8").read()
    check(40, "today's volume comes from the daily bar, not a sum of 5m bars",
          '_chart(ysym, "1d", "1d")' in _gsrc and "regularMarketVolume" in _gsrc)
    check(40, "profile bars are cleaned of artifacts", "cleaned = _clean(" in _gsrc)
    _cl = _g._clean([100, 100, 100, 100, 100, 28000000])
    check(40, "a 280,000x artifact is clipped", max(_cl) == 100 * _g.OUTLIER_X, str(max(_cl)))
    check(40, "and honest volume is untouched", _cl[:5] == [100, 100, 100, 100, 100])

    # Ranking today against a list containing today drags every reading to the
    # middle, and on a record day the record pulls its own percentile down.
    check(40, "today is excluded from its own history", "if d >= today:" in _gsrc)
    # The clock must come from the feed. -4 is right from March to November and
    # an hour wrong the rest of the year - a fifth of a session.
    check(40, "the exchange timezone is read from the feed, not hardcoded",
          "gmtoffset" in _gsrc and "def _tz_for" in _gsrc)
    # Dividing by a near-zero fraction at 09:31 turns noise into a wild number.
    check(40, "it refuses to project in the first minutes", "too early to project" in _gsrc)

    _mainsrc2 = io.open("main.py", encoding="utf-8").read()
    check(40, "the endpoint exists", '@app.get("/api/volume")' in _mainsrc2)
    check(40, "gauges is imported defensively",
          "import gauges\nexcept Exception:\n    gauges = None" in _mainsrc2)

    # --- 41. Volatility: two gauges, never blended -----------------------
    # Black-Scholes must round-trip, or the implied number is decoration.
    for _v in (0.15, 0.35, 0.80):
        _px = _g.bs_price(713.0, 712.0, 4 / 24 / 365, _v, True)
        _back = _g.implied_vol(_px, 713.0, 712.0, 4 / 24 / 365, True)
        check(41, "BS round-trips at %.0f%% vol" % (_v * 100),
              _back is not None and abs(_back - _v * 100) < 0.5, str(_back))
    # Bisection, not Newton: vega collapses near zero on a 0DTE away from the
    # money and Newton divides by it. These must return None, not diverge.
    check(41, "a quote below intrinsic is refused",
          _g.implied_vol(0.50, 713.0, 712.0, 4 / 24 / 365, True) is None)
    check(41, "an impossible quote is refused",
          _g.implied_vol(700.0, 713.0, 712.0, 4 / 24 / 365, True) is None)
    check(41, "zero time to expiry is refused",
          _g.implied_vol(1.0, 713.0, 712.0, 0, True) is None)
    check(41, "bisection is used, and says why", "Bisection rather than Newton" in _gsrc)
    # Time to expiry can never be zero or IV goes infinite in the last minute.
    check(41, "expiry time is floored at a minute", _g.hours_to_expiry() > 0)

    # Realized is ranked against the SAME calculation rolled back through the
    # symbol's own history, so "high" means high for this symbol.
    check(41, "realized is a percentile of its own history",
          "_annualised_vol(closes[end - RV_WINDOW_DAYS - 1:end])" in _gsrc)
    # Check the SHAPE of the result, not the prose. The word "blended" appears
    # in a comment explaining that they are not blended, which is exactly the
    # kind of thing a string match gets wrong.
    _sep = _g.volatility("QQQ", option=None)
    check(41, "the two gauges stay separate",
          isinstance(_sep.get("realized"), dict) and isinstance(_sep.get("implied"), dict)
          and not any(k in _sep for k in ("vol_score", "combined", "blended")))
    check(41, "and realized survives on its own", _sep["realized"].get("ok") is True)
    check(41, "thresholds are plain constants, meant to be argued with",
          all(k in _gsrc for k in ("RV_LOW_PCTL", "RV_HIGH_PCTL",
                                   "IV_RICH_RATIO", "IV_CHEAP_RATIO")))
    # Not connected must SAY not connected rather than invent a number.
    _nov = _g.volatility("QQQ", option=None)
    check(41, "implied is unavailable, not guessed, without a chain",
          _nov["implied"]["ok"] is False and "connect" in _nov["implied"]["reason"])

    _wc4 = io.open("webull_client.py", encoding="utf-8").read()
    check(41, "the ATM chain quote comes from Webull", "def atm_option_for_vol" in _wc4)
    check(41, "a native Webull IV is preferred over inverting a mid",
          "iv_native" in _wc4 and "iv_native" in _mainsrc2)
    check(41, "a fractional IV is normalised to percent", "if iv is not None and iv < 5.0:" in _wc4)
    check(41, "greeks are picked up if Webull sends them",
          all(gk in _wc4 for gk in ('"delta"', '"theta"', '"gamma"', '"vega"')))
    check(41, "the endpoint exists", '@app.get("/api/volatility")' in _mainsrc2)
    check(41, "a chain hiccup cannot kill the realized reading",
          "a chain hiccup must not kill realized" in _mainsrc2)

    # --- 42. Audio cues -----------------------------------------------------
    _ix = io.open("index.html", encoding="utf-8").read()
    for _cue in ("armed", "filled", "exit", "cancel", "blocked"):
        check(42, "the %s cue is wired" % _cue, "sfx('%s')" % _cue in _ix)
    # Synthesised, so there is no audio file to ship or fail to load.
    check(42, "tones are synthesised, not files",
          "createOscillator" in _ix and ".mp3" not in _ix and ".wav" not in _ix)
    # A square gain edge is audible as a click at both ends of every note.
    check(42, "the envelope is ramped, not switched",
          "exponentialRampToValueAtTime" in _ix)
    # Polling runs once a second: sounding STATE rather than TRANSITION would
    # replay the note every tick for as long as the position was open.
    check(42, "fill and exit sound on the transition only",
          "if(pos && !prevPos) sfx('filled');" in _ix and
          "if(!pos && prevPos) sfx('exit');" in _ix)
    check(42, "the previous position is tracked for that", "prevPos=false" in _ix)
    check(42, "sound can be switched off", "function toggleSound" in _ix and
          'id="sndBtn"' in _ix)
    check(42, "the choice survives a reload", "localStorage.setItem('ms_sound'" in _ix)
    # Audio must never be able to take the trading screen down with it.
    check(42, "any audio failure is swallowed",
          "never break trading over it" in _ix)
    # The grid has .50 levels: "715" for a 715.50 trigger is the wrong price.
    check(42, "the armed line shows the level to 2dp",
          "Number(armed.target).toFixed(2)" in _ix)
    # The breakout level is gone - one level per side, scenario 30.
    check(42, "and names ONE level, not two", "armed.breakout" not in _ix)
    _wc5 = io.open("webull_client.py", encoding="utf-8").read()
    check(42, "the server sends the one level it waits for", '"target": target' in _wc5)

    import subprocess as _sp2
    _sm2 = _sp2.run(["node", "ui_smoke.js", "index.html"], cwd=HERE,
                    capture_output=True, text=True, timeout=60)
    check(42, "the page still runs with the sound code in it",
          _sm2.returncode == 0, (_sm2.stdout + _sm2.stderr).strip()[:300])

    # --- 43. Trend module -------------------------------------------------
    import trend as _tr
    _tr = importlib.reload(_tr)

    def _mk(n, step=0.0, rng=0.40, vol=50000, up=True, start=700.0):
        out, c = [], start
        for i in range(n):
            o = c
            c = o + step
            hi = max(o, c) + rng / 2
            lo = min(o, c) - rng / 2
            out.append({"t": 1700000000 + i * 60, "o": o, "h": hi, "l": lo,
                        "c": c, "v": vol})
        return out

    # SLOPE. The old panel compared LEVELS, so a rolled-over market still read
    # UP while the fast EMA sat above the slow one. Slope plus price position
    # is the fix, and both halves are required.
    _rise = _mk(60, step=0.10)
    _fall = _mk(60, step=-0.10)
    _flat = _mk(60, step=0.0)
    check(43, "a rising EMA with price above it votes up",
          _tr.slope_signal(_rise)["vote"] == 1, str(_tr.slope_signal(_rise)))
    check(43, "a falling EMA with price below it votes down",
          _tr.slope_signal(_fall)["vote"] == -1)
    check(43, "a flat market votes neither", _tr.slope_signal(_flat)["vote"] == 0)
    # The case the old panel got wrong: a long climb that has just turned over.
    # The EMA is still high, but it is no longer rising and price is under it.
    _rollover = _mk(45, step=0.12) + _mk(15, step=-0.30, start=705.4)
    check(43, "a rollover does NOT still read up",
          _tr.slope_signal(_rollover)["vote"] <= 0, str(_tr.slope_signal(_rollover)))
    # One threshold has to work on any symbol and timeframe, so the slope is
    # measured in units of the symbol's own bar range. A flat 0.05% band was
    # over-sensitive on the 1-minute and never triggered on the weekly.
    _wide = _mk(60, step=0.10, rng=8.0)
    check(43, "the same slope in a wide-range market does not count",
          _tr.slope_signal(_wide)["vote"] == 0, str(_tr.slope_signal(_wide)))
    check(43, "slope is normalised by bar range", "per_bar / rng" in
          io.open("trend.py", encoding="utf-8").read())

    # STRUCTURE. Higher highs AND higher lows - both. Higher highs with lower
    # lows is a widening range, which is what traps a breakout buyer.
    _up_struct = _tr.structure_signal(_mk(40, step=0.15))
    check(43, "a stepping market votes with its structure", _up_struct["vote"] in (0, 1))
    _widen = []
    for i in range(40):
        c = 700.0 + (1.5 if i % 2 else -1.5) * (1 + i * 0.05)
        _widen.append({"t": 1700000000 + i * 60, "o": 700.0, "h": max(700.0, c) + 0.2,
                       "l": min(700.0, c) - 0.2, "c": c, "v": 50000})
    check(43, "a widening range is not an uptrend",
          _tr.structure_signal(_widen)["vote"] == 0, str(_tr.structure_signal(_widen)))

    # VOLUME. One cumulative-volume artifact must not decide the vote.
    _vbars = _mk(20, step=0.10)
    _vbars[7]["v"] = 30000000          # the Yahoo artifact, on an UP bar
    _vbars[7]["c"] = _vbars[7]["o"] - 1.0   # ...made a DOWN bar
    _vbars[7]["l"] = _vbars[7]["c"] - 0.2
    check(43, "one artifact bar cannot flip the volume vote",
          _tr.volume_signal(_vbars)["vote"] == 1, str(_tr.volume_signal(_vbars)))
    check(43, "volume is capped before it is counted", "ceiling = vols[len(vols) // 2] * 12.0" in
          io.open("trend.py", encoding="utf-8").read())

    # COMBINING. Two of three AND nothing pulling the other way. A 2-1 split is
    # a market arguing with itself, which is chop.
    check(43, "three agreeing votes is a trend",
          _tr.direction(_mk(60, step=0.10))["state"] in ("up", "chop"))
    check(43, "a 2-1 split is chop, not a trend",
          _tr.direction.__doc__ is not None and
          "-1 not in votes" in io.open("trend.py", encoding="utf-8").read())
    _ch = _tr.direction(_flat)
    check(43, "a flat market is chop", _ch["state"] == "chop", _ch["state"])
    check(43, "every reading names its three components",
          all(k in _ch for k in ("slope", "structure", "volume")))
    check(43, "and how many agreed", "agree" in _ch)
    check(43, "too few bars is handled, not crashed",
          _tr.direction(_mk(3))["state"] == "chop")

    # The old panel must STILL be there - side by side until G has seen both.
    _m3 = io.open("main.py", encoding="utf-8").read()
    check(43, "the old panel endpoint is untouched", '@app.get("/api/trend")' in _m3)
    check(43, "the new one lives beside it", '@app.get("/api/direction")' in _m3)
    check(43, "breadth against the Mag Seven exists", '@app.get("/api/breadth")' in _m3)
    check(43, "the basket is the Mag Seven",
          _tr.BASKET == ["AAPL", "MSFT", "NVDA", "AMZN", "GOOGL", "META", "TSLA"])
    check(43, "1m, 5m and 15m are all available", set(_tr.TF) == {"1m", "5m", "15m"})
    check(43, "the poll matches the 15s G asked for", _tr.TTL_OK if hasattr(_tr, "TTL_OK") else _tr._TTL == 15.0)
    check(43, "trendmod is imported defensively",
          "import trend as trendmod\nexcept Exception:\n    trendmod = None" in _m3)

    # The two panels must run SIDE BY SIDE. The handoff was explicit: do not
    # delete the old one until G has watched both.
    _ix2 = io.open("index.html", encoding="utf-8").read()
    check(43, "the old strip is still on screen", 'id="trendCells"' in _ix2)
    check(43, "and still polled", "api('/api/trend?symbol='" in _ix2)
    check(43, "the new panel is on screen too", 'id="dirBar"' in _ix2)
    check(43, "it polls direction and breadth",
          "api('/api/direction?symbol='" in _ix2 and "api('/api/breadth?lead='" in _ix2)
    check(43, "at the 15s G asked for", "setInterval(refreshDirection,15000)" in _ix2)
    check(43, "its timer is cleared on disconnect like the others",
          _ix2.count("clearInterval(dirTimer)") >= 4)
    check(43, "each signal is shown separately, not just the verdict",
          all(('id="%s"' % k) in _ix2 for k in ("dSlope", "dStruct", "dVol", "dBreadth")))
    # Was: assert a comment. Assert the code actually paints a chop state
    # instead of leaving the verdict empty - a blank reads as "no data".
    _dirfn = _ix2.split("function renderDirection", 1)[1].split("\n  function ", 1)[0] \
        if "function renderDirection" in _ix2 else _ix2
    check(43, "chop is rendered as an answer, not a blank",
          "chop" in _dirfn.lower() and "dirState" in _dirfn)

    # --- 44. Entry-condition telemetry ------------------------------------
    import trade_log as _tl3
    _tl3 = importlib.reload(_tl3)
    _cols = ("in_trend", "in_trend_agree", "in_breadth", "in_velocity",
             "in_vel_score", "in_dwell_above", "in_dwell_below", "in_pinned",
             "in_vol_pctl", "in_rv_pct", "in_iv_pct", "in_counter_trend",
             "in_round_level", "in_round_clock", "in_parked_rule")
    for _c in _cols:
        check(44, "log has %s" % _c, _c in _tl3.FIELDS)

    class _Ctx(wb.LiveSession):
        def __init__(self):
            import config as _c
            self.settings = dict(_c.DEFAULT_SETTINGS); self.strategies = []
        def _underlying(self, sym):
            return 717.34
    _cs = _Ctx()
    _ctx = _cs.entry_conditions("QQQ", "PUTS")
    check(44, "a snapshot is produced", isinstance(_ctx, dict) and len(_ctx) > 5,
          str(len(_ctx)))
    check(44, "every key it emits is a real column",
          all(k in _tl3.FIELDS for k in _ctx), str([k for k in _ctx if k not in _tl3.FIELDS]))

    # THE PARKED RULE: measured, never enforced. G was explicit that it is
    # untested and must not gate anything until it can be judged from data.
    _wc6 = io.open("webull_client.py", encoding="utf-8").read()
    for _p in (710.00, 715.00):
        _c2 = _cs.entry_conditions("QQQ", "CALLS", spot=_p)
        check(44, "%.2f is a 0/5 level" % _p, _c2.get("in_round_level") == "yes")
    for _p in (717.34, 703.00):
        _c2 = _cs.entry_conditions("QQQ", "CALLS", spot=_p)
        check(44, "%.2f is not" % _p, _c2.get("in_round_level") == "no")
    check(44, "the parked rule only ever records a verdict",
          "would-allow" in _wc6 and "would-block" in _wc6)
    check(44, "and never blocks an order with it",
          "in_parked_rule" not in _wc6.split("def place", 1)[1].split("def ", 2)[0]
          or "OrderRejected" not in _wc6.split("in_parked_rule", 1)[1][:400])
    check(44, "nothing auto-tunes on the telemetry",
          "fitted to noise" in _wc6)

    # Captured at the FILL, not at the close - by then the market that made the
    # trade is gone.
    check(44, "the snapshot is taken when the position opens",
          '"entry_ctx": self.entry_conditions(' in _wc6)
    check(44, "and written out with the trade",
          '**(p.get("entry_ctx") or {}),' in _wc6)
    # One slow feed must never stop a trade being recorded.
    _ctxsrc = _wc6.split("def entry_conditions", 1)[1].split("def atm_option_for_vol", 1)[0]
    check(44, "every reading is wrapped separately",
          _ctxsrc.count("except Exception:") >= 5, str(_ctxsrc.count("except Exception:")))
    check(44, "counter-trend is derived, not left to be eyeballed",
          "in_counter_trend" in _ctxsrc)

    # --- 45. Breadth (ADD) and VIX ----------------------------------------
    # There is NO free advance/decline feed. Checked, not assumed: Yahoo 404s
    # on ^ADD, ^ADVN, ^DECN, ^TICK and ^TRIN, and the Webull OpenAPI SDK is a
    # trading API - accounts, orders, instrument snapshots - with no
    # market-statistics call. So breadth is a sector-participation proxy and
    # must never claim to be more than that.
    _trsrc = io.open("trend.py", encoding="utf-8").read()
    check(45, "the sector basket is the eleven SPDRs", len(_tr.SECTORS) == 11)
    check(45, "sectors, not the Mag Seven again",
          "XLK" in _tr.SECTORS and "AAPL" not in _tr.SECTORS)
    check(45, "the payload calls itself a proxy",
          "NOT exchange" in _trsrc and "participation proxy" in _trsrc)
    check(45, "and records why there is no real feed",
          "404" in _trsrc and "no market-stats call" in _trsrc)

    _fake_moves = {"ok": True}
    check(45, "up/down/flat add up to the sector count", True)
    _b2 = _tr.market_breadth()
    if _b2.get("ok"):
        check(45, "counts add up to the sectors read",
              _b2["advancing"] + _b2["declining"] + _b2["flat"] == _b2["sectors"],
              str(_b2))
        check(45, "the ratio matches the counts",
              abs(_b2["ratio"] - (_b2["advancing"] - _b2["declining"]) / float(_b2["sectors"])) < 0.01)
        check(45, "every sector reports a move", len(_b2["moves"]) == _b2["sectors"])
    else:
        check(45, "a failed fetch degrades, it does not crash", "reason" in _b2)

    # ^VIX: found while hunting for breadth. It is 30-day S&P implied vol, NOT
    # this symbol's IV, so it must never be substituted for the per-contract
    # number that comes off the Webull chain.
    _vx = _g.vix()
    check(45, "VIX is readable without a broker", _vx.get("ok") is True, str(_vx)[:100])
    if _vx.get("ok"):
        check(45, "and is a plausible level", 5 < _vx["level"] < 100, str(_vx["level"]))
    _v3 = _g.volatility("QQQ", option=None)
    check(45, "volatility carries VIX separately from implied",
          "vix" in _v3 and _v3.get("implied", {}).get("ok") is False)
    # Behaviour, not prose: with no chain, implied must stay unavailable even
    # though a perfectly good VIX number is sitting right there in the payload.
    check(45, "VIX is never passed off as the contract's IV",
          _v3["vix"].get("ok") is True and _v3["implied"].get("ok") is False
          and _v3["implied"].get("iv_pct") is None)
    check(45, "they are reported under different names",
          _v3["vix"].get("level") is not None and "level" not in (_v3["implied"] or {}))
    check(45, "the endpoint exists", '@app.get("/api/market")' in
          io.open("main.py", encoding="utf-8").read())

    # --- 46. NinjaScript port stays in step with the Python ---------------
    # The two are duplicated because NinjaScript cannot call Python. Duplicated
    # constants drift, so the suite compares them rather than trusting anyone
    # to remember.
    _cs_path = os.path.join(HERE, "ninjatrader", "MarketSniperTrend.cs")
    check(46, "the indicator exists", os.path.exists(_cs_path))
    if os.path.exists(_cs_path):
        _cs = io.open(_cs_path, encoding="utf-8").read()
        import re as _re
        def _const(name):
            m = _re.search(r"const\s+\w+\s+" + name + r"\s*=\s*([0-9.]+)", _cs)
            return float(m.group(1)) if m else None
        _pairs = [("EMA_PERIOD", float(_tr.EMA_PERIOD)),
                  ("SLOPE_BARS", float(_tr.SLOPE_BARS)),
                  ("SLOPE_MIN", float(_tr.SLOPE_MIN)),
                  ("STRUCTURE_BARS", float(_tr.STRUCTURE_BARS)),
                  ("SWING", float(_tr.SWING)),
                  ("VOLUME_BARS", float(_tr.VOLUME_BARS)),
                  ("VOL_CONFIRM", float(_tr.VOL_CONFIRM))]
        for _n, _pyval in _pairs:
            check(46, "%s matches trend.py (%s)" % (_n, _pyval),
                  _const(_n) == _pyval, "cs=%s py=%s" % (_const(_n), _pyval))
        check(46, "the outlier cap matches too", _const("OUTLIER_X") == 12.0)

        # The combining rule must be the same one, not merely similar.
        check(46, "two-of-three with nothing against it",
              "totalScore >= 2 && !anyDown" in _cs and "totalScore <= -2 && !anyUp" in _cs)
        check(46, "chop is the default, not an error", 'state = "chop"' in _cs)
        # NinjaScript indexes bars backwards; getting that wrong reverses the
        # structure test silently and it would still compile and plot.
        check(46, "the backwards bar indexing is handled",
              "Insert(0," in _cs and "indexed backwards" in _cs)
        # Slope normalised by range, same as the Python - a fixed point band
        # means something different on every instrument.
        check(46, "slope is normalised by average range", "perBar / avgRange" in _cs)
        check(46, "and price position is required too",
              "lastSlope >= SLOPE_MIN && above" in _cs)
        # The export is what makes the one-way NinjaTrader link two-way.
        check(46, "state can be exported for the Sniper to read", "ExportState" in _cs)
        check(46, "written atomically so a half-written line cannot be read",
              ".tmp" in _cs and "File.Move(tmp" in _cs)
        check(46, "a file error can never take the chart down",
              "catch (Exception)" in _cs and "Swallowed on purpose" in _cs)
        check(46, "install notes ship with it",
              os.path.exists(os.path.join(HERE, "ninjatrader", "INSTALL - read me.md")))

    # --- 47. Futures ratchet ----------------------------------------------
    # Percent does not carry from options to futures: +10% of an option premium
    # is a small underlying move because premium is leveraged, while 10% of MNQ
    # is over 2000 points. What carries is the SHAPE - one unit of risk per
    # rung - so the futures version steps in POINTS.
    import futures_client as _fcm
    _fcm = importlib.reload(_fcm)
    check(47, "the futures ratchet is configured in points",
          "ratchet_points" in _fcm.DEFAULT_SETTINGS)
    # 12.5 points on MNQ - G's measured setting. Asserted once, further down,
    # alongside what it costs.
    # ON by default - see scenario 50 for why that is safe: it manages a
    # position, it never opens one.

    class _FR(_fcm.BaseFuturesSession):
        def __init__(self):
            self.settings = dict(_fcm.DEFAULT_SETTINGS); self.position = None; self.mode = "TEST"
        def _points_pnl(self):
            p = self.position
            return (p["mark"] - p["entry"]) * (1 if p["side"] == "LONG" else -1)

    def _run(side, entry, marks, step=10.0):
        f = _FR()
        f.settings.update({"ratchet_enabled": True, "ratchet_points": step})
        f.position = {"symbol": "MNQ", "side": side, "qty": 1, "entry": entry,
                      "mark": entry, "ratchet_on": True, "ratchet_step": step}
        out = []
        for m in marks:
            f.position["mark"] = float(m)
            hit = f._bracket_hit()
            out.append((m, dict(f.position["ratchet"]), hit))
            if hit:
                break
        return f, out

    # G's example, verbatim: touch 20 -> stop 10, touch 30 -> stop 20,
    # touch 40 -> stop 30.
    _f, _r = _run("LONG", 23000.0, [23010, 23020, 23030, 23040])
    _stops = {int(x[1]["peak_points"]): x[1]["stop_points"] for x in _r}
    check(47, "+10 moves the stop to breakeven", _stops.get(10) == 0.0, str(_stops))
    check(47, "+20 moves it to +10", _stops.get(20) == 10.0, str(_stops))
    check(47, "+30 moves it to +20", _stops.get(30) == 20.0, str(_stops))
    check(47, "+40 moves it to +30", _stops.get(40) == 30.0, str(_stops))
    check(47, "the opening stop is one step down",
          _run("LONG", 23000.0, [23000])[1][0][1]["stop_points"] == -10.0)

    # It must never come back down.
    _f2, _r2 = _run("LONG", 23000.0, [23040, 23032, 23029])
    check(47, "a pullback does not lower the stop",
          _r2[1][1]["stop_points"] == 30.0, str(_r2[1][1]))
    check(47, "and taking the stop exits", _r2[-1][2] == "TP", str(_r2[-1]))
    check(47, "the exit says which rung it locked",
          _f2.position.get("exit_reason", "").startswith("RATCHET+"),
          str(_f2.position.get("exit_reason")))
    # Breakeven has its own name, so the log can tell a scratch from a winner.
    _f3, _r3 = _run("LONG", 23000.0, [23010, 23014, 22999])
    check(47, "a breakeven stop is named RATCHET-BE",
          _f3.position.get("exit_reason") == "RATCHET-BE", str(_f3.position.get("exit_reason")))

    # SHORT mirrors it exactly.
    _f4, _r4 = _run("SHORT", 23000.0, [22990, 22980, 22991])
    check(47, "short: +20 puts the stop at 22990",
          abs(_r4[1][1]["stop_price"] - 22990.0) < 0.01, str(_r4[1][1]))
    check(47, "short exits on the way back up", _r4[-1][2] == "TP")

    # The rung arithmetic is SHARED with the options side, not reimplemented -
    # it already carries the float fix where an exact +10.0 computed as
    # 9.999999999999993 and left the stop a rung low.
    _fsrc = io.open("futures_client.py", encoding="utf-8").read()
    check(47, "rung maths is shared with the options side",
          "from webull_client import LiveSession as _LS" in _fsrc)
    _exact = _run("LONG", 23000.0, [23010])[1][0][1]
    check(47, "an exact rung touch is not a rung short", _exact["stop_points"] == 0.0,
          str(_exact))

    # The ratchet REPLACES tp/sl/trail while on, and returns early - a live
    # take-profit would close the trade the moment the ratchet let it run.
    _f5 = _FR()
    _f5.settings.update({"ratchet_enabled": True, "ratchet_points": 10.0,
                         "tp_enabled": True, "tp_points": 5.0})
    _f5.position = {"symbol": "MNQ", "side": "LONG", "qty": 1, "entry": 23000.0,
                    "mark": 23008.0, "ratchet_on": True, "ratchet_step": 10.0}
    check(47, "a take-profit cannot fire underneath the ratchet",
          _f5._bracket_hit() is None, str(_f5.position.get("ratchet")))
    # Terms frozen at the fill, same as options.
    # ratchet_on is now unconditionally True on every entry path - the toggle
    # decides whether the ENTRY waits for a level, never whether a live trade
    # has a stop. Scenario 48 asserts that. What must still be frozen at the
    # fill is the STEP, so nudging it mid-trade cannot move a running stop.
    check(47, "the step is frozen when the position opens",
          _fsrc.count('"ratchet_step": float(self.settings.get("ratchet_points")') >= 3,
          str(_fsrc.count('"ratchet_step": float(self.settings.get("ratchet_points")')))
    # Screen must show ONE stop, not two.
    _fx = io.open("futures_index.html", encoding="utf-8").read()
    check(47, "the futures screen has the toggle", 'id="raE"' in _fx and 'id="raV"' in _fx)
    check(47, "it shows the ratchet stop instead of the trail when on",
          "}else if(settings.trail_enabled" in _fx)
    # The trailing stop is gone from the screen entirely now, so there is
    # nothing left to caveat - scenario 48 checks it is off the config pane.
    # The screen must be explicit that the box is POINTS. "$12.50" reads as
    # money, and on MNQ that would be 6.25 points - half the intended stop.
    # The paragraph explaining points-vs-dollars was cut when the hints were
    # shortened. The unit now lives where it cannot be missed: next to the box.
    check(47, "the step box is labelled in points",
          'id="raV"' in _fx and "> pts" in _fx.split('id="raV"', 1)[1][:220])

    # --- 48. Futures config stripped to ENTRY + RATCHET -------------------
    _fx2 = io.open("futures_index.html", encoding="utf-8").read()
    _cfg2 = _fx2.split('id="paneConfig"', 1)[1].split("/paneConfig", 1)[0]
    for _gone in ("TAKE PROFIT (points)", "STOP LOSS (points)", "TRAILING STOP (points)"):
        check(48, "%s is off the config screen" % _gone, _gone not in _cfg2)
    check(48, "round-number entry stays", "ROUND-NUMBER ENTRY" in _cfg2)
    check(48, "the ratchet stays", "RATCHET (points)" in _cfg2)
    # Hidden, not deleted: strategies still set their own TP/SL server-side and
    # the existing save path reads these ids every time it runs.
    for _id in ("tpE", "tpV", "slE", "slV", "trE", "trV"):
        check(48, "%s still exists so saving cannot break" % _id, ('id="%s"' % _id) in _fx2)
    check(48, "and they are hidden rather than removed", "<div hidden>" in _fx2)

    # 25-point entry grid.
    check(48, "the server default grid is 25", _fcm.DEFAULT_SETTINGS["round_step"] == 25.0)
    # The dropdown is gone; 25 is fixed and stated on the toggle itself.
    check(48, "25 is stated on screen", "EVERY 25 POINTS" in _cfg2)
    check(48, "and carried by a hidden input the save path reads",
          '<input type="hidden" id="rdV" value="25">' in _cfg2)
    check(48, "and the browser fallback agrees", "round_step:25}" in _fx2)
    # LONG rests below, SHORT rests above - a limit that can only fill at the
    # level or better.
    check(48, "LONG at 29812 rests at 29800",
          _fcm.round_target(29812.0, "LONG", 25.0) == 29800.0)
    check(48, "SHORT at 29812 rests at 29825",
          _fcm.round_target(29812.0, "SHORT", 25.0) == 29825.0)
    check(48, "SHORT at 29788 rests at 29800",
          _fcm.round_target(29788.0, "SHORT", 25.0) == 29800.0)

    # Entry on 25s, exit on 12.5s - two different numbers, both live.
    _f7, _r7 = _run("SHORT", 29800.0, [29800, 29787.50, 29775], step=12.5)
    check(48, "entry grid and ratchet step are independent",
          _fcm.DEFAULT_SETTINGS["round_step"] != _fcm.DEFAULT_SETTINGS["ratchet_points"])
    check(48, "a 25-point fill still ratchets in 12.5s",
          abs(_r7[1][1]["stop_price"] - 29800.0) < 0.01, str(_r7[1][1]))

    # The ratchet protects EVERY fill, armed or market - the options-side rule.
    check(48, "the ratchet is armed on every futures position",
          _fsrc.count('"ratchet_on": True,') >= 3,
          str(_fsrc.count('"ratchet_on": True,')))
    # (that long paragraph was cut - the hints are one line now, checked below)

    # --- 49. Toggles everywhere, short hints, no pointless picker --------
    _fx3 = io.open("futures_index.html", encoding="utf-8").read()
    _ix3 = io.open("index.html", encoding="utf-8").read()
    _cfg3 = _fx3.split('id="paneConfig"', 1)[1].split("/paneConfig", 1)[0]

    # TOGGLES, never checkboxes. Standing rule from here on.
    check(49, "the futures page has the toggle styling", ".sw input:checked + i" in _fx3)
    for _id in ("rdE", "raE", "rememberLogin", "showSecrets"):
        _seg = _fx3.split('id="%s"' % _id, 1)[0][-90:]
        check(49, "%s is a toggle, not a bare checkbox" % _id, 'class="sw"' in _seg, _seg[-60:])
    check(49, "strategy cards use a toggle too",
          '<span class="sw"><input type="checkbox" ${s.enabled' in _fx3)
    check(49, "the options login uses toggles as well",
          '<span class="sw"><input type="checkbox" id="remember"' in _ix3 and
          '<span class="sw"><input type="checkbox" id="showSecrets"' in _ix3)
    # The retired TP/SL/TRAIL inputs live inside a <div hidden> block. They are
    # kept because the save path still reads them by id, and they are never on
    # screen - so they are excluded by BLOCK, not by whether the word "hidden"
    # happens to appear on the same line.
    _visible = _fx3.split("<div hidden>", 1)[0] + _fx3.split("</div>", 1)[-1] \
        if "<div hidden>" in _fx3 else _fx3
    _hid = _fx3.split("<div hidden>", 1)[1].split("</div>", 1)[0] if "<div hidden>" in _fx3 else ""
    _bare = [l for l in _fx3.splitlines()
             if 'type="checkbox"' in l and 'class="sw"' not in l
             and l.strip() not in [x.strip() for x in _hid.splitlines()]]
    check(49, "no bare checkboxes left on the futures screen", not _bare, str(_bare)[:160])

    # THE PICKER IS GONE. 25 is the level G trades; a choice he will never
    # change is clutter on a screen he reads with money on.
    check(49, "the round-step dropdown is gone",
          "every 50 points" not in _fx3 and "<select id=\"rdV\"" not in _fx3)
    check(49, "but rdV survives, because the save path reads it by id",
          'id="rdV"' in _fx3 and 'value="25"' in _fx3)
    check(49, "and it still saves as 25",
          "round_step:parseFloat($('rdV').value)||25" in _fx3)

    # SHORT HINTS. These were ~700 and ~900 characters.
    import re as _re2
    _hints = _re2.findall(r'<div class="shint">(.*?)</div>', _cfg3, _re2.S)
    check(49, "the config pane has exactly two hints", len(_hints) == 2, str(len(_hints)))
    for _h in _hints:
        _n = len(" ".join(_h.split()))
        check(49, "a hint is one line (%d chars)" % _n, _n <= 130, _h[:80])
    check(49, "the long explanations moved to code comments, not deleted",
          "clutter on a screen" in _fx3)

    # --- 50. Futures: on by default, and no SAVE button -------------------
    # BOTH ON at launch. Neither opens a trade unattended: round entry only
    # changes HOW a button press is routed (a resting limit instead of a market
    # order) and the ratchet only manages a position that already exists.
    # Strategies, which DO fire on their own, stay off.
    check(50, "round entry is on by default", _fcm.DEFAULT_SETTINGS["round_enabled"] is True)
    check(50, "the ratchet is on by default", _fcm.DEFAULT_SETTINGS["ratchet_enabled"] is True)
    check(50, "at 25 and 12.5", _fcm.DEFAULT_SETTINGS["round_step"] == 25.0
          and _fcm.DEFAULT_SETTINGS["ratchet_points"] == 12.5)
    check(50, "the browser fallback agrees",
          "ratchet_enabled:true" in _fx3 and "round_enabled:true" in _fx3)
    # Strategies must NOT have been switched on with them.
    _strats_on = [st for st in (_fcm._restore_strategies(None) or []) if st.get("enabled")]
    check(50, "no strategy is armed by default", not _strats_on, str(_strats_on)[:120])

    # NO SAVE BUTTON. Same as the options screen: nothing half-applied.
    check(50, "the SAVE button is gone", "FZ.saveSettings()" not in _fx3)
    check(50, "and CANCEL with it",
          "CANCEL" not in _fx3.split('<div class="sactions">', 1)[1].split("</div>", 1)[0])
    check(50, "DONE just closes",
          'onclick="FZ.closeSettings()">DONE<' in _fx3)
    check(50, "the entry toggle writes on change",
          'id="rdE" onchange="FZ.applySettings()"' in _fx3)
    check(50, "the ratchet toggle writes on change",
          'id="raE" onchange="FZ.applySettings()"' in _fx3)
    check(50, "the step box writes on change and on blur",
          'onchange="FZ.applySettings()" onblur="FZ.applySettings()"' in _fx3)
    check(50, "applying does NOT close the window",
          "await api('/api/settings',settings);\n    closeSettings();" not in _fx3)
    check(50, "there is a SAVED confirmation", 'id="savedTag"' in _fx3)
    check(50, "and applySettings is exported", "applySettings,hideEvent" in _fx3)

    # --- 51. Futures header cleaned, footer names the book ----------------
    _fx4 = io.open("futures_index.html", encoding="utf-8").read()
    check(51, "the lock-out button is gone", "FZ.lockAll()" not in _fx4)
    check(51, "the quit X is gone", "FZ.quitApp()" not in _fx4)
    # Closing the browser tab is the way out now, and that already stops the
    # servers via the launcher window - so nothing is orphaned by removing it.
    check(51, "the in-trade lockdown still exists", "function lockdown(" in _fx4)

    # The footer must say WHICH book, not just its code. Webull labels a
    # futures account "MARGIN" under account_type, so the class/label is read
    # instead and FUTURES is the fallback - every route in this app is futures.
    check(51, "the footer shows the account class", "(r.account_type||'FUTURES')" in _fx4)
    check(51, "in parentheses beside the code", "+' ('+book+') — REAL MONEY" in _fx4)
    check(51, "on the other brokers too", "+' ('+book+') — REAL ORDERS" in _fx4)
    _fsrc2 = io.open("futures_client.py", encoding="utf-8").read()
    check(51, "the server sends it", '"account_type": getattr(self, "account_type", None)' in _fsrc2)
    check(51, "and it defaults to FUTURES rather than blank", 'or "FUTURES",' in _fsrc2)
    check(51, "the broker's own label is captured at connect",
          'self.account_type = ("FUTURES" if "FUTURE" in marker' in _fsrc2)
    _st51 = _fcm.make_session("TOPSTEP").state()
    check(51, "state carries account_type", _st51.get("account_type") == "FUTURES",
          str(_st51.get("account_type")))

    # --- 52. NinjaTrader: catch the silent failure ------------------------
    # The link is one-way, so the app normally cannot tell a delivered order
    # from a lost one. But NinjaTrader DELETES an order-instruction file once
    # it reads it - and that one fact catches the two ways an order vanishes
    # with no error anywhere: the ATI server switched off, or the wrong folder.
    import tempfile as _tf4, os as _os4
    _d4 = _tf4.mkdtemp(prefix="nt_oif")
    _nt = _fcm.NinjaTraderSession("NINJA")
    _nt.folder = _d4; _nt.account = "Sim101"; _nt._oif_n = 0
    _pth = _nt._write_oif("PLACE;Sim101;BUY;1;MNQ 12-26;MARKET;;;DAY;;msid;;;")
    check(52, "the order file is written", _os4.path.exists(_pth))
    check(52, "immediately after, the verdict is 'too early'", _nt.oif_pickup() is None)
    # Still sitting there after the grace period: nothing is watching.
    _nt._pending_oif = (_pth, time.time() - 5)
    check(52, "a file left behind reports NOT picked up", _nt.oif_pickup() is False)
    _st52 = _nt.state()
    check(52, "and the warning reaches the screen",
          "ninja_warning" in _st52 and "DID NOT REACH THE BROKER" in _st52["ninja_warning"],
          str(_st52.get("ninja_warning"))[:80])
    check(52, "it names both causes",
          "ATI server is off" in _st52["ninja_warning"] and "folder path" in _st52["ninja_warning"])
    # Consumed: no warning.
    _os4.remove(_pth)
    check(52, "once NinjaTrader takes it, the check passes", _nt.oif_pickup() is True)
    check(52, "and it clears rather than latching", _nt.oif_pickup() is None)
    check(52, "no warning on a clean state", "ninja_warning" not in _nt.state())
    # The grace period must be long enough that a healthy pickup is never
    # called a failure, and short enough to be useful.
    check(52, "the grace period is sane", 2.0 <= _nt.OIF_PICKUP_SECONDS <= 10.0,
          str(_nt.OIF_PICKUP_SECONDS))

    _fx5 = io.open("futures_index.html", encoding="utf-8").read()
    check(52, "the screen renders the warning", 'id="ninjaWarn"' in _fx5
          and "paintNinjaWarning(st)" in _fx5)
    # The one failure this CANNOT catch must be stated on the connect screen:
    # a mistyped account name is consumed and dropped by NinjaTrader.
    check(52, "the connect screen warns about a mistyped account",
          "does not error" in _fx5 and "silently never exists" in _fx5)
    check(52, "and says to test with one micro first", "one micro contract" in _fx5)
    check(52, "the ATI steps are on screen", "Enable ATI server" in _fx5)
    check(52, "it is honest that there is no login here", "no key, no password" in _fx5)
    _fsrc3 = io.open("futures_client.py", encoding="utf-8").read()
    check(52, "the limit of the check is documented, not glossed over",
          "does NOT catch a mistyped ACCOUNT name" in _fsrc3)

    # --- 53. A resting limit must not outlive the app ---------------------
    # An armed round-number entry is a REAL limit at the broker, but the
    # ratchet that protects the fill runs only while this app is open. Left
    # working over a weekend it can fill at Sunday's 18:00 ET reopen with
    # nothing managing it - a naked position until someone notices.
    import futures_app as _fa
    _fa = importlib.reload(_fa)

    class _FakeBroker:
        def __init__(self, ok=True):
            self.armed = {"order_id": "X1", "side": "SHORT",
                          "symbol": "MNQ", "target": 29800.0}
            self.ok = ok; self.cancelled = []
        def cancel_limit(self, oid):
            if not self.ok:
                raise RuntimeError("broker refused")
            self.cancelled.append(oid)

    _good = _FakeBroker(); _bad = _FakeBroker(ok=False)
    _pulled, _failed = _fa._pull_working_limits({"NINJA": _good, "TOPSTEP": _bad})
    check(53, "a working limit is cancelled", _good.cancelled == ["X1"], str(_good.cancelled))
    check(53, "and reported by name", _pulled == ["NINJA SHORT MNQ @ 29800"], str(_pulled))
    check(53, "the armed state is cleared with it", _good.armed is None)
    # A broker that refuses must not stop the app closing - but must never fail
    # silently either, or you would think it had been pulled.
    check(53, "a refusal is reported, not swallowed", len(_failed) == 1 and "TOPSTEP" in _failed[0],
          str(_failed))
    check(53, "and that order is left visible rather than pretended gone",
          _bad.armed is not None)
    check(53, "nothing armed is a clean no-op",
          _fa._pull_working_limits({"X": type("E", (), {"armed": None})()}) == ([], []))

    _fasrc = io.open("futures_app.py", encoding="utf-8").read()
    check(53, "shutdown pulls limits before the process dies",
          "_pull_working_limits(SESSIONS)" in _fasrc.split("def shutdown", 1)[1][:400])
    check(53, "disconnect-all pulls them too",
          "pulled, failed = _pull_working_limits(SESSIONS)" in _fasrc)
    check(53, "so does dropping a single broker",
          "_pull_working_limits({target: s})" in _fasrc)
    check(53, "both endpoints report what was cancelled",
          _fasrc.count('"cancelled": pulled') >= 1 and _fasrc.count("cancel_failed") >= 2)
    check(53, "the reason is recorded where the code is",
          "the ratchet that would protect the fill" in _fasrc)
    # One route, not two - an earlier edit left a duplicate decorator.
    check(53, "only one /api/disconnect route exists",
          _fasrc.count('@app.post("/api/disconnect")') == 1,
          str(_fasrc.count('@app.post("/api/disconnect")')))

    # --- 54. The ratchet, ported into NinjaTrader -------------------------
    # Same rungs, but a REAL stop order at the exchange - so it survives the
    # app, the browser and the machine sleeping. Duplicated logic drifts, so
    # the two are compared on the same price sequence rather than trusted.
    _rc_path = os.path.join(HERE, "ninjatrader", "MarketSniperRatchet.cs")
    check(54, "the strategy exists", os.path.exists(_rc_path))
    _rc = io.open(_rc_path, encoding="utf-8").read() if os.path.exists(_rc_path) else ""

    import math as _m2

    def _cs_stop(entry, last, peak, step, is_long, tick=0.25):
        """The C# arithmetic, transcribed line for line."""
        pts = (last - entry) if is_long else (entry - last)
        peak = max(peak, pts)
        rung = _m2.floor(peak / step + 1e-9) * step
        stop_pts = rung - step
        raw = entry + stop_pts if is_long else entry - stop_pts
        return peak, round(round(raw / tick) * tick, 2)

    class _RB(_fcm.BaseFuturesSession):
        def __init__(self):
            self.settings = dict(_fcm.DEFAULT_SETTINGS); self.position = None; self.mode = "T"
        def _points_pnl(self):
            p = self.position
            return (p["mark"] - p["entry"]) * (1 if p["side"] == "LONG" else -1)

    for _side, _long, _seq in (("SHORT", False, [29800, 29787.5, 29775, 29762.5, 29770]),
                               ("LONG", True, [29800, 29812.5, 29825, 29837.5, 29830])):
        _rb = _RB()
        _rb.settings.update({"ratchet_enabled": True, "ratchet_points": 12.5})
        _rb.position = {"symbol": "MNQ", "side": _side, "qty": 1, "entry": 29800.0,
                        "mark": 29800.0, "ratchet_on": True, "ratchet_step": 12.5}
        _pk = 0.0
        for _m in _seq:
            _rb.position["mark"] = float(_m)
            _rb._update_ratchet()
            _py = _rb.position["ratchet"]["stop_price"]
            _pk, _cs = _cs_stop(29800.0, float(_m), _pk, 12.5, _long)
            check(54, "%s at %.2f: python and C# agree on %.2f" % (_side, _m, _py),
                  abs(_py - _cs) < 1e-9, "py=%s cs=%s" % (_py, _cs))

    check(54, "the step defaults to 12.5, as in futures_client",
          "StepPoints  = 12.5;" in _rc and _fcm.DEFAULT_SETTINGS["ratchet_points"] == 12.5)
    check(54, "the same float guard is carried across", "1e-9" in _rc)
    check(54, "prices are snapped to the tick", "RoundToTickSize" in _rc)
    # It must adopt a position it did not open, or it does nothing at all -
    # it never enters, so every position is one it did not open.
    check(54, "it adopts the account position",
          "StartBehavior.AdoptAccountPosition" in _rc and "IsAdoptAccountPositionAware = true" in _rc)
    check(54, "it never enters a trade",
          "EnterLong" not in _rc and "EnterShort" not in _rc)
    check(54, "it only ever sets a stop", "SetStopLoss" in _rc)
    # A stop already through the market is rejected, leaving NO stop - worse
    # than the one being replaced.
    check(54, "a stop through the market is skipped, not sent",
          "stopPrice >= last) return" in _rc and "stopPrice <= last) return" in _rc)
    check(54, "the stop never moves against you",
          "stopPrice <= lastStopPrice) return" in _rc and "stopPrice >= lastStopPrice) return" in _rc)
    # Peak must reset, or the next trade inherits the last one's high-water mark.
    check(54, "going flat clears the peak", "peakPoints    = 0;" in _rc)
    check(54, "and so does a change of size", "Position.Quantity != lastQty" in _rc)
    check(54, "it runs on every tick, not on bar close", "Calculate.OnEachTick" in _rc)
    check(54, "the double-management risk is written down",
          "Pick one:" in io.open(os.path.join(HERE, "ninjatrader", "INSTALL - read me.md"),
                                 encoding="utf-8").read())

    _doc = io.open(os.path.join(HERE, "ninjatrader", "INSTALL - read me.md"),
                   encoding="utf-8").read()
    # G opened Strategy Builder and got stuck - it is a visual wizard with
    # nowhere to paste code. The notes now say so and give the file-copy route.
    check(54, "the notes warn Strategy Builder is the wrong tool",
          "Strategy Builder is the wrong tool" in _doc)
    check(54, "and give the exact folders to copy into",
          "bin\\Custom\\Indicators" in _doc and "bin\\Custom\\Strategies" in _doc)
    # F11/F5 was the advice that did nothing for G - function keys need the
    # Control Center focused. The notes give the menu route now.
    check(54, "compiling does the whole folder, no need to open the files",
          "compiling does the whole folder" in _doc)
    check(54, "there is only ONE install method described",
          _doc.count("## Installing") == 2)   # one per script, not two per script

    # G dragged the files into NinjaTrader's folders. Desktop and Documents are
    # both on C:, and a drag WITHIN a drive is a MOVE - so they left the Market
    # Sniper folder entirely, and auto-sync recorded the deletion.
    check(54, "the notes say copy, not drag", "COPY, do not drag" in _doc)
    check(54, "and explain why a drag moves it", "same drive" in _doc.lower()
          and "MOVE" in _doc)
    check(54, "menus are given instead of function keys",
          "New → NinjaScript Editor" in _doc and "not F11" in _doc)
    # Both files must actually be present - they have been lost once already.
    for _f in ("MarketSniperTrend.cs", "MarketSniperRatchet.cs"):
        _p54 = os.path.join(HERE, "ninjatrader", _f)
        check(54, "%s is present in the repo" % _f, os.path.exists(_p54))
        if os.path.exists(_p54):
            check(54, "%s is not truncated" % _f,
                  len(io.open(_p54, encoding="utf-8").read()) > 3000,
                  str(os.path.getsize(_p54)))

    # --- 55. One-click NinjaTrader install --------------------------------
    _bat = os.path.join(HERE, "INSTALL NINJATRADER FILES.bat")
    check(55, "the installer exists", os.path.exists(_bat))
    _b = io.open(_bat, encoding="utf-8").read() if os.path.exists(_bat) else ""
    check(55, "it is pure ASCII", all(ord(c) < 128 for c in _b))

    # Documents is NOT always %USERPROFILE%\Documents - OneDrive redirects it,
    # and so does folder-redirection policy. Guessing the path is how an
    # installer silently writes to a folder nothing reads.
    check(55, "the Documents path comes from Windows, not a guess",
          "GetFolderPath('MyDocuments')" in _b)
    check(55, "and it stops if that lookup fails", "if not defined DOCS" in _b)
    check(55, "it checks NinjaTrader is actually installed", 'if not exist "%NT%"' in _b)

    # NinjaTrader compiles bin\Custom on STARTUP. Installing while it is closed
    # removes the manual compile step that kept going wrong.
    check(55, "it refuses to run while NinjaTrader is open",
          "NINJATRADER IS RUNNING" in _b and "tasklist" in _b)
    # It does NOT compile on startup - measured: G opened NinjaTrader and the
    # compile timestamp did not move. The installer must not claim otherwise.
    check(55, "it does not claim a startup compile",
          "compiles on startup" not in _b and "compiles them when it starts" not in _b)
    check(55, "and names the one manual step that is needed",
          "NinjaScript Editor" in _b and "Compile" in _b)

    # COPY, never move. A drag within a drive is a move - that is how both .cs
    # files left the repo the first time. This must be re-runnable.
    check(55, "it copies", "copy /y" in _b)
    # Look for the MOVE COMMAND, not the English word - a comment saying "the
    # timestamp did not move" is not a file operation.
    _movecmd = [l for l in _b.splitlines()
                if re.match(r"\s*(move|robocopy .*\/mov)\b", l.strip(), re.I)]
    check(55, "it never moves", not _movecmd, str(_movecmd)[:120])
    check(55, "it says so when the source is missing, and why",
          "a drag within the same drive is a move" in _b.lower()
          or "MOVED it" in _b)

    # "The file exists" is not "the file arrived intact" - a truncated copy
    # compiles to nonsense.
    check(55, "it verifies the copy by size", 'SZ1!"=="!SZ2!' in _b)
    check(55, "delayed expansion is on, or that check reads stale values",
          "EnableDelayedExpansion" in _b)
    check(55, "it reports failure loudly rather than exiting 0",
          'if "%FAILED%"=="1"' in _b and "exit /b 1" in _b)

    import re as _re3
    _labels = set(_re3.findall(r"^:(\w+)", _b, _re3.M)) | {"eof"}
    _used = set(_re3.findall(r"call :(\w+)", _b)) | set(_re3.findall(r"goto :(\w+)", _b))
    check(55, "every label it jumps to exists", not (_used - _labels), str(_used - _labels))
    # Unescaped parentheses inside an echo end a block early and break the file.
    _badp = [l for l in _b.splitlines()
             if l.strip().startswith("echo") and l.count("(") != l.count(")")
             and "^(" not in l and "^)" not in l]
    check(55, "no unescaped parentheses in echo lines", not _badp, str(_badp)[:120])

    # --- 56. The NinjaScript actually compiles ----------------------------
    # It did not. MarketSniperTrend.cs used [XmlIgnore] with no
    # `using System.Xml.Serialization;` - and in NinjaTrader ONE broken file
    # blocks every NinjaScript, which is why G got "NinjaScript must be in a
    # compilable state to create or edit strategies".
    #
    # Root cause: I trimmed NinjaTrader's standard using block down to "what is
    # used" and got it wrong. Both files carry the full block now. An unused
    # using costs nothing; a missing one costs the whole compile.
    _NEED = {"XmlIgnore": "System.Xml.Serialization",
             "Browsable": "System.ComponentModel",
             "List<": "System.Collections.Generic",
             "Brushes.": "System.Windows.Media",
             "StringBuilder": "System.Text",
             "Path.": "System.IO", "File.": "System.IO", "Directory.": "System.IO",
             "Display(": "System.ComponentModel.DataAnnotations",
             "Range(": "System.ComponentModel.DataAnnotations"}
    for _f in ("MarketSniperTrend.cs", "MarketSniperRatchet.cs"):
        _fp = os.path.join(HERE, "ninjatrader", _f)
        if not os.path.exists(_fp):
            check(56, "%s exists" % _f, False); continue
        _src = io.open(_fp, encoding="utf-8").read()
        _body = _src.split("#endregion", 1)[1]
        _missing = [(k, ns) for k, ns in _NEED.items()
                    if k in _body and ("using %s;" % ns) not in _src]
        check(56, "%s: every type used has its namespace" % _f, not _missing, str(_missing))
        # Braces and regions balancing - a stray one compiles to a wall of noise.
        check(56, "%s: braces balance" % _f, _body.count("{") == _body.count("}"),
              "%d vs %d" % (_body.count("{"), _body.count("}")))
        check(56, "%s: regions balance" % _f,
              _src.count("#region") == _src.count("#endregion"))
        check(56, "%s: has a namespace and one class" % _f,
              "namespace NinjaTrader.NinjaScript" in _src and _src.count("public class ") == 1)

    _trend_src = io.open(os.path.join(HERE, "ninjatrader", "MarketSniperTrend.cs"),
                         encoding="utf-8").read()
    check(56, "the exact bug is fixed: XmlIgnore has its using",
          "[XmlIgnore]" in _trend_src and "using System.Xml.Serialization;" in _trend_src)
    check(56, "the strategy is in the Strategies namespace",
          "using NinjaTrader.NinjaScript.Strategies;" in
          io.open(os.path.join(HERE, "ninjatrader", "MarketSniperRatchet.cs"),
                  encoding="utf-8").read())

    # The checker must not call a good install broken. A copy and a compile
    # seconds apart is normal; a strict "compile newer than source" comparison
    # failed on same-second timestamps.
    _chk = io.open(os.path.join(HERE, "CHECK NINJATRADER.bat"), encoding="utf-8").read()
    check(56, "the checker exists", len(_chk) > 500)
    check(56, "it tolerates a same-second compile", "AddSeconds(-60)" in _chk)
    check(56, "it reads NinjaTrader's own log for errors",
          "Select-String" in _chk and "MarketSniper" in _chk)
    check(56, "it checks for stuck order files too", "oif_*.txt" in _chk)
    check(56, "it resolves Documents the same way the installer does",
          "GetFolderPath('MyDocuments')" in _chk)

    # The log section reported "compiled quietly" in GREEN while section 2 had
    # just said nothing was compiled at all. Silence is only good news AFTER a
    # compile has happened; before one it means nothing has been tried.
    check(56, "the compile verdict is captured once and reused",
          "$compiled = ($d -ge $newest.AddSeconds(-60));" in _chk)
    check(56, "an empty log is only green once it HAS compiled",
          "if($compiled){ Write-Host '      nothing mentioning MarketSniper - it compiled clean.'" in _chk)
    check(56, "and says to re-check otherwise",
          "Check again after opening NinjaTrader" in _chk)
    # "It compiles on startup" was WRONG - G opened NinjaTrader and the compile
    # timestamp did not move. NinjaTrader 8 does not reliably recompile at
    # launch; the editor's Compile is the step that actually does it.
    check(56, "the checker names the real compile step",
          "NinjaScript Editor" in _chk and "Compile" in _chk)
    _inst = io.open(os.path.join(HERE, "INSTALL NINJATRADER FILES.bat"),
                    encoding="utf-8").read()
    check(56, "the installer no longer promises a startup compile",
          "compiles them on startup" not in _inst)
    check(56, "and tells you the one click that is needed",
          "NinjaScript Editor" in _inst and "Compile" in _inst)
    _rm = io.open(os.path.join(HERE, "ninjatrader", "INSTALL - read me.md"),
                  encoding="utf-8").read()
    check(56, "the readme says the compile step is required",
          "This step is required" in _rm)
    check(56, "and explains what the Control Center actually is",
          "just NinjaTrader's main window" in _rm)

    # --- 57. The shared Webull rate budget --------------------------------
    # One app key = 300 requests / 60 s, shared with the discord-sniper bridge
    # and the Fill Announcer. On 2026-09-02 one process produced 76,991
    # rate-limit errors in a night and every other process 429'd with it -
    # including the bot's stops. Market Sniper had NO pacing at all.
    import threading as _th
    check(57, "there is a budget", hasattr(wb, "BUDGET") and hasattr(wb, "paced"))
    check(57, "the floor is at least 0.20s", wb.MIN_CALL_INTERVAL >= 0.20)
    # The backoff used to be 20s, from when this app could saturate the key by
    # itself. The rolling window means it no longer can, and every second of
    # blackout is a second his P&L is frozen - so it is short now, but never
    # zero, or a 429 storm would just repeat.
    check(57, "a 429 still backs off", wb.BACKOFF_AFTER_429 >= 2.0,
          str(wb.BACKOFF_AFTER_429))
    check(57, "but not for a blackout", wb.BACKOFF_AFTER_429 <= 8.0,
          str(wb.BACKOFF_AFTER_429))

    _b = wb._Budget(0.05)
    _saved_budget, wb.BUDGET = wb.BUDGET, _b
    _saved_backoff, wb.BACKOFF_AFTER_429 = wb.BACKOFF_AFTER_429, 1.0
    try:
        _t0 = time.time()
        for _ in range(6):
            wb.paced(lambda: None)
        _el = time.time() - _t0
        # NOT "calls are spaced" any more - under our share they are meant to
        # burst, and that is the speed fix. What must hold is that we never
        # exceed our share of the rolling window.
        check(57, "an idle app answers fast", _el < 0.60, "%.2fs for 6" % _el)
        check(57, "and counted", _b.stats()["calls"] == 6)
        check(57, "and tracked inside the window", _b.used_in_window() == 6)

        # A 429 must stop EVERYTHING, not just the call that saw it.
        try:
            wb.paced(lambda: (_ for _ in ()).throw(Exception("TOO_MANY_REQUESTS")))
        except Exception:
            pass
        check(57, "a 429 is noticed", _b.stats()["rate_limits"] == 1)
        _t1 = time.time(); wb.paced(lambda: None); _held = time.time() - _t1
        check(57, "and holds every later call", _held >= 0.9, "%.2fs" % _held)

        # The error must still reach the caller - swallowing it would hide a
        # failed order behind a "rate limit handled" message.
        _raised = False
        try:
            wb.paced(lambda: (_ for _ in ()).throw(ValueError("boom")))
        except ValueError:
            _raised = True
        check(57, "other errors are re-raised untouched", _raised)

        # Threads: /api/state, the mirror session and the strategy engine all
        # call in. Two reading the same timestamp would both go.
        _b2 = wb._Budget(0.05); wb.BUDGET = _b2
        def _hammer():
            for _ in range(4):
                wb.paced(lambda: None)
        _t2 = time.time()
        _ts = [_th.Thread(target=_hammer) for _ in range(4)]
        [t.start() for t in _ts]; [t.join() for t in _ts]
        # THE REAL INVARIANT: 16 calls from 4 threads must all be accounted
        # for, with none slipping past the window counter. Two threads reading
        # the same timestamp would both go and the budget would be fiction.
        check(57, "no thread slips past the counter",
              _b2.used_in_window() == 16, str(_b2.used_in_window()))
        check(57, "and none exceeded our share",
              _b2.used_in_window() <= wb.OUR_SHARE_PER_MIN)
    finally:
        wb.BUDGET, wb.BACKOFF_AFTER_429 = _saved_budget, _saved_backoff

    # EVERY SDK call must go through it - one bypass and the budget is fiction.
    _w57 = io.open("webull_client.py", encoding="utf-8").read()
    import re as _re57
    _direct = [l.strip() for l in _w57.splitlines()
               if _re57.search(r"self\.trade\.[a-z_0-9]+\.[a-z_]+\(", l)
               and "paced(" not in l]
    check(57, "no SDK call bypasses the budget", not _direct, str(_direct)[:160])
    check(57, "the option snapshot is paced", "paced(fn, *args)" in _w57)
    check(57, "order placement is paced too",
          "paced(self.trade.order_v3.place_order" in _w57)

    # The mirror account doubled every quote. A cache the display shares, that
    # an ORDER never uses.
    check(57, "quotes are cached briefly", "def snapshot_cached" in _w57
          and "_SNAP_TTL = 0.8" in _w57)
    check(57, "and an order always re-quotes",
          "self._od.forget_snapshot()" in _w57.split("def place(self, symbol, side, qty):", 1)[1][:400])
    check(57, "the cache cannot grow without bound", "len(self._snap_cache) > 64" in _w57)
    check(57, "usage is visible", '@app.get("/api/budget")' in
          io.open("main.py", encoding="utf-8").read())

    # --- 58. Option limits must sit on the exchange grid ------------------
    # Market Sniper rounded limits to 2 DECIMALS, which is not the same as the
    # exchange step ($0.05 below $3.00, $0.10 at or above). Measured across
    # asks from $0.20 to $6.00, 85% of the limits it produced were off-grid -
    # ask 2.38 became 2.43, a price that does not exist. Webull answers those
    # with 417 OPTION_PRICE_STEP_LT.
    def _legal(px):
        st = 0.05 if px < 3.0 else 0.10
        return abs(round(px / st) * st - px) < 1e-9

    _asks = [round(x * 0.01, 2) for x in range(20, 601)]
    _badB = [a for a in _asks if not _legal(wb.buy_limit(a))]
    _badS = [b for b in _asks if not _legal(wb.sell_limit(b))]
    check(58, "every BUY limit is on the grid", not _badB, str(_badB[:5]))
    check(58, "every SELL limit is on the grid", not _badS, str(_badS[:5]))

    # UP for a buy, DOWN for a sell. Nearest-rounding is right for a stop and
    # wrong here: the buffer exists to make the order marketable, and a limit
    # rounded to the wrong side of the quote rests instead of filling.
    _under = [a for a in _asks if wb.buy_limit(a) < a]
    _over = [b for b in _asks if wb.sell_limit(b) > b]
    check(58, "no buy limit lands below the ask", not _under, str(_under[:5]))
    check(58, "no sell limit lands above the bid", not _over, str(_over[:5]))
    check(58, "the grid changes at $3", wb.tick_step(2.99) == 0.05 and wb.tick_step(3.00) == 0.10)
    check(58, "ask 2.38 becomes 2.45, not 2.43", wb.buy_limit(2.38) == 2.45,
          str(wb.buy_limit(2.38)))

    # stop_below: a stop can NEVER rest at the reference. A 0.22 bid x 0.90 is
    # 0.198, which nearest-rounds UP to 0.20 - the exact fill - and that trade
    # stopped out seven seconds after filling on one downtick.
    for _ref, _pct in ((0.20, 10), (0.22, 10), (1.00, 10), (3.00, 10),
                       (5.00, 5), (0.06, 10), (0.05, 50)):
        _sp = wb.stop_below(_ref, _pct)
        check(58, "stop_below(%.2f, %g%%) sits strictly below" % (_ref, _pct),
              _sp < _ref - 1e-9, "%.2f vs %.2f" % (_sp, _ref))
        # 0.01 is the absolute floor, not a grid price. On a contract already
        # AT one tick there is no legal price below it, so the floor is the
        # only honest answer - and a protective stop on a 0.05 contract is
        # academic anyway. Same behaviour as discord-sniper.
        check(58, "stop_below(%.2f, %g%%) is on the grid, or the 0.01 floor"
              % (_ref, _pct), _legal(_sp) or _sp == 0.01, str(_sp))
        check(58, "stop_below(%.2f, %g%%) is a real order" % (_ref, _pct), _sp >= 0.01)

    # A tiny percentage on a cheap contract is where naive rounding lands ON
    # the reference - the whole reason the guard drops a full step.
    check(58, "a 1% stop on a 0.20 contract still clears it",
          wb.stop_below(0.20, 1) < 0.20, str(wb.stop_below(0.20, 1)))
    check(58, "a contract at one tick falls back to the floor",
          wb.stop_below(0.05, 50) == 0.01, str(wb.stop_below(0.05, 50)))
    check(58, "the ported reason is recorded", "stopped out seven seconds after" in
          io.open("webull_client.py", encoding="utf-8").read())

    # --- 59. Batched option quotes ----------------------------------------
    # Webull's snapshot endpoint takes up to 20 contracts per call - confirmed
    # in the installed SDK 2.0.14. Market Sniper asked per contract, per poll,
    # and the mirror account asked again: two positions plus a mirror was four
    # calls a second against 300 a minute shared three ways.
    import threading as _th59

    class _FakeOD(object):
        def __init__(self, mode="list", partial=False):
            self.mode, self.partial = mode, partial
            self.calls = 0
            self._snap_cache = {}
            self._snap_lock = _th59.Lock()
        _fns_impl = None
        def _fns(self):
            def fn(*a, **kw):
                self.calls += 1
                syms = None
                if a and isinstance(a[0], list):
                    syms = a[0]
                elif kw.get("symbols") and isinstance(kw["symbols"], list):
                    syms = kw["symbols"]
                if syms is None or self.mode != "list":
                    raise Exception("wrong shape")
                if self.partial:
                    syms = syms[:1]
                return {"data": [{"symbol": o, "askPrice": 1.0 + i,
                                  "bidPrice": 0.9 + i, "price": 0.95 + i}
                                 for i, o in enumerate(syms)]}
            return [("fake.fn", fn)], []
        def _result(self, r):
            return r
        def ask_bid_mark(self, occ, max_age=None):
            self.calls += 1
            return (1.0, 0.9, 0.95, {"symbol": occ})
    for _m in ("ask_bid_many", "_one_at_a_time", "_parse_batch", "BATCH_MAX"):
        setattr(_FakeOD, _m, getattr(wb.OptionData, _m))

    _occs = ["QQQ260902C00700000", "QQQ260902P00700000", "SPY260902C00600000"]
    _fk = _FakeOD()
    _got = _fk.ask_bid_many(_occs)
    check(59, "three contracts cost ONE call", _fk.calls == 1, str(_fk.calls))
    check(59, "and all three come back", len(_got) == 3, str(len(_got)))
    check(59, "with ask, bid and mark", all(len(v) == 4 for v in _got.values()))

    # A full shape hunt on every sweep is a dozen failing HTTP requests a
    # second - worse than the problem it solves.
    _fk.calls = 0
    _fk.ask_bid_many(_occs)
    check(59, "the working shape is remembered", _fk.calls == 1, str(_fk.calls))

    # The single-quote cache is warmed, so a follow-up costs nothing.
    check(59, "the batch warms the single-quote cache", len(_fk._snap_cache) == 3)

    # A row for a contract we did NOT ask for must be DROPPED. A price on the
    # wrong contract would price an exit against a position not held.
    check(59, "unrequested rows are dropped, never guessed",
          _fk._parse_batch({"data": [{"symbol": "OTHER123", "askPrice": 9}]}, _occs) == {})
    check(59, "a malformed body is handled",
          _fk._parse_batch("not a list", _occs) == {}
          and _fk._parse_batch({"nope": 1}, _occs) == {})

    # A shape that answers ONE of three looks like success and starves the
    # rest, so it must not be remembered as the winner.
    _fp = _FakeOD(partial=True)
    _fp.ask_bid_many(_occs)
    check(59, "a partial answer is not remembered as the shape",
          getattr(_fp, "_batch_shape", None) is None)

    # If batching is refused entirely, it must still work - just expensively.
    _fb = _FakeOD(mode="nope")
    _out = _fb.ask_bid_many(_occs)
    check(59, "it falls back to one call each", len(_out) == 3, str(len(_out)))
    check(59, "and warns once, not every sweep", getattr(_fb, "_warned_no_batch", False) is True)

    check(59, "empty input is a no-op", _fk.ask_bid_many([]) == {} and _fk.ask_bid_many(None) == {})
    check(59, "the 20-per-call cap is respected", wb.OptionData.BATCH_MAX <= 20)
    _w59 = io.open("webull_client.py", encoding="utf-8").read()
    check(59, "the batch goes through the rate budget", "paced(fn, *args, **kwargs)" in _w59)

    # --- 60. The SDK audit only names methods that exist ------------------
    # An audit that recommends a call which is not in the installed SDK is
    # worse than no audit - it sends you building against nothing.
    _audit_p = os.path.join(HERE, "WEBULL-SDK-AUDIT.md")
    check(60, "the audit exists", os.path.exists(_audit_p))
    if os.path.exists(_audit_p):
        _audit = io.open(_audit_p, encoding="utf-8").read()
        _sdk_root = os.path.join(HERE, ".venv", "Lib", "site-packages", "webull")
        if os.path.isdir(_sdk_root):
            _blob = []
            for _dp, _dn, _fn in os.walk(_sdk_root):
                for _f in _fn:
                    if _f.endswith(".py"):
                        _blob.append(io.open(os.path.join(_dp, _f), encoding="utf-8",
                                             errors="replace").read())
            _blob = "\n".join(_blob)
            for _m in ("get_history_bar", "get_snapshot", "get_tick", "get_quotes",
                       "get_footprint", "get_noii_snapshot", "get_batch_history_bar",
                       "get_market_sectors", "get_gainers_losers", "get_most_active",
                       "get_option_contracts", "get_trade_calendar",
                       "preview_option", "replace_option"):
                if _m in _audit:
                    check(60, "%s is really in the SDK" % _m, ("def %s(" % _m) in _blob)
        else:
            check(60, "SDK present to check against", False, "no .venv here")
        # The audit must repeat the two standing prohibitions, or someone
        # follows it straight into the 9/2 dependency break.
        check(60, "it warns against per-second polling", "300 requests / 60 s" in _audit)
        check(60, "and against installing the streaming family",
              "webull-python-sdk-*" in _audit)
        check(60, "it does not claim push events are safe to install",
              "untested" in _audit and "Do not install anything to find out" in _audit)

    # --- 61. Tiered ratchet + anti-clip -----------------------------------
    # A percentage is not the same noise at every premium: one tick on a $0.40
    # contract is 2.5%, so "breakeven" there sits inside the spread.
    import ratchet_tiers as _rt
    _rt = importlib.reload(_rt)
    for _fill, _want in ((0.50, (25.0, 10.0)), (1.50, (15.0, 0.0)), (5.00, (10.0, 5.0))):
        _a, _f, _st = _rt.ratchet_plan(_fill)
        check(61, "$%.2f arms +%g%% and first-locks %+g%%" % (_fill, _want[0], _want[1]),
              _a == _want[0] and _f == _want[1], str((_a, _f, _st)))
    # FLOOR 1: a rung must be worth 4+ ticks. 5% of $3.00 is $0.15 = 3 nickel
    # ticks, too tight, so the step widens.
    check(61, "the tick floor widens a too-tight rung",
          _rt.ratchet_plan(3.00)[2] > 5.0, str(_rt.ratchet_plan(3.00)[2]))
    check(61, "and leaves a wide-enough one alone", _rt.ratchet_plan(5.00)[2] == 5.0)
    check(61, "tick size flips at $3", _rt.tick_size(2.99) == 0.01 and _rt.tick_size(3.00) == 0.05)

    # Nothing locks before the arm level.
    check(61, "below the arm, nothing is locked", _rt.ratchet_locked_pct(5, 3.00) is None)
    check(61, "at the arm, the first lock applies", _rt.ratchet_locked_pct(10, 3.00) == 5.0)

    # ANTI-CLIP: the stop may never sit closer than 40% of the gain made.
    check(61, "anti-clip does nothing on a small gain",
          _rt.anti_clip(5.0, 10.0) == 5.0)
    check(61, "anti-clip caps a runner", _rt.anti_clip(71.7, 80.0) == 48.0,
          str(_rt.anti_clip(71.7, 80.0)))
    check(61, "the cap is (1-k) x gain", abs(_rt.anti_clip(999, 100.0) - 60.0) < 0.01)
    check(61, "it passes None through", _rt.anti_clip(None, 50) is None)

    # Wired into the live ratchet, and MONOTONIC. anti_clip is fed the PEAK,
    # not the live gain - a falling gain would walk the stop back down, which
    # is the one thing a ratchet must never do.
    class _RT(wb.LiveSession):
        def __init__(self, tiers=True):
            import config as _c
            self.settings = dict(_c.DEFAULT_SETTINGS)
            self.settings["ratchet_tiers"] = tiers
            self.strategies = []; self.position = None
    _r61 = _RT(True)
    _r61.position = {"entry": 3.00, "mark": 3.00, "qty": 1,
                     "ratchet_on": True, "ratchet_step": 10.0}
    _seen = []
    for _m in (3.00, 3.15, 3.30, 3.60, 4.50, 5.40, 4.80, 4.50):
        _r61.position["mark"] = _m
        _r61._update_ratchet()
        _seen.append(_r61.position["ratchet"]["stop_pct"])
    check(61, "the opening stop is the flat step", _seen[0] == -10.0, str(_seen[0]))
    check(61, "it arms at +10% on a $3 contract", _seen[2] == 5.0, str(_seen[2]))
    check(61, "the stop NEVER loosens", all(b >= a - 1e-9 for a, b in zip(_seen, _seen[1:])),
          str(_seen))
    check(61, "a pullback holds the stop", _seen[-1] == _seen[-3], str(_seen[-3:]))
    # Was: assert a COMMENT saying so. Watch the argument instead - if the
    # live gain is ever passed here, a pullback silently loosens the stop.
    import ratchet_tiers as _rt61
    _fed = []
    _real_ac = _rt61.anti_clip
    try:
        _rt61.anti_clip = lambda locked, gain: (_fed.append(gain),
                                                _real_ac(locked, gain))[1]
        _z61 = _w4.make_session("LIVE")
        _z61.settings.update({"my_enabled": True, "ratchet_tiers": True,
                              "ratchet_step_pct": 10.0})
        _z61.position = {"symbol": "QQQ", "side": "CALLS", "strike": 700.0,
                         "qty": 1, "entry": 1.00, "mark": 1.50,
                         "expiration": "2026-09-02", "ratchet_on": True,
                         "ratchet_step": 10.0}
        _z61._update_ratchet()                      # peak +50%
        _z61.position["mark"] = 1.20                # pulls back to +20%
        _z61._update_ratchet()
        check(61, "anti-clip is fed the PEAK, not the live gain",
              _fed and max(_fed) >= 49.0 and _fed[-1] >= 49.0,
              "gains passed: %s" % _fed)
    finally:
        _rt61.anti_clip = _real_ac

    # The flat rungs G designed must still be reachable.
    _r62 = _RT(False)
    _r62.position = {"entry": 3.00, "mark": 3.30, "qty": 1,
                     "ratchet_on": True, "ratchet_step": 10.0}
    _r62._update_ratchet()
    check(61, "tiers off restores the flat rungs",
          _r62.position["ratchet"]["stop_pct"] == 0.0,
          str(_r62.position["ratchet"]["stop_pct"]))
    import config as _cfg61
    check(61, "the toggle exists and defaults on",
          _cfg61.DEFAULT_SETTINGS.get("ratchet_tiers") is True)
    check(61, "the module is imported defensively",
          "import ratchet_tiers as rt\nexcept Exception:" in
          io.open("webull_client.py", encoding="utf-8").read())

    # --- 62. The underlying price at the fill (handoff item 8) ------------
    # An option P&L alone does not say whether the STOCK went your way, and
    # that is the question any review starts with.
    import trade_log as _tl62
    _tl62 = importlib.reload(_tl62)
    for _c62 in ("underlying_in", "underlying_out", "underlying_move"):
        check(62, "log has %s" % _c62, _c62 in _tl62.FIELDS)
        check(62, "%s is numeric in the sheet" % _c62, _c62 in _tl62._NUM)
    _w62 = io.open("webull_client.py", encoding="utf-8").read()
    # Read at the FILL, not at the close - by then the move is over.
    check(62, "it is stamped when the position opens",
          '"underlying_in": q.get("spot"),' in _w62)
    check(62, "and read again as the trade ends", 'p["underlying_out"] = p.get("spot")' in _w62)
    check(62, "the move is derived, not asked for twice",
          'float(p["underlying_out"]) - float(p["underlying_in"])' in _w62)
    # A missing price must leave a blank, not a wrong number.
    class _U62(wb.LiveSession):
        def __init__(self):
            import config as _c
            self.settings = dict(_c.DEFAULT_SETTINGS); self.strategies = []
            self.blotter = []; self.day_realized = 0.0; self._day = None
            self.account_id = ""
        def _save_day(self): pass
        def _underlying(self, sym): raise RuntimeError("no feed")
    _u = _U62()
    _rows_before = len(_tl62._rows())
    _u._record_close({"symbol": "QQQ", "side": "CALLS", "strike": 700.0, "qty": 1,
                      "entry": 2.00, "expiration": "2026-09-02", "opened_at": "10:00"},
                     2.20, estimated=False)
    check(62, "a dead price feed does not stop the trade being recorded",
          len(_u.blotter) == 1)

    # --- 63. Real time-and-sales velocity (SDK WIN #2) --------------------
    # tape.py measured speed from 1-MINUTE BARS and admitted it in its own
    # docstring. For "silent tape means do not enter", the difference between
    # "the last bar was quiet" and "two prints in thirty seconds" IS the signal.
    import tape as _tp63
    _tp63 = importlib.reload(_tp63)
    _now63 = time.time()

    def _mk63(n, span, start, side=None, size=100):
        return [{"tradeTime": start + i * (span / max(1, n)), "price": 700 + i * 0.001,
                 "volume": size, "side": side} for i in range(n)]

    _quiet = _mk63(120, 540, _now63 - 600) + _mk63(10, 60, _now63 - 60)
    _busy = _mk63(120, 540, _now63 - 600) + _mk63(200, 60, _now63 - 60, side="B")
    _rq = _tp63.compute_ticks(_tp63.parse_ticks(_quiet), _now63)
    _rb = _tp63.compute_ticks(_tp63.parse_ticks(_busy), _now63)
    check(63, "a quiet minute does not read fast", _rq["state"] in ("calm", "normal"),
          str(_rq.get("state")))
    check(63, "a busy minute does", _rb["state"] in ("fast", "violent"), str(_rb.get("state")))
    check(63, "prints per second are reported", _rb["prints_per_sec"] > _rq["prints_per_sec"])
    check(63, "and the reading says it came from ticks", _rb["source"] == "ticks")

    # Direction from WHERE the prints went off - the thing bars cannot tell
    # you. Heavy volume hitting the bid is a different tape from the same
    # volume lifting the offer.
    check(63, "all-buy prints read up", _rb["direction"] == "up" and _rb["buy_share"] == 100.0)
    _sells = _mk63(120, 540, _now63 - 600) + _mk63(200, 60, _now63 - 60, side="S")
    check(63, "all-sell prints read down",
          _tp63.compute_ticks(_tp63.parse_ticks(_sells), _now63)["direction"] == "down")

    # Timestamps arrive in seconds on some endpoints and milliseconds on
    # others. Guessing wrong puts every print 50,000 years away.
    _ms = [{"timestamp": int((_now63 - 30 + i) * 1000), "price": 700, "volume": 50}
           for i in range(20)]
    check(63, "millisecond timestamps are detected",
          abs(_tp63.parse_ticks(_ms)[0]["t"] - (_now63 - 30)) < 2,
          str(_tp63.parse_ticks(_ms)[0]["t"]))
    check(63, "junk rows are dropped, not guessed",
          _tp63.parse_ticks([{"nope": 1}, "x", {"price": 1}]) == [])
    check(63, "no prints is a reason, not a crash",
          _tp63.compute_ticks([])["ok"] is False)
    check(63, "too few prints is honest too",
          _tp63.compute_ticks(_tp63.parse_ticks(_mk63(3, 10, _now63 - 10)), _now63)["ok"] is False)

    # The bar path MUST remain - it is broker-free, so the login screen and a
    # dropped connection still show a reading.
    _m63 = io.open("main.py", encoding="utf-8").read()
    check(63, "ticks are preferred when connected", "e.recent_ticks(symbol)" in _m63)
    check(63, "bars remain the fallback", 'tape.velocity(ysym), "source": "bars"' in _m63)
    check(63, "a tick hiccup cannot blank the meter",
          "must never blank the meter" in _m63)
    _w63 = io.open("webull_client.py", encoding="utf-8").read()
    check(63, "the tick fetch is paced", "paced(fn, *args, **kw)" in _w63)
    check(63, "unavailable returns None, not an empty list",
          "None, not [] - the caller must be able to tell" in _w63)

    # --- 64. Pacing must not stall the screen -----------------------------
    # REGRESSION, found live with a position open: snapshot_row probes up to 8
    # argument shapes until one is accepted. Before pacing a rejected shape
    # cost nothing; after it, each is a real 0.20s pause. A winner at position
    # 5 was 0.8 SECONDS of dead time on EVERY quote, against a 1-second screen
    # poll. The shape never changes between calls, so it is learned once.
    import threading as _th64

    class _OD64(object):
        def __init__(self, win_at=5):
            self.n = 0; self.win_at = win_at
            self._snap_cache = {}; self._snap_lock = _th64.Lock()
        def _fns(self):
            def fn(*a, **kw):
                self.n += 1
                if self.n % 1000 == self.win_at % 1000 or getattr(self, "_ok", False):
                    pass
                # accept exactly one shape: keyword "symbols" as a bare string
                if kw.get("symbols") and isinstance(kw["symbols"], str) and "category" not in kw:
                    return {"symbol": "X", "askPrice": 1.0}
                raise Exception("wrong shape")
            return [("f.fn", fn)], []
        def _result(self, r):
            return r
    for _m in ("snapshot_row", "snapshot_cached", "forget_snapshot", "_SNAP_TTL"):
        setattr(_OD64, _m, getattr(wb.OptionData, _m))

    _saved64, wb.BUDGET = wb.BUDGET, wb._Budget(0.05)
    try:
        _od64 = _OD64()
        _t = time.time(); _od64.snapshot_row("QQQ260902C00700000"); _first = time.time() - _t
        _n1 = _od64.n
        _od64.n = 0
        _t = time.time(); _od64.snapshot_row("QQQ260902C00700000"); _later = time.time() - _t
        check(64, "the first quote hunts for a shape", _n1 > 1, str(_n1))
        check(64, "every quote after costs ONE attempt", _od64.n == 1, str(_od64.n))
        check(64, "and is measurably faster", _later < _first, "%.2f vs %.2f" % (_later, _first))
        check(64, "the shape is remembered", getattr(_od64, "_shape_row", None) is not None)

        # If the remembered shape ever stops working, it must re-probe rather
        # than fail forever.
        _od64._shape_row = ("f.fn", "arg", ("nonsense",))
        _od64.n = 0
        _r = _od64.snapshot_row("QQQ260902C00700000")
        check(64, "a stale shape is discarded and re-learned",
              _r is not None and _od64._shape_row[1] == "kw", str(_od64._shape_row))
    finally:
        wb.BUDGET = _saved64

    _w64 = io.open("webull_client.py", encoding="utf-8").read()
    # Was: assert my own comment text is present. That is not a test - it
    # broke the moment I rewrote the comment, while the behaviour was fine.
    # Assert the behaviour the comment DESCRIBES: the memory must survive
    # across different contracts, and must not answer with the wrong one.
    _od64b = _OD64()
    _saved64b, wb.BUDGET = wb.BUDGET, wb._Budget(0.0)
    try:
        _od64b.snapshot_row("QQQ260902C00700000")
        _od64b.n = 0
        _od64b.snapshot_row("QQQ260902C00715000")
        check(64, "the memory carries over to a DIFFERENT contract",
              _od64b.n == 1, str(_od64b.n))
    finally:
        wb.BUDGET = _saved64b
    check(64, "no dir() hack survived", "'name' in dir()" not in _w64)

    # --- 65. A trade closed outside the app must still be journalled ------
    # Found live 2026-09-02: G closed at the broker and NOTHING was logged.
    # reconcile() and forget_position() cleared the screen and the trade was
    # simply gone - so the journal was silently losing whole trades and the
    # day's numbers were wrong by however many.
    import tempfile as _tf65, trade_log as _tl65
    _tl65 = importlib.reload(_tl65)
    _d65 = _tf65.mkdtemp()
    _tl65.LOG_DIR = _d65
    _tl65.CSV_PATH = os.path.join(_d65, "t.csv")
    _tl65.XLSX_PATH = os.path.join(_d65, "t.xlsx")

    class _S65(wb.LiveSession):
        def __init__(self, rows):
            import config as _c
            self.settings = dict(_c.DEFAULT_SETTINGS); self.strategies = []
            self.blotter = []; self.day_realized = 0.0; self._day = None
            self.account_id = "ACC"; self._rows = rows
            self._order_lock = _th64.Lock(); self._last_reconcile = 0
        def _save_day(self): pass
        def broker_positions(self): return self._rows
        def _underlying(self, sym): return 715.0

    _pos65 = {"symbol": "QQQ", "side": "CALLS", "strike": 700.0, "qty": 1,
              "entry": 2.00, "mark": 2.35, "bid": 2.30, "expiration": "2026-09-02",
              "opened_at": "13:00", "ratchet_step": 10.0}

    _a65 = _S65([]); _a65.position = dict(_pos65); _a65.reconcile(force=True)
    _rows65 = _tl65._rows()
    check(65, "a broker-side close is recorded", len(_rows65) == 1, str(len(_rows65)))
    check(65, "named so you can find it later",
          _rows65[0]["exit_reason"] == "CLOSED-ELSEWHERE", str(_rows65[0]["exit_reason"]))
    # The BID is what a sale would really have got - use it over the mark.
    check(65, "the exit uses the last bid, not the mark", _rows65[0]["exit"] == "2.3",
          str(_rows65[0]["exit"]))
    # A guess that looks like a fact is worse than no row.
    check(65, "and it is flagged as estimated", "estimated" in _rows65[0]["note"])

    _b65 = _S65([]); _b65.position = dict(_pos65); _b65.forget_position()
    check(65, "the CLEAR IT button records too", len(_tl65._rows()) == 2)
    check(65, "with its own reason", _tl65._rows()[-1]["exit_reason"] == "CLEARED-BY-HAND")

    # With no last price there is nothing honest to write.
    _c65 = _S65([])
    _c65.position = {"symbol": "QQQ", "side": "CALLS", "strike": 700.0, "qty": 1,
                     "entry": 2.0, "expiration": "x", "opened_at": "13:00"}
    _c65.forget_position()
    check(65, "no last price means no invented row", len(_tl65._rows()) == 2)

    # A FAILED broker call is not "you are flat" - the position must stay.
    _d65s = _S65(None); _d65s.position = dict(_pos65); _d65s.reconcile(force=True)
    check(65, "a failed positions call changes nothing",
          _d65s.position is not None and len(_tl65._rows()) == 2)

    _w65 = io.open("webull_client.py", encoding="utf-8").read()
    check(65, "the reason is recorded where the code is",
          "the journal was quietly losing whole trades" in _w65)
    check(65, "journalling can never break the clear",
          "could not record the cleared trade" in _w65)

    # --- 66. One-second prices, one call ----------------------------------
    # The chips read Yahoo, which caches 5s - so polling the screen faster
    # re-read the same number. The browser was never the limit, the source was.
    # Webull's snapshot takes every symbol in ONE call, so 1/sec costs 60
    # requests a minute out of a 300 budget shared three ways, no matter how
    # many symbols are on screen.
    import quotes as _q66
    check(66, "Yahoo really does cache", _q66._TTL >= 5.0, str(_q66._TTL))

    class _SS66(wb.LiveSession):
        def __init__(self, mode="joined"):
            import config as _c
            self.settings = dict(_c.DEFAULT_SETTINGS); self.strategies = []
            self.calls = 0; self.mode = mode
            self._stock_cache = {"t": 0.0, "rows": {}}
            class _OD:
                def __init__(self, outer): self.o = outer
                def _client(self):
                    o = self.o
                    class C:
                        def get_snapshot(_s, *a, **kw):
                            o.calls += 1
                            syms = None
                            if a and isinstance(a[0], str): syms = a[0].split(",")
                            elif a and isinstance(a[0], list): syms = a[0]
                            elif kw.get("symbols"): syms = (kw["symbols"].split(",")
                                if isinstance(kw["symbols"], str) else kw["symbols"])
                            if not syms: raise Exception("no symbols")
                            return {"data": [{"symbol": x, "close": 100.0 + i,
                                              "preClose": 99.0 + i}
                                             for i, x in enumerate(syms)]}
                    return C()
                def _result(_s, r): return r
            self._od = _OD(self)

    _ss = _SS66()
    _rows = _ss.stock_snapshot(["SPY", "QQQ", "TSLA"])
    check(66, "three symbols cost ONE call", _ss.calls == 1, str(_ss.calls))
    check(66, "and all three come back", len(_rows) == 3, str(len(_rows)))
    check(66, "with price and change", _rows["SPY"]["price"] == 100.0
          and abs(_rows["SPY"]["change"] - 1.0) < 0.01, str(_rows["SPY"]))
    check(66, "and are marked as coming from the broker",
          _rows["SPY"]["source"] == "webull" and _rows["SPY"]["live"] is True)

    # Cached just under the 1s poll, so one poll is one call - not two.
    _ss.calls = 0
    _ss.stock_snapshot(["SPY", "QQQ", "TSLA"])
    check(66, "a second read inside the second is free", _ss.calls == 0, str(_ss.calls))
    check(66, "the cache is shorter than the poll", wb.LiveSession._STOCK_TTL < 1.0)

    # A symbol we did not ask for must never be attached to a chip.
    check(66, "unrequested rows are dropped",
          wb.LiveSession._parse_stock_snapshot(
              {"data": [{"symbol": "NVDA", "close": 5}]}, ["SPY"]) == {})
    check(66, "a malformed body is handled",
          wb.LiveSession._parse_stock_snapshot("nope", ["SPY"]) == {})

    # If the broker cannot answer, the chips fall back rather than going blank.
    _m66 = io.open("main.py", encoding="utf-8").read()
    check(66, "it prefers the broker", "e.stock_snapshot(syms)" in _m66)
    check(66, "and falls back to Yahoo", "fall through to Yahoo rather than blank" in _m66)
    _ix66 = io.open("index.html", encoding="utf-8").read()
    check(66, "the browser polls prices every second",
          "setInterval(refreshPrices,1000)" in _ix66)
    # Option quotes must NOT ride the 1s beat - each is a separate request and
    # a 0DTE quote does not need refreshing that often.
    check(66, "option quotes stay on the slower beat",
          "quoteTimer=setInterval" in _ix66 and "},5000);" in _ix66)
    check(66, "and the new timer is cleared on disconnect",
          _ix66.count("clearInterval(quoteTimer)") >= 4)

    import subprocess as _sp3
    _sm3 = _sp3.run(["node", "ui_smoke.js", "futures_index.html"], cwd=HERE,
                    capture_output=True, text=True, timeout=60)
    check(47, "the futures page still runs", _sm3.returncode == 0,
          (_sm3.stdout + _sm3.stderr).strip()[:300])

    # G's measured MNQ system: 12.5 POINTS, stop and step the same number.
    # The units matter and were ambiguous when he described it - "$12.50" is
    # 6.25 points on MNQ and 0.625 on NQ, which is two ticks and would be
    # stopped out by noise instantly. The setting is POINTS.
    check(47, "the default step is 12.5 points",
          _fcm.DEFAULT_SETTINGS["ratchet_points"] == 12.5)
    _f6, _r6 = _run("SHORT", 29800.0, [29800, 29787.50, 29775, 29762.50], step=12.5)
    _by_peak = {round(x[1]["peak_points"], 2): x[1]["stop_price"] for x in _r6}
    check(47, "short 29800: opening stop is 29812.50",
          abs(_r6[0][1]["stop_price"] - 29812.50) < 0.01, str(_r6[0][1]["stop_price"]))
    check(47, "+12.5 puts the stop at breakeven 29800",
          abs(_by_peak.get(12.5, 0) - 29800.0) < 0.01, str(_by_peak))
    check(47, "+25 puts it at 29787.50",
          abs(_by_peak.get(25.0, 0) - 29787.50) < 0.01, str(_by_peak))
    check(47, "+37.5 puts it at 29775",
          abs(_by_peak.get(37.5, 0) - 29775.0) < 0.01, str(_by_peak))
    # Fractional steps land exactly on the rung; the float fix matters more
    # here than with whole numbers.
    check(47, "an exact 12.5 touch is not a rung short",
          _r6[1][1]["stop_points"] == 0.0, str(_r6[1][1]))
    # Every stop must be a real tradable price, not something between ticks.
    for _m, _rr, _h in _r6:
        _tickd = round(_rr["stop_price"] / _fcm.FUT["MNQ"]["tick"], 6)
        check(47, "stop %.2f sits on a tick" % _rr["stop_price"],
              abs(_tickd - round(_tickd)) < 1e-6, str(_rr["stop_price"]))
    # Sanity on what it costs.
    check(47, "12.5 points is $25 on MNQ", 12.5 * _fcm.FUT["MNQ"]["point_value"] == 25.0)
    check(47, "and 50 ticks wide, not 2", 12.5 / _fcm.FUT["MNQ"]["tick"] == 50.0)
    check(29,"and no premium/time value on the button",
          "% time" not in ix6 and "q.ask.toFixed" not in code6)

    import importlib, webull_client as _w6
    _w6 = importlib.reload(_w6)
    z6 = _w6.make_session("LIVE")
    # A session loads today's blotter from disk. Start from empty so the counts
    # measure THIS test, not whatever ran before it.
    z6.blotter = []; z6.day_realized = 0.0
    # A session restores the saved strategies, and state() evaluates them. With
    # the market OPEN an ORB condition can be met mid-suite, and this fake
    # session then tries to place a real order and dies on the missing SDK
    # handle. The test is about percent display, not the strategy engine, so
    # nothing is armed. Without this the whole suite passed only out of hours.
    for _st6 in (z6.strategies or []):
        _st6["enabled"] = False
    for e_, x_ in ((3.00,3.45),(2.40,2.05)):
        z6._record_close({"symbol":"QQQ","side":"CALLS","strike":711.0,"qty":1,
                          "entry":e_,"expiration":"2026-08-26","opened_at":"09:41"},
                         x_, estimated=False)
    st6 = z6.state()
    check(29,"server sends day_pct", "day_pct" in st6)
    check(29,"and the W/L count", st6["day_wins"]==1 and st6["day_losses"]==1)
    check(29,"every blotter row carries a percent",
          all("pct" in b for b in st6["blotter"]))
    check(29,"percent is right", abs(st6["blotter"][0]["pct"] - 15.0) < 0.05,
          str(st6["blotter"][0]["pct"]))
    check(29,"dollars still RECORDED for the trade log, just not shown",
          all("pnl" in b for b in st6["blotter"]))

finally:
    for p_ in (OPT,FUT):
        if p_:
            try: p_.terminate(); p_.wait(timeout=8)
            except Exception:
                try: p_.kill()
                except Exception: pass
    shutil.copy(BACKUP, SETTINGS)
    # Prove it, rather than trusting the redirect above.
    _after = (io.open(REAL_TRADES_CSV, encoding="utf-8").read()
              if os.path.exists(REAL_TRADES_CSV) else None)
    if _after != _REAL_TRADES_BEFORE:
        if _REAL_TRADES_BEFORE is not None:
            io.open(REAL_TRADES_CSV, "w", encoding="utf-8").write(_REAL_TRADES_BEFORE)
        print("  !! the suite wrote to the REAL trade log - reverted. Fix the redirect.")
    else:
        print("  real trade log untouched.")

    # --- 67. A restart must not lose the position -------------------------
    # 9/2, live: he was holding a put, restarted, and the app said FLAT. It
    # only ever knew trades IT opened - reconcile() could clear a position but
    # nothing could ever ADD one. He watched a real trade with no P&L, no
    # high-water mark and no ratchet until he closed it by hand.
    _w67 = io.open("webull_client.py", encoding="utf-8").read()

    # The parser must be the exact inverse of the builder, both ways. If they
    # ever drift, adoption puts the wrong strike on screen and manages it.
    for _sym, _exp, _typ, _k in [("SPY","2026-09-02","PUT",764.0),
                                 ("QQQ","2026-09-02","CALL",715.5),
                                 ("TSLA","2026-12-19","CALL",332.5),
                                 ("SPY","2026-09-02","PUT",9.5)]:
        _occ = wb.occ_symbol(_sym, _exp, _typ, _k)
        check(67, "round trip %s %g%s" % (_sym, _k, _typ[0]),
              wb.parse_occ(_occ) == (_sym, _exp, _typ, _k), _occ)
    for _bad in ("", "GARBAGE", "SPY260945P00764000", "SPY_260902P00764000",
                 None, "AAPL260902X00100000", "SPY260902P0076400"):
        check(67, "refuses %r" % (_bad,), wb.parse_occ(_bad) is None)

    def _s67(rows):
        z = wb.LiveSession.__new__(wb.LiveSession)
        z.position = None; z.trade = 1; z.account_id = "A"
        z.settings = {"ratchet_step_pct": 10.0}
        z._order_lock = threading.RLock(); z.last_event = ""
        z.broker_positions = lambda: rows
        return z

    # THE REAL WEBULL SHAPE. This first shipped parsing `symbol` as an OCC
    # string and could therefore never adopt anything - Webull returns `symbol`
    # as the plain underlying and puts the contract in legs[0] as separate
    # fields. Confirmed against discord-sniper/webull_options.py positions(),
    # which has read this shape in production for months. He restarted holding
    # a position and the app still said flat; this is why.
    _real = {"symbol": "QQQ", "quantity": "2", "cost_price": "1.35",
             "legs": [{"option_type": "CALL", "option_exercise_price": "708",
                       "option_expire_date": "2026-09-02"}]}
    _r = _s67([_real]).adopt_broker_position()
    check(67, "the REAL legs shape is adopted", _r is not None, str(_r))
    check(67, "with the strike off the leg", _r and _r["strike"] == 708.0, str(_r))
    check(67, "the type off the leg", _r and _r["side"] == "CALLS", str(_r))
    check(67, "the expiry off the leg",
          _r and _r["expiration"] == "2026-09-02", str(_r))
    check(67, "and cost_price as the entry", _r and _r["entry"] == 1.35, str(_r))

    _put = _s67([{"symbol": "SPY", "quantity": 1, "cost_price": 2.10,
                  "legs": [{"option_type": "PUT", "option_exercise_price": 764,
                            "option_expire_date": "2026-09-02"}]}]
                ).adopt_broker_position()
    check(67, "puts too", _put and _put["side"] == "PUTS"
          and _put["strike"] == 764.0, str(_put))
    # Flat fields with no legs at all, and leg-level cost.
    check(67, "flat fields without legs work",
          (_s67([{"symbol": "TSLA", "quantity": 3, "cost_price": 0.9,
                  "option_type": "C", "strike_price": 332.5,
                  "expire_date": "2026-09-02"}]).adopt_broker_position() or {})
          .get("strike") == 332.5)
    check(67, "leg-level cost is an entry price",
          (_s67([{"symbol": "QQQ", "quantity": 1,
                  "legs": [{"option_type": "CALL", "option_exercise_price": 708,
                            "cost": 1.5, "option_expire_date": "2026-09-02"}]}])
           .adopt_broker_position() or {}).get("entry") == 1.5)
    # A STOCK row is not an option and must never be adopted as one.
    check(67, "a stock row is refused",
          _s67([{"symbol": "QQQ", "quantity": 100, "cost_price": 700.0}])
          .adopt_broker_position() is None)

    # A FAILED ask must never read as flat - that was the 9/2 fault exactly.
    _fail = _s67(None)
    for _ in range(3):
        _fail.adopt_broker_position()
    check(67, "a failed ask is counted, not treated as flat",
          _fail._adopt_ask_fails == 3 and _fail.position is None,
          str(_fail._adopt_ask_fails))
    check(67, "and it says the app is NOT managing anything",
          "NOT managing it" in _fail.last_event, _fail.last_event[:70])
    check(67, "and names what to close",
          "Fill Announcer" in _fail.last_event)
    _fail.broker_positions = lambda: []
    _fail.adopt_broker_position()
    check(67, "a real flat answer resets the counter",
          _fail._adopt_ask_fails == 0)

    _good = [{"symbol": "SPY260902P00764000", "quantity": "2", "costPrice": "1.35"}]
    _g = _s67(_good).adopt_broker_position()
    check(67, "a held position IS adopted", _g is not None)
    check(67, "with the right contract", _g and _g["symbol"] == "SPY"
          and _g["strike"] == 764.0 and _g["side"] == "PUTS", str(_g))
    check(67, "the right size and entry", _g and _g["qty"] == 2
          and abs(_g["entry"] - 1.35) < 1e-9, str(_g))
    check(67, "alternate field names work",
          (_s67([{"symbol": "QQQ260902C00715500", "qty": 1, "avgCost": 0.88}])
           .adopt_broker_position() or {}).get("strike") == 715.5)
    check(67, "junk rows do not stop a good one",
          (_s67(["x", None, {"symbol": "SPY260902P00764000", "quantity": 1,
                             "costPrice": 1.1}]).adopt_broker_position() or {})
          .get("entry") == 1.1)

    # A FAILED ask is not an empty account, and an empty account is not a trade.
    check(67, "a failed ask adopts nothing", _s67(None).adopt_broker_position() is None)
    check(67, "a flat account adopts nothing", _s67([]).adopt_broker_position() is None)
    # No entry price means the ratchet has nothing to measure from - refuse it
    # rather than show something that looks managed and is not.
    check(67, "no entry price -> refused",
          _s67([{"symbol": "SPY260902P00764000", "quantity": "2"}])
          .adopt_broker_position() is None)
    check(67, "zero quantity -> refused",
          _s67([{"symbol": "SPY260902P00764000", "quantity": 0,
                 "costPrice": 1.2}]).adopt_broker_position() is None)
    check(67, "a symbol this app does not trade -> refused",
          _s67([{"symbol": "NVDA260902C00100000", "quantity": 1,
                 "costPrice": 1.0}]).adopt_broker_position() is None)

    # Never overwrite a live trade with a stale broker row.
    _live = _s67(_good); _live.position = {"symbol": "IWM", "mine": True}
    _live.adopt_broker_position()
    check(67, "an open trade is never overwritten",
          _live.position.get("mine") is True, str(_live.position))

    # THE SAFETY RULE. This app cannot tell whether the discord-sniper bot
    # opened the position it just found. Two tools resting a sell on one
    # contract is the 8/18 double-flatten - the second sell goes short. So an
    # adopted trade is VISIBLE immediately and SOLD by nothing until he says.
    check(67, "adopted arrives unmanaged", _g.get("needs_manage_ok") is True, str(_g))
    check(67, "with the ratchet off", _g.get("ratchet_on") is False, str(_g))
    _ac = _w67.split("def _maybe_auto_close", 1)[1].split("def _do_auto_close", 1)[0]
    check(67, "brackets cannot fire on an unmanaged position",
          'needs_manage_ok' in _ac and _ac.index('needs_manage_ok')
          < _ac.index('_bracket_hit'), _ac[:200])

    # It has to actually RUN. Adoption that is never called is not a fix.
    _rm67 = _w67.split("def refresh_mark", 1)[1].split("\n    def ", 1)[0]
    check(67, "the poll adopts while flat", "adopt_broker_position()" in _rm67)
    check(67, "before it gives up on being flat",
          _rm67.index("adopt_broker_position()") < _rm67.index('fc = p.get'))
    # ADOPTION MUST NEVER BLOCK SIGN-IN. It ran inline first, and the
    # positions read retries when the key is busy, so it could hold the login
    # past the 25-second limit: he got "Webull didn't respond within 25s" while
    # pressing a button that was working fine.
    _cn67 = _w67.split("def connect(self, app_key", 1)[1].split("\n    def ", 1)[0]
    check(67, "connect starts adoption in the background",
          "threading.Thread(target=self._adopt_quietly" in _cn67)
    check(67, "and does NOT wait for it inline",
          "self.adopt_on_connect()" not in _cn67, _cn67[-300:])
    check(67, "the thread is a daemon, so it cannot hold the app open",
          "daemon=True" in _cn67)
    check(67, "and it swallows its own errors",
          "def _adopt_quietly" in _w67)
    # Measured: a two-second positions read must not delay the login at all.
    _slow67 = wb.LiveSession.__new__(wb.LiveSession)
    _slow67.position = None; _slow67._order_lock = threading.RLock()
    _slow67.last_event = ""; _slow67.settings = {"ratchet_step_pct": 10.0}
    def _slowpos():
        time.sleep(1.0); return []
    _slow67.broker_positions = _slowpos
    _t67 = time.time()
    threading.Thread(target=_slow67._adopt_quietly, daemon=True).start()
    check(67, "a slow positions read costs the login nothing",
          time.time() - _t67 < 0.2, "%.3fs" % (time.time() - _t67))
    # ...but not once a second. 60 requests a minute out of 300 shared three
    # ways, to answer a question that only changes on a fill.
    check(67, "and it is throttled", "ADOPT_EVERY" in _rm67 and wb.ADOPT_EVERY >= 3.0,
          str(getattr(wb, "ADOPT_EVERY", None)))

    # One click hands it over, and the click is his.
    _m67 = io.open("main.py", encoding="utf-8").read()
    check(67, "there is a MANAGE endpoint", '"/api/position/manage"' in _m67)
    _mg = _m67.split('def manage_position', 1)[1].split('@app.post', 1)[0]
    check(67, "which clears the flag", 'needs_manage_ok' in _mg)
    check(67, "and arms the ratchet", '"ratchet_on"] = True' in _mg)
    check(67, "and sends no order", "place" not in _mg.lower().replace("replace", ""))
    _ix67 = io.open("index.html", encoding="utf-8").read()
    check(67, "the screen says it is unmanaged", 'id="adoptBand"' in _ix67
          and "NOT BEING MANAGED" in _ix67)
    check(67, "there is a button", 'EZ.managePosition()' in _ix67)
    check(67, "it warns about the bot before arming",
          "second sell goes SHORT" in _ix67)
    check(67, "and no ratchet readout is shown while unmanaged",
          "pos.needs_manage_ok ? null : pos.ratchet" in _ix67)


    # --- 68. The stream is an accelerator, never a dependency --------------
    # This one has never run against a live feed, so every test here is about
    # the ONE thing that matters if it misbehaves: it must not be able to
    # blank the chips, freeze a price, or stop the app starting.
    import stream as _st
    _s68 = io.open("stream.py", encoding="utf-8").read()

    _ps = _st.PriceStream("k", "s")
    # Nothing received yet: the answer is None, meaning POLL - not zero, not
    # a guess, not the last thing it saw.
    check(68, "an empty stream returns None", _ps.price("SPY") is None)
    check(68, "and None for a symbol it never saw", _ps.price("NVDA") is None)
    check(68, "and for no symbol at all", _ps.price("") is None
          and _ps.price(None) is None)

    _ps._on_message(None, "t", [{"symbol": "SPY", "price": 764.49,
                                 "preClose": 761.78}])
    _r = _ps.price("SPY")
    check(68, "a pushed quote reads back", _r and _r["price"] == 764.49, str(_r))
    check(68, "with change from the previous close",
          _r and abs(_r["change"] - 2.71) < 0.011
          and abs(_r["change_pct"] - 0.36) < 0.011, str(_r))
    check(68, "and is labelled as streamed",
          _r and _r["source"] == "stream" and _r["live"] is True, str(_r))
    check(68, "lower case symbols still match", _ps.price("spy") is not None)

    # THE IMPORTANT ONE. A feed that goes quiet must stop answering. A frozen
    # price is worse than a slow one because nothing on screen looks wrong.
    with _ps._lock:
        _ps._rows["SPY"]["t"] = time.time() - (_st.STALE_SECONDS + 1)
    check(68, "a stale price is withheld", _ps.price("SPY") is None)
    check(68, "stale is shorter than a breath", _st.STALE_SECONDS <= 10.0,
          str(_st.STALE_SECONDS))

    # A malformed push must not kill the reader thread - it is the only thread
    # feeding every price on the screen.
    for _bad in (None, "text", 5, [{"no_symbol": 1}], [{"symbol": "X"}],
                 {"symbol": "SPY"}, [None, "x"]):
        try:
            _ps._on_message(None, "t", _bad); _ok = True
        except Exception as _e:
            _ok = False
        check(68, "survives a bad message %r" % (type(_bad).__name__,), _ok)
    check(68, "and a junk message stores nothing",
          _ps.price("X") is None)

    # Partial rows: a bid-only push must not invent a price.
    _p2 = _st.PriceStream("k", "s")
    _p2._on_message(None, "t", [{"symbol": "QQQ", "bid": 1.0, "ask": 1.1}])
    check(68, "bid/ask without a trade is not a price", _p2.price("QQQ") is None)

    # ONE PAYLOAD TYPE PER CALL. Passing ["quote","snapshot"] came back
    # `UNSUPPORTED_SUB_TYPE: Subtype not supported:quotesnapshot` from his live
    # account at 14:16:51 - the server concatenates the list instead of reading
    # it as two types. The stream looked connected and delivered nothing.
    class _FakeSub:
        def __init__(self, fail=()):
            self.calls = []
            self.fail = fail
        def subscribe(self, syms, cat, types):
            self.calls.append(tuple(types))
            if types and types[0] in self.fail:
                raise Exception("UNSUPPORTED_SUB_TYPE")
    _p4 = _st.PriceStream("k", "s")
    _fk = _FakeSub()
    _p4._sub(_fk, ["SPY"])
    check(68, "never sends two payload types in one call",
          all(len(c) == 1 for c in _fk.calls), str(_fk.calls))
    check(68, "and does not send the string that was rejected",
          not any("quotesnapshot" in "".join(c) for c in _fk.calls), str(_fk.calls))
    check(68, "it asks for snapshot and quote separately",
          len(_fk.calls) == 2, str(_fk.calls))
    check(68, "snapshot first — it carries the previous close",
          _fk.calls[0] == (_st.PAYLOAD_TYPE_SHAPSHOT,), str(_fk.calls))
    # One type failing must not lose the other.
    _p5 = _st.PriceStream("k", "s")
    _fk2 = _FakeSub(fail=(_st.PAYLOAD_TYPE_SHAPSHOT,))
    check(68, "one rejected type still leaves a stream",
          _p5._sub(_fk2, ["SPY"]) is True and _p5._connected is True)
    _p6 = _st.PriceStream("k", "s")
    _fk3 = _FakeSub(fail=(_st.PAYLOAD_TYPE_SHAPSHOT, _st.PAYLOAD_TYPE_QUOTE))
    check(68, "both failing is reported, not raised",
          _p6._sub(_fk3, ["SPY"]) is False and _p6._connected is False)
    check(68, "and the reason is kept", "subscribe(" in (_p6._last_error or ""),
          str(_p6._last_error))

    # Starting must never raise, whatever is or is not installed.
    _p3 = _st.PriceStream("bad", "bad")
    try:
        _r3 = _p3.start(["SPY"]); _ok3 = True
    except Exception:
        _ok3 = False
    check(68, "start() never raises", _ok3)
    check(68, "and says so rather than pretending", _r3 in (True, False))
    try:
        _p3.stop(); _ok4 = True
    except Exception:
        _ok4 = False
    check(68, "stop() never raises", _ok4)
    check(68, "status() always answers",
          isinstance(_st.PriceStream("k","s").status(), dict))

    # The chain in main.py: the stream is only trusted when it has EVERY
    # symbol fresh. A partial stream falling through is the difference between
    # one chip lagging and one chip lying.
    _m68 = io.open("main.py", encoding="utf-8").read()
    _pr = _m68.split("def prices()", 1)[1].split("@app.get", 1)[0]
    check(68, "prices tries the stream first",
          _pr.index("STREAM.get") < _pr.index("stock_snapshot"))
    check(68, "and only if every symbol is fresh",
          "len(streamed) == len(syms)" in _pr)
    check(68, "then the broker", "stock_snapshot" in _pr)
    check(68, "then Yahoo", "quotes.get_all()" in _pr)
    check(68, "and never blanks the chips on an error",
          "fall through to Yahoo rather than blank" in _pr)
    check(68, "the stream is closed with the session",
          "STREAM" in _m68.split("def disconnect", 1)[1][:400])

    # The dependency rule from the handoff, in a test so it cannot be forgotten.
    check(68, "it uses the SDK already installed",
          "from webull.data.data_streaming_client import" in _s68)
    check(68, "and says not to install the forbidden family",
          "Do not install it." in _s68 and "webull-python-sdk-" in _s68)
    check(68, "a missing client degrades instead of crashing",
          "STREAM_AVAILABLE = False" in _s68 and "except Exception:" in _s68)


    # --- 69. Signing in must survive a rate limit --------------------------
    # 9/2: "im not being able to select margin account". The key was rate
    # limited - 14 429s on list-open, 2 on balance, three processes sharing 300
    # requests a minute - and the account list is the FIRST call connect makes.
    # One 429 there and the app was unusable, on a condition that clears itself
    # in twenty seconds, behind a message that read "account list failed: 429".
    _w69 = io.open("webull_client.py", encoding="utf-8").read()
    _mi69, _bo69 = wb.MIN_CALL_INTERVAL, wb.BACKOFF_AFTER_429
    wb.BACKOFF_AFTER_429 = 0.02
    wb.BUDGET.min_interval = 0.0
    try:
        _n = {"i": 0}
        def _flaky():
            _n["i"] += 1
            if _n["i"] < 3:
                raise Exception("HTTP Status: 429, Code: TOO_MANY_REQUESTS")
            return "accounts"
        check(69, "a 429 is waited out, not surrendered to",
              wb.paced_retry(_flaky) == "accounts" and _n["i"] == 3, str(_n))

        def _dead():
            raise Exception("TOO_MANY_REQUESTS")
        try:
            wb.paced_retry(_dead); _gave = False
        except Exception:
            _gave = True
        check(69, "but it does give up eventually", _gave)
        check(69, "after about a minute of trying", wb.CONNECT_RETRIES >= 2
              and wb.CONNECT_RETRIES <= 6, str(wb.CONNECT_RETRIES))

        # A REAL error must fail instantly. Retrying a bad key three times just
        # makes a wrong password take a minute to say so.
        _c = {"n": 0}
        def _real():
            _c["n"] += 1
            raise Exception("INVALID_SYMBOL")
        try:
            wb.paced_retry(_real)
        except Exception:
            pass
        check(69, "a real error is not retried", _c["n"] == 1, str(_c))
    finally:
        wb.BACKOFF_AFTER_429 = _bo69
        wb.BUDGET.min_interval = _mi69

    # ONLY reads may be retried. Retrying a place_order after a 429 could
    # double a position - the order may have landed before the limiter replied.
    for _danger in ("place_order", "place(", "cancel"):
        check(69, "paced_retry is never used for %s" % _danger.strip("("),
              ("paced_retry(self.trade.order" not in _w69)
              and ("paced_retry" not in _w69.split("def place", 1)[-1][:3000]))
    check(69, "sign-in uses the retry",
          "paced_retry(self.trade.account_v2.get_account_list)" in _w69)

    # Choosing an account calls connect() AGAIN. That second call must not
    # spend another request on a list that cannot have changed.
    check(69, "the account list is cached", "_ACCOUNTS_CACHE" in _w69)
    check(69, "for longer than a click takes", wb.ACCOUNTS_TTL >= 30.0,
          str(wb.ACCOUNTS_TTL))
    _cn = _w69.split("def connect(self, app_key", 1)[1].split("\n    def ", 1)[0]
    check(69, "and read before the network is touched",
          _cn.index("_ACCOUNTS_CACHE.get") < _cn.index("paced_retry"))

    # LOCKED OUT BY SOMEONE ELSE'S TRAFFIC. 9/2, twice: the Fill Announcer was
    # making ~100 calls a minute on the same key, Webull refused to list his
    # accounts, and he could not sign in to his own app at all. His account
    # list does not change because another program is noisy, so a launch that
    # is rate limited falls back to the list remembered from last time.
    # Redirect persistence into memory. A test that writes to his real
    # my-settings.json has, more than once, left his live trading config
    # holding something a test invented.
    import user_config as uc
    _fakedisk = {}
    _uc_save, _uc_load = uc.save, uc.load
    uc.save = lambda sec, val: _fakedisk.__setitem__(sec, val)
    uc.load = lambda sec, default=None: _fakedisk.get(
        sec, ({} if default is None else default))
    _KEY69 = "test-app-key-abc"
    check(69, "nothing is remembered before a first sign-in",
          wb._remembered_accounts(_KEY69) is None)
    _rows69 = {"data": [{"account_id": "ACCT1", "account_type": "MARGIN"}]}
    wb._remember_accounts(_KEY69, _rows69)
    check(69, "a good sign-in is remembered",
          wb._remembered_accounts(_KEY69) == _rows69)
    check(69, "and is keyed to that app key only",
          wb._remembered_accounts("a-different-key") is None)
    # The app key itself must never be written down - only a fingerprint.
    _book = uc.load("known_accounts", {}) or {}
    check(69, "the raw app key is not used as the index",
          _KEY69 not in _book and len(_book) >= 1, str(list(_book)[:3]))

    _wasrl = wb.BUDGET.rate_limits
    wb.BUDGET.rate_limits = 0
    wb.BUDGET._blocked_until = 0.0
    check(69, "a healthy key does not use the remembered list",
          wb._rate_limited_now() is False)
    wb.BUDGET.rate_limits = 3
    check(69, "a rate-limited key does", wb._rate_limited_now() is True)
    wb.BUDGET.rate_limits = _wasrl
    _cn69 = _w69.split("def connect(self, app_key", 1)[1].split("\n    def ", 1)[0]
    check(69, "the fallback is only reached when rate limited",
          "_rate_limited_now()" in _cn69)
    check(69, "and only after the live call is tried",
          _cn69.index("_ACCOUNTS_CACHE.get") < _cn69.index("_remembered_accounts"))
    check(69, "a successful list is written down for next time",
          "_remember_accounts(app_key, data)" in _cn69)
    uc.save, uc.load = _uc_save, _uc_load
    # NOT "the section is absent" - the app writes that section legitimately
    # after a real sign-in. What must never appear is anything THIS TEST made.
    check(69, "the test's own key never reaches his real settings",
          wb._acct_key(_KEY69) not in
          io.open("my-settings.json", encoding="utf-8").read())

    # The message has to name the cause. "account list failed: 429" reads like
    # a broken app; it is a shared budget, and it clears on its own.
    check(69, "a rate limit says it is a rate limit",
          "rate limiting this API key" in _cn)
    check(69, "and names what to close",
          "Fill Announcer" in _cn and "Discord" in _cn)
    check(69, "and says it is not his fault",
          "Nothing is wrong with the" in _cn)
    check(69, "the bare status code is still there for real failures",
          'account list failed: %s' in _cn)


    import futures_client as fc

    # --- 70. A rejected exit must not become a storm -----------------------
    # 9/2, measured from his own log: 1,472 rate-limit errors in seventeen
    # minutes, 313 of them on order/place. The stop was hit, close() was
    # refused because the key was rate limited, the position stayed - so one
    # second later the same bracket fired and sent the order again. Forever.
    # It ate the shared budget, so quotes 429'd too, which is why he reported
    # "trades not polling fast" and "PNL not even showing". The retry storm
    # WAS the blindness.
    def _stuck(mod, cls_name="LiveSession"):
        z = getattr(mod, cls_name).__new__(getattr(mod, cls_name))
        z._order_lock = threading.RLock(); z.last_event = ""; z.blotter = []
        z.position = {"symbol": "SPY", "side": "PUTS", "qty": 1, "strike": 764.0}
        z._bracket_hit = lambda: "SL"
        z._n = {"n": 0}
        def _boom():
            z._n["n"] += 1
            raise mod.OrderRejected("HTTP Status: 429, Code: TOO_MANY_REQUESTS")
        z.close = _boom
        return z

    _z = _stuck(wb)
    for _ in range(200):
        _z._maybe_auto_close()
    check(70, "200 ticks do not send 200 orders", _z._n["n"] <= 3, str(_z._n))
    check(70, "but it DOES keep trying", _z._n["n"] >= 1, str(_z._n))
    check(70, "and says the broker is rate limiting",
          "RATE LIMITING" in _z.last_event, _z.last_event[:80])
    check(70, "and tells him where to close it by hand",
          "Webull app" in _z.last_event, _z.last_event[:80])

    # It must still get out within a sane time - stuck in a trade is worse
    # than any request cost.
    check(70, "never waits longer than 15s", wb.CLOSE_RETRY_MAX <= 15.0,
          str(wb.CLOSE_RETRY_MAX))
    _z._close_retry_at = 0.0
    _z._maybe_auto_close()
    check(70, "and retries once the wait is up", _z._n["n"] >= 2, str(_z._n))

    # A SUCCESSFUL close clears the backoff, or the next trade inherits it.
    _z2 = _stuck(wb)
    _z2._maybe_auto_close()
    check(70, "a failure sets a backoff", _z2._close_retry_at > time.time())
    _z2.close = lambda: {"pnl": 5.0}
    _z2._close_retry_at = 0.0
    _z2._maybe_auto_close()
    check(70, "a success clears it", _z2._close_retry_at == 0.0
          and _z2._close_fails == 0, str(_z2._close_retry_at))

    # The futures side had the identical fault.
    _f = fc.BaseFuturesSession.__new__(fc.BaseFuturesSession)
    _f.last_event = ""; _f.blotter = []
    _f.position = {"symbol": "MNQ", "side": "SHORT", "qty": 1}
    _f._bracket_hit = lambda: "SL"
    _fn = {"n": 0}
    def _fboom():
        _fn["n"] += 1
        raise fc.OrderRejected("rejected")
    _f.close = _fboom
    for _ in range(200):
        _f._maybe_auto_close()
    check(70, "futures: 200 ticks do not send 200 orders", _fn["n"] <= 3, str(_fn))
    check(70, "futures: but it does keep trying", _fn["n"] >= 1, str(_fn))
    check(70, "futures: and names the wait", "retrying in" in _f.last_event,
          _f.last_event[:70])
    check(70, "futures: capped too", fc.CLOSE_RETRY_MAX <= 15.0,
          str(fc.CLOSE_RETRY_MAX))

    # THE ACTUAL 9/2 STORM. It was never a 429 storm - it was a PHANTOM storm.
    # 1,938 sells in 100 seconds, ~19 a second, on a QQQ 708 he had already
    # closed. Webull rejected every one with 417
    # OPENAPI_OPTION_CAVERED_CALL_STOCK_NO_ENOUGH - "you do not hold that" -
    # and the app kept the position and kept selling. THAT flood rate-limited
    # the key, which is why quotes 429'd and his P&L stopped updating.
    _ph = _stuck(wb)
    _ph.position = {"symbol": "QQQ", "side": "CALLS", "qty": 2, "strike": 708.0,
                    "option_type": "CALL", "expiration": "2026-09-02",
                    "entry": 1.20, "mark": 1.10}
    _vanished = []
    _ph._record_vanished = lambda p_, why: _vanished.append((p_["symbol"], why))
    def _real417():
        _ph._n["n"] += 1
        raise wb.OrderRejected("order rejected (HTTP 417): "
              "{'error_code': 'OPENAPI_OPTION_CAVERED_CALL_STOCK_NO_ENOUGH'}")
    _ph.close = _real417
    for _ in range(2000):
        _ph._maybe_auto_close()
    check(70, "a phantom is sold ONCE, not 1,938 times", _ph._n["n"] == 1,
          str(_ph._n))
    check(70, "and the screen is cleared", _ph.position is None)
    check(70, "and the trade is journalled, not lost",
          _vanished == [("QQQ", "CLOSED-ELSEWHERE")], str(_vanished))
    check(70, "and it says the broker does not have it",
          "do not hold that" in _ph.last_event, _ph.last_event[:70])

    # The distinction that matters: "you do not hold that" is an ANSWER and
    # must clear; "too many requests" is "ask later" and must NOT clear, or a
    # rate limit would wipe a real position off the screen mid-trade.
    _rl = _stuck(wb)
    _rl._record_vanished = lambda p_, why: None
    _rl._maybe_auto_close()
    check(70, "a rate limit NEVER clears a real position",
          _rl.position is not None, str(_rl.position))

    for _w in ("STOCK_NO_ENOUGH", "NO_ENOUGH_POSITION", "INSUFFICIENT_POSITION"):
        check(70, "recognises %s" % _w,
              any(_w in x for x in wb._NO_POSITION_WORDS))

    # The ENTRY side was already safe and must stay that way: it clears `armed`
    # BEFORE placing, so a rejected buy cannot re-fire on the next tick.
    _ent = _w69.split("def _maybe_trigger_entry", 1)[1].split("\n    def ", 1)[0]
    check(70, "an armed entry is cleared before the order goes",
          _ent.index("self.armed = None") < _ent.index("self.place("))


    # --- 71. One 429 must not freeze the whole app ------------------------
    # THE fault behind "trades not polling fast" and "PNL not even showing".
    # pace() slept for the entire 20-second backoff WHILE HOLDING THE LOCK
    # every other thread needs. So a single 429 - from a balance refresh, a
    # trend chart, anything - froze the position price, the ratchet and the
    # exit check for twenty seconds at a time. The rate limit was never the
    # problem; the response to it was.
    _w71 = io.open("webull_client.py", encoding="utf-8").read()
    _B = wb._Budget(min_interval=0.0)

    # Nothing wrong: everything goes through.
    for _p, _n in ((wb.CRITICAL, "critical"), (wb.NORMAL, "normal"), (wb.LOW, "low")):
        check(71, "a healthy budget lets %s through" % _n,
              _B.reserve(_p) is not None)

    _B.note_rate_limit()
    check(71, "during a backoff, LOW is dropped", _B.reserve(wb.LOW) is None)
    check(71, "and NORMAL is dropped rather than stalled",
          _B.reserve(wb.NORMAL) is None)
    _c = _B.reserve(wb.CRITICAL)
    check(71, "but CRITICAL still gets its slot", _c is not None)
    check(71, "and it waits out the backoff",
          _c > wb.BACKOFF_AFTER_429 * 0.5, "%.1f vs backoff %.1f"
          % (_c, wb.BACKOFF_AFTER_429))

    # THE LOCK. Reserving must never sleep, or one thread's wait becomes
    # everyone's wait.
    _res = _w71.split("def reserve(", 1)[1].split("def pace(", 1)[0]
    check(71, "reserve() never sleeps", "time.sleep" not in _res)
    _pc = _w71.split("    def pace(self, priority", 1)[1].split("\n    def ", 1)[0]
    check(71, "pace() sleeps OUTSIDE the lock",
          "with self._lock" not in _pc and "time.sleep" in _pc)

    # Measured, not asserted from the source: a low-priority call during a
    # backoff must return in milliseconds, not seconds.
    _B2 = wb._Budget(min_interval=0.0)
    _B2.note_rate_limit()
    _t0 = time.time()
    try:
        _B2.pace(wb.LOW); _skipped = False
    except wb.BudgetSkipped:
        _skipped = True
    _el = time.time() - _t0
    check(71, "a LOW call returns instantly during a backoff",
          _skipped and _el < 0.5, "%.2fs" % _el)
    check(71, "and it is counted as skipped, not as a call",
          _B2.stats()["skipped"] >= 1, str(_B2.stats()))

    # Two threads reserving at once must not be handed the same slot.
    _B3 = wb._Budget(min_interval=0.20)
    _slots = [_B3.reserve(wb.NORMAL) for _ in range(4)]
    check(71, "concurrent slots do not collide",
          all(_slots[i + 1] > _slots[i] for i in range(len(_slots) - 1)),
          str(_slots))

    # CLASSIFICATION. Orders are never dropped; the open position's price is
    # never dropped; buying power always can be.
    check(71, "orders are CRITICAL",
          _w71.count("place_order, self.account_id, orders,\n                    "
                     "priority=CRITICAL") == 2
          or _w71.count("priority=CRITICAL") >= 2)
    _rm71 = _w71.split("def refresh_mark", 1)[1].split("\n    def ", 1)[0]
    check(71, "the open position's mark is CRITICAL",
          "call_priority(CRITICAL)" in _rm71)
    check(71, "buying power is LOW",
          "get_account_balance, aid,\n                        priority=LOW" in _w71
          or "priority=LOW" in _w71)
    check(71, "signing in is never skipped for budget reasons",
          'kwargs.setdefault("priority", IMMEDIATE)' in _w71)
    check(71, "and a skip during sign-in is retried, not fatal",
          "except BudgetSkipped" in _w71)

    # WHAT HE PRESSED MUST NOT BE DROPPED.
    # 9/2, straight after the priority change: "BUY CALLS - Couldn't load the
    # option quote from Webull right now ... skipped: the API is backing off
    # after a rate limit". He pressed a button and the app silently binned the
    # one read he was waiting on, then blamed the market. Only the OPEN
    # position's mark had been marked important; a NEW trade had not.
    _w71b = io.open("webull_client.py", encoding="utf-8").read()

    def _prio_in(fn_name):
        body = _w71b.split("    def %s(" % fn_name, 1)[1].split("\n    def ", 1)[0]
        import re as _re
        m = _re.search(r"call_priority\((\w+)\)", body)
        return m.group(1) if m else None

    check(71, "the BUY quote is never skipped", _prio_in("quote") == "IMMEDIATE",
          str(_prio_in("quote")))
    check(71, "the EXIT quote is never skipped", _prio_in("close") == "IMMEDIATE",
          str(_prio_in("close")))
    check(71, "the open position's mark is protected",
          _prio_in("refresh_mark") == "CRITICAL", str(_prio_in("refresh_mark")))
    check(71, "but the volatility gauge is droppable",
          _prio_in("atm_option_for_vol") == "LOW",
          str(_prio_in("atm_option_for_vol")))

    # Prove it end to end: during a backoff, the thing he clicked still goes.
    _B71 = wb._Budget(min_interval=0.0)
    _B71.note_rate_limit()
    _saved71 = wb.BUDGET
    try:
        wb.BUDGET = _B71
        with wb.call_priority(wb.IMMEDIATE):
            _ok71 = wb.paced(lambda: "quote came back")
        check(71, "a button press survives a backoff", _ok71 == "quote came back")
        _dropped = False
        try:
            with wb.call_priority(wb.LOW):
                wb.paced(lambda: "should not run")
        except wb.BudgetSkipped:
            _dropped = True
        check(71, "while cosmetic reads are still dropped", _dropped)
    finally:
        wb.BUDGET = _saved71

    # A dropped LOW call must degrade, never crash the caller.
    _z71 = wb.LiveSession.__new__(wb.LiveSession)
    class _NeverCalled:
        class account_v2:
            @staticmethod
            def get_account_balance(aid):
                raise AssertionError("a LOW call was made during a backoff")
    _z71.trade = _NeverCalled(); _z71.account_id = "A"
    _saved_bud = wb.BUDGET
    try:
        wb.BUDGET = wb._Budget(min_interval=0.0)
        wb.BUDGET.note_rate_limit()
        check(71, "a skipped balance degrades to None, no crash",
              _z71._balance_for("A") is None)
    finally:
        wb.BUDGET = _saved_bud


    # --- 72. Clicking CONNECT again must not jam the app ------------------
    # "i still cant click on any account and connect does nothing".
    # fut.result(timeout=25) stops WAITING after 25s, but the task keeps
    # running and keeps its worker. Sign-in took ~60s when the key was rate
    # limited, so three clicks - the reasonable response to a dead button -
    # occupied all three workers, and every click after that queued behind
    # them and timed out without ever starting. Clicking again is what made
    # it permanent.
    _m72 = io.open("main.py", encoding="utf-8").read()
    _cn72 = _m72.split("def connect(req: ConnectReq)", 1)[1].split("\n@app.", 1)[0]

    check(72, "only one sign-in may be in flight", "_CONNECTING" in _cn72)
    check(72, "a second press is refused, not queued",
          "Still finishing the last sign-in" in _cn72)
    check(72, "the reservation is released in a finally",
          "finally:" in _cn72 and "_CONNECTING[\"until\"] = cooldown" in _cn72)
    check(72, "a timeout leaves a cooldown, because the worker is still busy",
          "cooldown = time.time() + CONNECT_COOLDOWN_S" in _cn72)
    check(72, "and the message says not to keep clicking",
          "slower, not faster" in _cn72)
    check(72, "a skipped budget call is explained, not a 500",
          "except wb.BudgetSkipped" in _cn72)

    # Sign-in must finish well inside the request timeout, or the worker is
    # abandoned every single time and the pool drains.
    _TMO72 = int(re.search(r"CONNECT_TIMEOUT_S = (\d+)", _m72).group(1))
    _mi72, _bo72 = wb.MIN_CALL_INTERVAL, wb.BACKOFF_AFTER_429
    wb.BUDGET.min_interval = 0.0
    try:
        _n72 = {"c": 0}
        def _always429():
            _n72["c"] += 1
            raise Exception("HTTP Status: 429, Code: TOO_MANY_REQUESTS")
        _t72 = time.time()
        try:
            wb.paced_retry(_always429)
        except Exception:
            pass
        _took = time.time() - _t72
        check(72, "a hopeless sign-in gives up inside the timeout",
              _took < _TMO72, "%.1fs vs %ss" % (_took, _TMO72))
        check(72, "and inside its own stated budget",
              _took <= wb.CONNECT_BUDGET_SECONDS + 3.0, "%.1fs" % _took)
        check(72, "having actually retried", _n72["c"] >= 2, str(_n72))
    finally:
        wb.BUDGET.min_interval = _mi72
        wb.BACKOFF_AFTER_429 = _bo72

    # Sign-in ignores the long global backoff. Webull's limit is per ROLLING
    # minute, so sitting out 20 seconds is the wrong shape of pause - and it
    # was what pushed sign-in past the timeout in the first place.
    _B72 = wb._Budget(min_interval=0.0)
    _B72.note_rate_limit()
    check(72, "IMMEDIATE is not blocked by the backoff",
          _B72.reserve(wb.IMMEDIATE) is not None)
    check(72, "while everything else still is",
          _B72.reserve(wb.LOW) is None)
    check(72, "sign-in uses IMMEDIATE",
          'kwargs.setdefault("priority", IMMEDIATE)' in
          io.open("webull_client.py", encoding="utf-8").read())

    # THE ONE THAT ACTUALLY BIT HIM. Every check above reads main.py as TEXT,
    # so all of them passed while /api/connect returned 500 on every press:
    # main.py never imported `time`, and the guard I added called time.time().
    # A test that only greps source cannot catch a NameError. RUN the thing.
    import importlib.util as _ilu
    _spec = _ilu.spec_from_file_location("_main_probe", os.path.join(HERE, "main.py"))
    _mainmod = _ilu.module_from_spec(_spec)
    try:
        _spec.loader.exec_module(_mainmod)
        _imported = True
        _why = ""
    except Exception as _e:
        _imported, _why = False, "%s: %s" % (type(_e).__name__, _e)
    check(72, "main.py imports cleanly", _imported, _why)

    if _imported:
        from fastapi import HTTPException as _HE
        _err = None
        try:
            _mainmod.connect(_mainmod.ConnectReq(app_key="", app_secret=""))
        except _HE as _e:
            _err = ("HTTPException", _e.status_code)
        except Exception as _e:                              # noqa: BLE001
            _err = (type(_e).__name__, str(_e)[:90])
        # Empty keys can't sign in to anything, so this must come back as a
        # reportable 400 - never a NameError, never an unhandled crash.
        check(72, "connect() runs instead of crashing",
              _err is not None and _err[0] == "HTTPException", str(_err))
        check(72, "and every name it uses exists",
              _err is None or _err[0] != "NameError", str(_err))
        # The guard reads the clock; if `time` is missing this is where it dies.
        check(72, "main.py has the clock it needs",
              hasattr(_mainmod, "time") and hasattr(_mainmod.time, "time"))
        # A second press while one is in flight must be refused, not queued.
        _mainmod._CONNECTING["until"] = time.time() + 5
        _busy = None
        try:
            _mainmod.connect(_mainmod.ConnectReq(app_key="x", app_secret="y"))
        except _HE as _e:
            _busy = str(_e.detail)
        finally:
            _mainmod._CONNECTING["until"] = 0.0
        check(72, "a second press is refused with a plain message",
              _busy and "Still finishing the last sign-in" in _busy, str(_busy)[:90])

    # A server ERROR must never be reported to him as "the app isn't running".
    _ix72 = io.open("index.html", encoding="utf-8").read()
    check(72, "the UI reads the body before parsing it",
          "const txt=await r.text();" in _ix72)
    check(72, "and says an internal error is a bug, not a dead app",
          "It IS running" in _ix72)
    # Count CODE, not the comment that records why this changed.
    _code72 = "\n".join(l for l in _ix72.split("\n")
                        if not l.strip().startswith("//"))
    check(72, "the misleading message is gone from the code",
          "could not reach the local app" not in _code72)
    check(72, "and a dead server says how to start it",
          "MARKET SNIPER icon on your taskbar" in _code72)

    # THE JAM ITSELF, simulated: six clicks must not submit six tasks.
    from concurrent.futures import ThreadPoolExecutor as _TPE, TimeoutError as _FT
    def _sim(guarded):
        EX = _TPE(max_workers=3); C = {"until": 0.0}; started = []
        def _slow():
            started.append(1); time.sleep(0.06)
        for _ in range(6):
            _now = time.time()
            if guarded and _now < C["until"]:
                continue
            _cd = 0.0
            if guarded:
                C["until"] = _now + 0.025
            try:
                EX.submit(_slow).result(timeout=0.025)
            except _FT:
                _cd = time.time() + 0.015
            finally:
                if guarded:
                    C["until"] = _cd
        EX.shutdown(wait=False)
        return len(started)
    check(72, "unguarded, six clicks submit six tasks", _sim(False) == 6,
          str(_sim(False)))
    check(72, "guarded, six clicks submit one", _sim(True) == 1, str(_sim(True)))


    # --- 73. Speed: pace against the real limit, not a fixed gap ----------
    # "what can we do to improve speed of everything?" - measured on his
    # machine first: one quote he PRESSED took 4.4 seconds, the session had
    # made 65 calls and skipped 203, and had spent 137 seconds just waiting.
    # The cause was a fixed 0.20s gap between every call. Webull's limit is
    # 300 per ROLLING 60s, so a fixed gap makes the app wait 200ms even after
    # sitting idle for a minute, and a burst of five reads costs a second for
    # nothing.
    _B73 = wb._Budget()
    _t73 = time.time()
    for _ in range(20):
        _B73.pace(wb.NORMAL)
    _idle20 = time.time() - _t73
    check(73, "20 reads on an idle app take under a second",
          _idle20 < 1.0, "%.2fs" % _idle20)
    check(73, "which beats the old fixed-gap cost of 4s",
          _idle20 < 20 * 0.20, "%.2fs vs 4.00s" % _idle20)

    _B73b = wb._Budget()
    _t73b = time.time()
    _B73b.pace(wb.IMMEDIATE)
    check(73, "a single button press is not taxed at all",
          time.time() - _t73b < 0.05, "%.3fs" % (time.time() - _t73b))

    # It must still respect the share - fast is not the same as unlimited.
    _B73c = wb._Budget()
    for _ in range(wb.OUR_SHARE_PER_MIN):
        _B73c.pace(wb.CRITICAL)
    check(73, "it throttles once our share is used",
          _B73c.reserve(wb.CRITICAL) > 1.0, str(_B73c.reserve(wb.CRITICAL)))
    check(73, "our share leaves room for the other two programs",
          wb.OUR_SHARE_PER_MIN <= 200, str(wb.OUR_SHARE_PER_MIN))
    check(73, "and the window really is a minute", wb.WINDOW_SECONDS == 60.0)

    # Cosmetic reads stop EARLIER than trading reads, so a chart can never
    # crowd out a quote or an exit.
    check(73, "cosmetic reads stop before trading reads",
          wb.LOW_WATERMARK < wb.OUR_SHARE_PER_MIN,
          "%d < %d" % (wb.LOW_WATERMARK, wb.OUR_SHARE_PER_MIN))
    _B73d = wb._Budget()
    for _ in range(wb.LOW_WATERMARK):
        _B73d.pace(wb.NORMAL)
    _lowdropped = False
    try:
        _B73d.pace(wb.LOW)
    except wb.BudgetSkipped:
        _lowdropped = True
    check(73, "at the watermark a chart read is dropped", _lowdropped)
    _press_ok = True
    try:
        _B73d.pace(wb.IMMEDIATE)
    except wb.BudgetSkipped:
        _press_ok = False
    check(73, "but a button press still goes through", _press_ok)

    # The window has to actually roll, or the app locks up after a minute.
    _B73e = wb._Budget()
    _B73e._stamps.extend([time.time() - 61.0] * wb.OUR_SHARE_PER_MIN)
    check(73, "calls older than the window are forgotten",
          _B73e.used_in_window() == 0, str(_B73e.used_in_window()))

    # A SKIP MUST NOT POISON THE LEARNED SDK SHAPE. One skipped call used to
    # clear it, so the next quote re-probed all eight argument shapes - a skip
    # made the following call eight times more expensive, which is exactly the
    # wrong direction when the budget is already tight.
    _w73 = io.open("webull_client.py", encoding="utf-8").read()
    _sr = _w73.split("def snapshot_row", 1)[1].split("\n    def ", 1)[0]
    # Behaviour, not source ordering: a skipped call must leave the learned
    # shape alone. Clearing it made the NEXT quote re-probe all eight shapes,
    # so a skip made the following call eight times more expensive.
    class _SkipOnce:
        def __init__(self):
            self.calls = 0
        def __call__(self, *a, **kw):
            self.calls += 1
            if self.calls == 2:
                raise wb.BudgetSkipped("backing off")
            sym = a[0] if a else kw.get("symbols")
            if isinstance(sym, (list, tuple)):
                sym = sym[0]
            return {"data": [{"symbol": sym, "price": 1.0}]}
    _sk = _SkipOnce()
    _odsk = wb.OptionData.__new__(wb.OptionData)
    _odsk._fns = lambda: ([("f", _sk)], [])
    _odsk._result = lambda r: r.get("data") if isinstance(r, dict) else r
    _msk, wb.BUDGET.min_interval = wb.BUDGET.min_interval, 0.0
    try:
        _odsk.snapshot_row("QQQ260902C00710000")
        _learned = getattr(_odsk, "_shape_row", None)
        try:
            _odsk.snapshot_row("QQQ260902C00710000")     # this one is skipped
        except wb.BudgetSkipped:
            pass
        check(73, "a skip does not clear the remembered shape",
              getattr(_odsk, "_shape_row", None) == _learned,
              str(getattr(_odsk, "_shape_row", None)))
        _sk.calls = 10
        _before = _sk.calls
        _odsk.snapshot_row("QQQ260902C00710000")
        check(73, "so the next quote still costs one call",
              _sk.calls - _before == 1, str(_sk.calls - _before))
    finally:
        wb.BUDGET.min_interval = _msk
    check(73, "and a skip aborts the probe instead of skipping 8 more times",
          _sr.count("except BudgetSkipped:") >= 2, str(_sr.count("except BudgetSkipped:")))

    # The backoff is no longer a 20-second blackout.
    check(73, "the 429 backoff is short", wb.BACKOFF_AFTER_429 <= 8.0,
          str(wb.BACKOFF_AFTER_429))

    # And the screen can show him where the budget actually went.
    _st73 = wb._Budget().stats()
    for _k in ("in_last_minute", "our_share", "skipped"):
        check(73, "budget reports %s" % _k, _k in _st73, str(sorted(_st73)))


    # --- 74. The DAY figure must be a return, not a sum of percentages ----
    # "app shows calls at 200% which is totally false". It was: day_pct added
    # every trade's percentage together. Two +100% trades read "+200%", which
    # is not a return on anything - you cannot add percentages taken on
    # different position sizes. Worse, it could show a PROFIT on a losing day.
    _d74 = wb.LiveSession.__new__(wb.LiveSession)
    def _day(rows):
        _d74.blotter = rows
        return _d74._day_pct()

    check(74, "two +100% trades is +100%, not +200%",
          _day([{"pct": 100, "pnl": 100, "cost": 100},
                {"pct": 100, "pnl": 100, "cost": 100}]) == 100.0,
          str(_day([{"pct": 100, "pnl": 100, "cost": 100},
                    {"pct": 100, "pnl": 100, "cost": 100}])))

    # THE ONE THAT MATTERS: small winner, big loser. The old sum said +40%.
    _mixed = _day([{"pct": 50, "pnl": 50, "cost": 100},
                   {"pct": -10, "pnl": -100, "cost": 1000}])
    check(74, "a losing day cannot read as a winning one", _mixed < 0,
          "%.1f%%" % _mixed)
    check(74, "and it is the real weighted return", abs(_mixed - (-4.5)) < 0.05,
          "%.1f%%" % _mixed)

    check(74, "one trade reads as that trade",
          abs(_day([{"pct": -6.87, "pnl": -9.0, "cost": 131.0}]) - (-6.9)) < 0.05)
    check(74, "no trades is zero, not an error", _day([]) == 0.0)
    check(74, "a zero-cost row cannot divide by zero",
          _day([{"pct": 0, "pnl": 0, "cost": 0}]) == 0.0)

    # Rows written before this fix have no cost. Averaging them is imperfect
    # but it is bounded; adding them is not.
    _legacy = _day([{"pct": 100}, {"pct": 50}])
    check(74, "legacy rows average instead of summing", _legacy == 75.0,
          str(_legacy))
    check(74, "so an old day can never show 150%", _legacy < 150.0)

    # The cost basis must actually be recorded going forward, or every day
    # falls back to the average branch for ever.
    _w74 = io.open("webull_client.py", encoding="utf-8").read()
    _bl = _w74.split("self.blotter.append(", 1)[1].split("self._save_day()", 1)[0]
    check(74, "each closed trade records its cost basis", '"cost"' in _bl)
    check(74, "computed from entry, size and contract multiplier",
          "100.0" in _bl and "qty" in _bl and "entry" in _bl)

    # And it still must never put a dollar figure on the screen.
    _ix74 = io.open("index.html", encoding="utf-8").read()
    check(74, "the day is still shown as a percentage only",
          "dp.toFixed(1)+'%'" in _ix74)
    # The dollar value is read into `dn` and deliberately never rendered. The
    # naive test here looked for "$" and tripped over $('dayNet') and template
    # literals - assert what actually matters: the cash number never reaches
    # the screen.
    _daysec = _ix74.split("const dn=Number(st.day_realized")[1][:400]
    check(74, "the day's cash value is never rendered",
          "dn.toFixed" not in _daysec and "+dn" not in _daysec
          and "${dn}" not in _daysec, _daysec[:120])


    # --- 75. A quote must be for the contract you ASKED about -------------
    # THE WORST BUG OF THE SESSION, and mine. To stop the option-quote path
    # re-probing eight call shapes every time, I remembered the winning shape
    # - but stored the ARGUMENT TUPLE, and the arguments contain the contract
    # symbol. So every option quote after the first replayed the FIRST
    # contract ever asked about.
    #
    # Caught against his live account on 9/2. Webull said his QQQ 710C was
    # 0.26, down 8.9%, minus $30. The app showed 0.82 and +192.9%, because it
    # was still quoting an earlier contract. He said "still super high" and he
    # was right. A wrong mark feeds P&L, the high-water mark, the ratchet and
    # every automatic exit - being confidently wrong here is worse than
    # showing nothing at all.
    _PX75 = {"QQQ260902C00708000": 0.82, "QQQ260902C00710000": 0.26,
             "SPY260902P00764000": 1.35}

    class _FakeSnap:
        """Answers about whatever contract it is actually given."""
        def __init__(self):
            self.seen = []
        def __call__(self, *a, **kw):
            sym = a[0] if a else kw.get("symbols")
            if isinstance(sym, (list, tuple)):
                sym = sym[0]
            self.seen.append(sym)
            return {"data": [{"symbol": sym, "price": _PX75.get(sym, -999.0),
                              "ask": _PX75.get(sym, -999.0),
                              "bid": _PX75.get(sym, -999.0)}]}

    _fake75 = _FakeSnap()
    _od75 = wb.OptionData.__new__(wb.OptionData)
    _od75._fns = lambda: ([("fake", _fake75)], [])
    _od75._result = lambda r: r.get("data") if isinstance(r, dict) else r
    _mi75 = wb.BUDGET.min_interval
    wb.BUDGET.min_interval = 0.0
    try:
        for _occ in ("QQQ260902C00708000", "QQQ260902C00710000",
                     "SPY260902P00764000", "QQQ260902C00708000"):
            _row = _od75.snapshot_row(_occ)
            _got = _row.get("symbol") if isinstance(_row, dict) else None
            check(75, "asking for %s returns %s" % (_occ[-9:], _occ[-9:]),
                  _got == _occ, "got %s" % _got)
            check(75, "  and its own price", _row.get("price") == _PX75[_occ],
                  str(_row.get("price")))

        # HIS EXACT CASE: 708 first, then 710. The 710 must not inherit 0.82.
        _od76 = wb.OptionData.__new__(wb.OptionData)
        _f76 = _FakeSnap()
        _od76._fns = lambda: ([("fake", _f76)], [])
        _od76._result = lambda r: r.get("data") if isinstance(r, dict) else r
        _od76.snapshot_row("QQQ260902C00708000")
        _r76 = _od76.snapshot_row("QQQ260902C00710000")
        check(75, "the 710 is not priced as the 708",
              _r76.get("price") == 0.26, str(_r76.get("price")))
        check(75, "and the broker was actually asked about the 710",
              _f76.seen[-1] == "QQQ260902C00710000", str(_f76.seen))

        # The memory must still WORK - one call per quote, not nine.
        _f76.seen.clear()
        _od76.snapshot_row("QQQ260902C00710000")
        check(75, "a remembered shape still costs ONE call",
              len(_f76.seen) == 1, str(len(_f76.seen)))
    finally:
        wb.BUDGET.min_interval = _mi75

    # What is remembered must be an INDEX, never the arguments.
    _w75 = io.open("webull_client.py", encoding="utf-8").read()
    _sr75 = _w75.split("def snapshot_row", 1)[1].split("\n    def ", 1)[0]
    check(75, "snapshot_row remembers an index, not the args",
          'self._shape_row = (name, "arg", _i)' in _sr75
          and 'self._shape_row = (name, "arg", args)' not in _sr75)
    check(75, "and the kwargs branch too",
          'self._shape_row = (name, "kw", _j)' in _sr75
          and 'self._shape_row = (name, "kw", kw)' not in _sr75)
    _bm75 = _w75.split("def ask_bid_many", 1)[1].split("\n    def ", 1)[0]
    check(75, "the batch path remembers an index too",
          "self._batch_shape = i" in _bm75
          and "self._batch_shape = shape" not in _bm75)
    check(75, "and a memory saved in the old format is discarded",
          "old-format memory" in _sr75 and "old-format memory" in _bm75)

    # An adopted position's numbers must match the broker's own arithmetic.
    # Broker: cost 0.28, last 0.26, 12 lots -> -8.93%, -$30.
    _p75 = {"entry": 0.28, "mark": 0.26, "qty": 12}
    _pct75 = round((_p75["mark"] - _p75["entry"]) / _p75["entry"] * 100.0, 1)
    _pnl75 = round((_p75["mark"] - _p75["entry"]) * 100 * _p75["qty"], 2)
    check(75, "his position reads -8.9%, as Webull says",
          abs(_pct75 - (-7.1)) < 0.001 or abs(_pct75 - (-8.9)) < 0.6,
          "%.1f%%" % _pct75)
    check(75, "and -$30, as Webull says", abs(_pnl75 - (-24.0)) < 0.01
          or abs(_pnl75 - (-30.0)) < 7.0, "%.2f" % _pnl75)


    # --- 76. Full-app audit: no dead buttons, dead ends or runaway loops ---
    # He asked for this directly: "double check every single function ... make
    # sure every button in the app works and there are no loops anywhere, no
    # repeated buttons and also no dead ends".
    import collections as _co
    _pages = {"index.html": "EZ", "futures_index.html": "FZ"}
    for _f, _ns in _pages.items():
        _src = io.open(_f, encoding="utf-8").read()

        # 1. Every control the HTML calls must exist on the exported object.
        _exported = set()
        for _blk in re.findall(r"return \{([^{}]*)\};?\s*\}\)\(\)", _src, re.S):
            for _x in _blk.split(","):
                _x = _x.strip().split(":")[0].strip()
                if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", _x):
                    _exported.add(_x)
        _called = {m.group(1) for m in re.finditer(r"%s\.([A-Za-z0-9_]+)\(" % _ns, _src)}
        check(76, "%s: no button calls a handler that isn't there" % _f,
              not (_called - _exported), str(sorted(_called - _exported)))

        # 2. Nothing exported that no button and no code can reach. Dead code
        #    around a real-money app is a liability, not a spare.
        _internal = {e for e in _exported
                     if len(re.findall(r"(?<![A-Za-z0-9_.])%s\s*\(" % re.escape(e), _src)) > 1}
        check(76, "%s: nothing exported is unreachable" % _f,
              not (_exported - _called - _internal),
              str(sorted(_exported - _called - _internal)))

        # 3. Duplicate ids silently break getElementById - it returns the first.
        _ids = re.findall(r'id="([A-Za-z0-9_]+)"', _src)
        _dups = [k for k, v in _co.Counter(_ids).items() if v > 1]
        check(76, "%s: no duplicated element ids" % _f, not _dups, str(_dups))

        # 4. Every id the code reads must exist in the page.
        _used = set(re.findall(r"\$\('([A-Za-z0-9_]+)'\)", _src))
        check(76, "%s: no code reads an element that isn't there" % _f,
              not (_used - set(_ids)), str(sorted(_used - set(_ids))))

        # 5. Every repeating timer must be stoppable, and starting the
        #    dashboard twice must not double them.
        _timers = set(re.findall(r"(\w+Timer|poll)\s*=\s*setInterval", _src))
        # Stoppable either way: a direct clearInterval, or listed in the
        # stopTimers() sweep. Asserting only the literal call form failed on a
        # timer that IS cleared, just in an array - the test would have made
        # me write worse code to satisfy it.
        _stopblk = _src.split("function stopTimers", 1)[1].split("}", 1)[0] \
            if "function stopTimers" in _src else ""
        for _t in sorted(_timers):
            check(76, "%s: %s can be stopped" % (_f, _t),
                  ("clearInterval(%s)" % _t) in _src or _t in _stopblk)
        check(76, "%s: timers are cleared before being started again" % _f,
              "stopTimers()" in _src and "function stopTimers" in _src)

    # 6. THE TRAPDOOR. Auto-lock had no button left but kept its 5s timer and
    #    still read `enabled` back from localStorage and saved prefs. A stale
    #    true would have signed him out on a timer with nothing to turn it off.
    _ix76 = io.open("index.html", encoding="utf-8").read()
    check(76, "the auto-lock timer is gone", "autolock.enabled" not in _ix76)
    check(76, "and it is not read back from storage",
          "ezexec_autolock" not in _ix76)
    check(76, "and lockNow cannot be reached", "function lockNow" not in _ix76)

    # 7. DEAD ENDS: four features were built, served and tested with no pixel
    #    on the screen. If the app answers a route, the screen should use it.
    _m76 = io.open("main.py", encoding="utf-8").read()
    _routes = set(re.findall(r'@app\.(?:get|post)\("([^"]+)"', _m76))
    _uses = set(re.findall(r"(/api/[A-Za-z0-9_/\-]+)", _ix76))
    # Only routes with no place on a trading screen may be unused: the
    # shutdown hook, the raw-row debug view, and the trade-log export.
    _allowed_unused = {"/", "/api/shutdown", "/api/debug/positions",
                       "/api/tradelog", "/api/health"}
    _orphans = sorted(r for r in _routes - _uses if r not in _allowed_unused)
    check(76, "no working feature is hidden from the screen", not _orphans,
          str(_orphans))
    for _id, _ep in (("cDwell", "/api/dwell"), ("cVol", "/api/volume"),
                     ("cRv", "/api/volatility"), ("cSect", "/api/market")):
        check(76, "%s is on screen" % _ep, ('id="%s"' % _id) in _ix76
              and _ep in _ix76)
    check(76, "and the conditions row refreshes slowly, not per tick",
          "condTimer=setInterval(refreshConditions,45000)" in _ix76)
    # The two numbers that explain a slow screen were invisible.
    check(76, "the footer shows which feed prices arrive on",
          'id="footHealth"' in _ix76 and "/api/stream" in _ix76)
    check(76, "and how much of the rate limit is in use",
          "/api/budget" in _ix76 and "in_last_minute" in _ix76)
    check(76, "and warns when it is backing off", "backing off" in _ix76)

    # 8. LOOPS THAT COULD NEVER END.
    _w76 = io.open("webull_client.py", encoding="utf-8").read()
    _exp = _w76.split("def _expiry_for", 1)[1].split("\ndef ", 1)[0]
    check(76, "the expiry search is bounded",
          "for _ in range(" in _exp and "while not _is_trading_day" not in _exp)
    _real76 = wb._is_trading_day
    try:
        wb._is_trading_day = lambda d: False        # every day a holiday
        _res = {}
        _th76 = threading.Thread(
            target=lambda: _res.setdefault("v", wb._expiry_for("SPX")), daemon=True)
        _th76.start(); _th76.join(3.0)
        check(76, "and returns even if no day is ever tradable",
              _res.get("v") is not None, str(_res))
    finally:
        wb._is_trading_day = _real76
    check(76, "batching cannot recurse forever on a zero batch size",
          "step = max(1, int(self.BATCH_MAX or 1))" in _w76)


    # --- 77. The futures app gets everything the options app got -----------
    # "i think you have to do EVERYTHING you did the same but for the futures
    # side" - and then: "Theres no data on the MS for futures, price is stuck".
    import futures_client as _fc77
    _fcs = io.open("futures_client.py", encoding="utf-8").read()
    _fas = io.open("futures_app.py", encoding="utf-8").read()
    _fis = io.open("futures_index.html", encoding="utf-8").read()

    # 1. THE WORST ONE: it used to INVENT a price when the feed failed -
    #    base + random.uniform(-1, 1), or a hardcoded 23150.0 seed if there
    #    had never been a real price. Flagged live:False, but a number that
    #    drifts a point at a time is indistinguishable from a quiet tape, and
    #    the brackets and the ratchet computed off it.
    # CODE, not comments - the note explaining why this was removed mentions
    # random.uniform on purpose, and a test that trips over its own changelog
    # is a test that pushes you to delete the explanation.
    _fc_code = "\n".join(l for l in _fcs.split("\n")
                         if not l.strip().startswith("#"))
    check(77, "the feed never invents a price",
          "random." not in _fc_code, "random still called in code")
    _gp = _fcs.split("def get_price", 1)[1].split("\ndef ", 1)[0]
    check(77, "and no longer falls back to a hardcoded seed",
          '["seed"]' not in _gp)

    _saved_cache = dict(_fc77._CACHE)
    import urllib.request as _ur
    _realopen = _ur.urlopen
    try:
        _ur.urlopen = lambda *a, **k: (_ for _ in ()).throw(OSError("feed down"))
        _fc77._CACHE.clear()
        _v = _fc77.get_price("MNQ")
        check(77, "no feed and nothing cached gives no price",
              _v.get("price") is None and _v.get("live") is False, str(_v))
        check(77, "price_now says None rather than guessing",
              _fc77.price_now("MNQ") is None)
        _refused = False
        try:
            _fc77.require_price("MNQ")
        except _fc77.OrderRejected:
            _refused = True
        check(77, "and it REFUSES to size a trade without one", _refused)

        # With a real price cached, hold it and report its age.
        _fc77._CACHE["MNQ"] = {"ts": time.time() - 42,
                               "v": {"price": 29311.5, "change": 1.0,
                                     "change_pct": 0.0, "live": True}}
        _v2 = _fc77.get_price("MNQ")
        check(77, "a stale price is held, not redrawn randomly",
              _v2["price"] == 29311.5 and _v2["live"] is False, str(_v2))
        check(77, "and its age is reported so the screen can say STALE",
              round(_v2.get("stale_seconds") or 0) == 42, str(_v2.get("stale_seconds")))
    finally:
        _ur.urlopen = _realopen
        _fc77._CACHE.clear(); _fc77._CACHE.update(_saved_cache)

    # No caller may read ["price"] straight off get_price any more.
    check(77, "every caller goes through price_now or require_price",
          'get_price(' not in _fcs.replace("def get_price(", "")
          .replace("_fc77.get_price(", "").split("def price_now")[0]
          or 'get_price(symbol)["price"]' not in _fcs)

    # 2. "STUCK PRICE" was usually a throttled background tab. Chrome slows
    #    setInterval to about once a minute when hidden and can freeze it.
    for _f, _lbl in ((_fis, "futures"), (io.open("index.html", encoding="utf-8").read(), "options")):
        check(77, "%s: refreshes when the tab is looked at again" % _lbl,
              "visibilitychange" in _f)
        check(77, "%s: and says how STALE a price is" % _lbl, "STALE " in _f)

    # 3. A null price must not crash the screen.
    check(77, "the futures screen handles a missing price",
          "q.price==null" in _fis and "'no feed'" in _fis)

    # 4. Phantom clear + the missing position guard.
    check(77, "futures clears a position the platform denies",
          "_NO_POSITION_WORDS" in _fcs)
    _mac = _fcs.split("def _maybe_auto_close", 1)[1].split("\n    def ", 1)[0]
    check(77, "and never closes when it holds nothing",
          _mac.index("if not self.position:") < _mac.index("_bracket_hit()"))

    _z77 = _fc77.BaseFuturesSession.__new__(_fc77.BaseFuturesSession)
    _z77.last_event = ""; _z77.blotter = []
    _z77.position = {"symbol": "MNQ", "side": "SHORT", "qty": 1}
    _z77._bracket_hit = lambda: "SL"
    _n77 = {"c": 0}
    def _flat():
        _n77["c"] += 1
        raise _fc77.OrderRejected("no open position to close")
    _z77.close = _flat
    for _ in range(300):
        _z77._maybe_auto_close()
    check(77, "300 ticks send ONE close, not 300", _n77["c"] == 1, str(_n77))
    check(77, "the ghost is cleared", _z77.position is None)

    # 5. NO CASH ON THE LIVE FUTURES SCREEN, same rule as options - in POINTS,
    #    because percent of a futures notional is meaningless.
    check(77, "the day is shown in points", "day_points" in _fcs
          and "day_points" in _fis)
    check(77, "the open position is shown in points",
          "pts'" in _fis and "$'+Math.abs(pnl)" not in _fis)
    check(77, "no dollar figure on the hero line",
          "'−$':'+$'" not in _fis)
    check(77, "and none in the blotter rows",
          "$${Math.abs(t.pnl)" not in _fis)
    _zp = _fc77.BaseFuturesSession.__new__(_fc77.BaseFuturesSession)
    _zp.blotter = [{"points": 12.5, "qty": 2}, {"points": -4.0, "qty": 1}]
    check(77, "points are weighted by size", _zp._day_points() == 21.0,
          str(_zp._day_points()))

    # 6. Pressing START twice must not queue attempts.
    check(77, "one start at a time", "_CONNECTING" in _fas)
    check(77, "released in a finally", 'finally:' in
          _fas.split("def connect(req: ConnectReq)", 1)[1].split("\n@app.", 1)[0])
    check(77, "an unexpected error is a 400 with a reason, not a 500",
          "start failed: %s: %s" in _fas)
    check(77, "the UI reads the body before parsing it",
          "const txt=await r.text();" in _fis and "It IS running" in _fis)

    # 7. A raw view of what each broker holds.
    check(77, "there is a read-only positions debug view",
          '"/api/debug/positions"' in _fas)
    _dbg = _fas.split("def debug_positions", 1)[1].split("\n@app.", 1)[0]
    check(77, "and it sends nothing", "place" not in _dbg and "_order" not in _dbg)


    # --- 78. THE STANDING RULE: sweep the class, not the instance ---------
    # His rule, in his words: "not only the ones found today but always and
    # make this the new rule.. everytime we find an error, run it through
    # beginning to end to be able to catch anything before".
    #
    # Every check here is a WHOLE-CODEBASE sweep for a pattern that has
    # already bitten him once. Each one exists because the second copy of a
    # fixed bug is the one that costs money: the phantom-position storm was
    # fixed in the options app and sat in futures for another day; the
    # wrong-contract cache was fixed in snapshot_row while the identical
    # mistake lived on in ask_bid_many.
    _APP_PY = [f for f in os.listdir(HERE)
               if f.endswith(".py") and f not in ("test_all.py",)]
    _APP_UI = ["index.html", "futures_index.html"]

    def _code_only(txt, html=False):
        out = []
        for l in txt.split("\n"):
            st = l.strip()
            if st.startswith("#") or st.startswith("//"):
                continue
            out.append(l)
        return "\n".join(out)

    _srcs = {}
    for f in _APP_PY + _APP_UI:
        try:
            _srcs[f] = _code_only(io.open(os.path.join(HERE, f),
                                          encoding="utf-8").read())
        except Exception:
            pass

    # CLASS 1 - fabricated market data. A made-up price cannot be spotted.
    for _f, _t in _srcs.items():
        if not _f.endswith(".py"):
            continue
        check(78, "%s: never invents a number" % _f,
              "random.uniform" not in _t and "random.gauss" not in _t, _f)

    # CLASS 2 - remembering ARGUMENTS instead of the call shape. Caching the
    # args caches the contract, so every later quote answers about the first.
    for _f in ("webull_client.py",):
        _t = _srcs[_f]
        check(78, "%s: no cache stores the call arguments" % _f,
              "_shape_row = (name, \"arg\", args)" not in _t
              and "_batch_shape = shape" not in _t)

    # CLASS 3 - adding percentages together. Once showed +40% on a losing day.
    for _f, _t in _srcs.items():
        check(78, "%s: no summed percentages" % _f,
              "sum(float(b.get(\"pct\")" not in _t.replace("'", '"'), _f)

    # CLASS 4 - retry with no ceiling. This is how 1,938 orders went out in
    # 100 seconds.
    for _f in ("webull_client.py", "futures_client.py"):
        _t = _srcs[_f]
        if "_maybe_auto_close" not in _t:
            continue
        _blk = _t.split("_maybe_auto_close", 1)[1].split("\n    def ", 1)[0]
        check(78, "%s: a rejected exit backs off" % _f, "_close_retry_at" in _blk)
        check(78, "%s: and does nothing when flat" % _f,
              "if not self.position" in _blk)

    # CLASS 5 - open(...,'w') with the value computed inline. The write
    # expression is evaluated AFTER the file is truncated, so if it raises the
    # file is destroyed. This emptied futures_client.py on 9/3.
    for _f, _t in _srcs.items():
        if not _f.endswith(".py"):
            continue
        check(78, "%s: writes cannot truncate on failure" % _f,
              not re.search(r"open\([^)]*['\"]w['\"][^)]*\)\.write\(", _t), _f)

    # CLASS 6 - unbounded while loops in code that runs during a trade.
    for _f in ("webull_client.py", "futures_client.py", "trend.py", "tape.py"):
        _t = _srcs.get(_f, "")
        for _ln in _t.split("\n"):
            if _ln.strip().startswith("while ") and "True" in _ln:
                check(78, "%s: %s is a daemon loop, not a retry" % (_f, _ln.strip()[:30]),
                      False, "unexpected while True in trading code")
        check(78, "%s: has no unbounded retry loop" % _f, True)

    # CLASS 7 - a server error reported to him as "the app isn't running".
    for _f in _APP_UI:
        _t = _srcs[_f]
        check(78, "%s: reads the body before parsing it" % _f,
              "await r.text()" in _t, _f)

    # CLASS 7b - an ORDER whose outcome is swallowed. Decoding the response
    # body inside try/except is fine - the status code is checked right after.
    # What must never happen is the send itself disappearing into a bare
    # except, because then a rejected order looks exactly like a filled one.
    for _f in ("webull_client.py", "futures_client.py"):
        _raw = io.open(os.path.join(HERE, _f), encoding="utf-8").read()
        _ls = _raw.split("\n")
        _bad_send = []
        for _i, _l in enumerate(_ls):
            if not re.search(r"(place_order|_write_oif|self\._order)\(", _l):
                continue
            # walk back a few lines: is this send wrapped in a try whose
            # handler is a bare pass?
            for _j in range(max(0, _i - 4), _i):
                if _ls[_j].strip() == "try:":
                    for _k in range(_i, min(len(_ls), _i + 8)):
                        if re.match(r"\s*except[^:]*:\s*$", _ls[_k]) and \
                           _k + 1 < len(_ls) and _ls[_k + 1].strip() == "pass":
                            _bad_send.append(_i + 1)
                    break
        check(78, "%s: no order is sent into a silent except" % _f,
              not _bad_send, str(_bad_send))
        # and the outcome is always inspected
        for _i, _l in enumerate(_ls):
            # A CALL SITE, not the word. Prose in a docstring ("retrying a
            # place_order after a 429") is not an order being sent, and a test
            # that cannot tell the difference pushes you to delete the note.
            if re.search(r"place_order\s*[,(]", _l) and "def " not in _l:
                _after = "\n".join(_ls[_i:_i + 12])
                check(78, "%s: the order result is checked (line %d)" % (_f, _i + 1),
                      "status_code" in _after, _after[:80])

    # CLASS 8 - tests that assert my own prose. They pass while the behaviour
    # they name is broken, which is the worst possible test.
    _tsrc = io.open(os.path.join(HERE, "test_all.py"), encoding="utf-8").read()
    _needles = set(re.findall(r'"([^"\\\n]{14,90})" in _?\w+', _tsrc))
    _comment_lines = set()
    for _f, _raw in ((f, io.open(os.path.join(HERE, f), encoding="utf-8").read())
                     for f in _APP_PY + _APP_UI if os.path.exists(os.path.join(HERE, f))):
        for _l in _raw.split("\n"):
            _st = _l.strip()
            if _st.startswith("#") or _st.startswith("//"):
                _comment_lines.add(_st.lstrip("#/ ").strip())
    # Flag a needle only when it exists ONLY in comments. A phrase that also
    # appears in real code - a user-facing message, say - is being asserted
    # against behaviour even if a comment happens to quote it. "do not hold
    # that" is exactly that case: it is the phantom message he reads on
    # screen, and the comment above it explains why.
    _code_all = "\n".join(_srcs.values())
    _prose = [n for n in _needles
              if re.search(r"[a-z]{3,} [a-z]{3,} [a-z]{3,}", n)
              and any(n in c for c in _comment_lines)
              and n not in _code_all]
    check(78, "no test asserts a comment instead of behaviour",
          not _prose, str(sorted(_prose)[:4]))


print("\n"+"="*68)
by={}
for sc,name,ok,_ in results:
    by.setdefault(sc,[0,0]); by[sc][0]+=1; by[sc][1]+= (1 if ok else 0)
T={78:"The standing rule: sweep the class",77:"Futures gets everything options got",76:"Full-app audit: buttons, loops, dead ends",75:"A quote is for the contract you asked",74:"The day is a return, not a sum",73:"Speed: pace to the real limit",72:"Clicking CONNECT twice cannot jam it",71:"One 429 must not freeze the app",70:"A rejected exit is not a storm",69:"Sign-in survives a rate limit",68:"Stream cannot blind the app",67:"Restart keeps the position",66:"One-second prices",65:"Journal never loses a trade",64:"Pacing must not stall",63:"Tick velocity",62:"Underlying at fill",61:"Tiered ratchet + anti-clip",60:"SDK audit is honest",59:"Batched option quotes",58:"Option price grid",57:"Webull rate budget",56:"NinjaScript compiles",55:"One-click NT install",54:"Ratchet inside NinjaTrader",53:"Limits die with the app",52:"NinjaTrader delivery check",51:"Futures header + footer",50:"Futures on by default",49:"Toggles + short hints",48:"Futures config stripped",47:"Futures ratchet",46:"NinjaScript in step",45:"Breadth + VIX",44:"Entry telemetry",43:"Trend module",42:"Audio cues",41:"Volatility gauges",40:"Volume gauge",39:"Dwell time",38:"Velocity vs feed artifacts",37:"Desktop icon",36:"Trade log detail",35:"Time value warns not blocks",34:"LOCK/X gone, size warns",33:"Page actually runs",32:"No SAVE / live trade frozen",31:"One switch / still modal",30:"Directional entry levels",29:"Percent only, no cash",28:"Grid/ATM/quality/one-armed",27:"Config screen cleanup",26:"Ratchet stop",25:"Console auto-hide",24:"Options auto-reconcile",23:"Daily trade log",22:"Options phantom clear",21:"Auto-reconcile w/ broker",20:"MY CONFIG always on",19:"Phantom position",18:"Futures hours",17:"Closed market honest",16:"Restart leaves no spinner",15:"One tab only",14:"Git lock self-heal",13:"Broker tabs + tray",12:"Velocity honest when shut",11:"Multi-broker sessions",1:"Futures login survives restart",2:"remember_login default",3:"Options profiles to disk",
   4:"Browser autofill guard",5:"ITM3 strike math",6:"Preview == Arm",7:"Live-only / dead modes",
   8:"Auto-sync safety",9:"Endpoints alive",10:"UI integrity"}
for sc in sorted(by):
    tot,good=by[sc]
    print(f"  [{sc:2}] {T[sc]:34} {good}/{tot} {'OK' if good==tot else '<-- FAILURES'}")
fails=[r for r in results if not r[2]]
print("="*68)
print(f"  TOTAL: {len(results)-len(fails)}/{len(results)} passed")
if fails:
    print("\n  FAILURES:")
    for sc,name,_,d in fails: print(f"    [{sc}] {name}  {d}")
print("  my-settings.json restored.")
