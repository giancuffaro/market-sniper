"""
TRADE LOG — every entry and exit, appended to ONE Excel workbook forever.

    logs/Market Sniper Trade Log.xlsx      <- the file you open
    logs/trades.csv                        <- the real record

WHY TWO FILES
    The CSV is the source of truth and the Excel file is rebuilt from it.
    That is deliberate: on Windows, having the workbook open in Excel LOCKS it,
    and a write fails. If the xlsx were the only copy, closing a trade while the
    sheet was open would silently lose that trade forever. The CSV append is
    tiny and happens first, so the trade is safe before Excel is ever touched.
    If the rebuild fails because the file is open, nothing is lost - the next
    trade (or REBUILD) picks it up.

WHAT YOU GET
    TRADES   one row per closed trade, newest last, with a Date column
    BY DAY   one row per trading day: trades, wins, losses, win rate, net P&L

    Same workbook every day. Nothing is ever overwritten or rotated.

openpyxl is optional. Without it the CSV still records everything and you can
open that in Excel directly.
"""

import os
import csv
import datetime as dt

HERE = os.path.dirname(os.path.abspath(__file__))
LOG_DIR = os.path.join(HERE, "logs")
CSV_PATH = os.path.join(LOG_DIR, "trades.csv")
XLSX_PATH = os.path.join(LOG_DIR, "Market Sniper Trade Log.xlsx")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except Exception:
    XLSX_AVAILABLE = False

FIELDS = [
    "date", "time_in", "time_out", "held_secs", "app", "broker", "account",
    "symbol", "side", "strike", "strike_mode", "expiry", "qty",
    "entry", "exit", "pnl", "pnl_pct",
    # How the trade actually travelled, not just where it started and stopped.
    # best/worst are the high and low water marks while it was open; give_back
    # is how much of the best was handed back by the time it closed. A ratchet
    # that is working shows a small gave_back_pct. A big one means the stop was
    # too far behind the move.
    "best_pct", "worst_pct", "best_price", "worst_price", "gave_back_pct",
    "ratchet_stop_pct", "ratchet_step",
    "exit_reason", "note",
]

# Columns that should be numbers in Excel, not text.
_NUM = {"strike", "qty", "entry", "exit", "pnl", "pnl_pct", "held_secs",
        "best_pct", "worst_pct", "best_price", "worst_price", "gave_back_pct",
        "ratchet_stop_pct", "ratchet_step"}


def _rows():
    if not os.path.exists(CSV_PATH):
        return []
    try:
        with open(CSV_PATH, "r", encoding="utf-8", newline="") as f:
            return [r for r in csv.DictReader(f) if r.get("date")]
    except Exception:
        return []


def _migrate_header():
    """Rewrite the CSV if its header is older than FIELDS.

    Appending a row in FIELDS order to a file whose header has FEWER columns
    silently shifts every value out of its column - the file still opens, it is
    just wrong, and you would not notice until you tried to read your own
    numbers back. So the header is checked before every append, and old rows
    are rewritten with blanks in the new columns.
    """
    if not os.path.exists(CSV_PATH):
        return
    try:
        with open(CSV_PATH, "r", encoding="utf-8", newline="") as f:
            rdr = csv.reader(f)
            head = next(rdr, None)
        if head is None or head == FIELDS:
            return
        with open(CSV_PATH, "r", encoding="utf-8", newline="") as f:
            old = list(csv.DictReader(f))
        with open(CSV_PATH, "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=FIELDS)
            w.writeheader()
            for r in old:
                w.writerow({k: r.get(k, "") for k in FIELDS})
        print("[trade_log] header upgraded: %d -> %d columns, %d rows kept"
              % (len(head), len(FIELDS), len(old)), flush=True)
    except Exception as e:                                   # noqa: BLE001
        print("[trade_log] header migration skipped: %s" % str(e)[:120], flush=True)


def record(trade):
    """Append ONE closed trade. Never raises - logging must not break trading."""
    try:
        os.makedirs(LOG_DIR, exist_ok=True)
        _migrate_header()
        row = {k: trade.get(k, "") for k in FIELDS}
        row["date"] = row["date"] or dt.date.today().isoformat()
        new_file = not os.path.exists(CSV_PATH)
        with open(CSV_PATH, "a", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=FIELDS)
            if new_file:
                w.writeheader()
            w.writerow(row)
    except Exception as e:                                   # noqa: BLE001
        print("[trade_log] could not write CSV: %s" % str(e)[:120], flush=True)
        return False

    # Excel is a convenience on top. If it is open in Excel this fails, and that
    # is fine - the CSV already has the trade and the next rebuild catches up.
    try:
        rebuild_xlsx()
    except PermissionError:
        print("[trade_log] the workbook is open in Excel - trade saved to CSV, "
              "the sheet updates once you close it.", flush=True)
    except Exception as e:                                   # noqa: BLE001
        print("[trade_log] xlsx rebuild skipped: %s" % str(e)[:120], flush=True)
    return True


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def by_day(rows=None):
    """Per-day totals, oldest first."""
    rows = _rows() if rows is None else rows
    days = {}
    for r in rows:
        d = days.setdefault(r["date"], {"date": r["date"], "trades": 0, "wins": 0,
                                        "losses": 0, "net": 0.0})
        d["trades"] += 1
        pnl = _num(r.get("pnl"))
        pnl = pnl if isinstance(pnl, float) else 0.0
        d["net"] += pnl
        if pnl > 0:
            d["wins"] += 1
        elif pnl < 0:
            d["losses"] += 1
    out = []
    for d in sorted(days.values(), key=lambda x: x["date"]):
        d["net"] = round(d["net"], 2)
        d["win_rate"] = round(d["wins"] / d["trades"] * 100, 1) if d["trades"] else 0.0
        out.append(d)
    return out


def rebuild_xlsx():
    """Regenerate the workbook from the CSV. Same file, every time."""
    if not XLSX_AVAILABLE:
        return False
    rows = _rows()
    if not rows:
        return False
    os.makedirs(LOG_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3A2E")
    green = Font(color="1B7F4B")
    red = Font(color="B00020")

    ws = wb.active
    ws.title = "TRADES"
    ws.append([f.replace("_", " ").upper() for f in FIELDS])
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([_num(r.get(f)) if f in _NUM else r.get(f, "") for f in FIELDS])
    pnl_col = FIELDS.index("pnl") + 1
    for i in range(2, ws.max_row + 1):
        v = ws.cell(row=i, column=pnl_col).value
        if isinstance(v, (int, float)):
            ws.cell(row=i, column=pnl_col).font = green if v >= 0 else red
            ws.cell(row=i, column=pnl_col).number_format = '"$"#,##0.00'
    ws.freeze_panes = "A2"

    ds = wb.create_sheet("BY DAY")
    ds.append(["DATE", "TRADES", "WINS", "LOSSES", "WIN RATE %", "NET P&L"])
    for c in ds[1]:
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")
    for d in by_day(rows):
        ds.append([d["date"], d["trades"], d["wins"], d["losses"], d["win_rate"], d["net"]])
    for i in range(2, ds.max_row + 1):
        cell = ds.cell(row=i, column=6)
        cell.number_format = '"$"#,##0.00'
        if isinstance(cell.value, (int, float)):
            cell.font = green if cell.value >= 0 else red
    ds.freeze_panes = "A2"

    for sheet in (ws, ds):
        for col in range(1, sheet.max_column + 1):
            width = max((len(str(sheet.cell(row=r, column=col).value or ""))
                         for r in range(1, min(sheet.max_row, 400) + 1)), default=8)
            sheet.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 9), 30)

    # Write to a temp file first, then swap. A crash mid-write can never leave
    # you with a half-written workbook where your trade history used to be.
    tmp = XLSX_PATH + ".tmp"
    wb.save(tmp)
    os.replace(tmp, XLSX_PATH)
    return True


def summary(days=1):
    """Quick totals for the app to display."""
    d = by_day()
    return d[-days:] if d else []


if __name__ == "__main__":
    n = len(_rows())
    print("%d trade(s) in %s" % (n, CSV_PATH))
    if rebuild_xlsx():
        print("rebuilt %s" % XLSX_PATH)
    else:
        print("nothing to rebuild" if n else "no trades yet")
    for d in by_day()[-10:]:
        print("  %s  %2d trades  %2dW/%2dL  %5.1f%%  $%.2f"
              % (d["date"], d["trades"], d["wins"], d["losses"], d["win_rate"], d["net"]))
